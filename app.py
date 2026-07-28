# -*- coding: utf-8 -*-
"""
와플랫 공공 지표 - 지자체 운영 대시보드
Google Sheets 실데이터 기반 인터랙티브 Streamlit 대시보드
"""
import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import plotly.io as pio
from plotly.subplots import make_subplots

# Plotly 전역 폰트: Pretendard (한국어 가독성 최적화)
pio.templates["waplat"] = go.layout.Template(
    layout=go.Layout(
        font=dict(family="Pretendard, Noto Sans KR, sans-serif", size=13, color="#1E293B"),
        title=dict(font=dict(size=15, color="#1E293B", family="Pretendard, Noto Sans KR")),
        paper_bgcolor="white",
        plot_bgcolor="white",
        xaxis=dict(gridcolor="#F1F5F9", linecolor="#E2E8F0"),
        yaxis=dict(gridcolor="#F1F5F9", linecolor="#E2E8F0"),
    )
)
pio.templates.default = "waplat"

from sheets_data import (
    fetch_all_sheets, fetch_sheet, SHEET_GIDS,
    build_dashboard_data, build_municipality_heatmap_data,
    get_week_summary, get_weekly_municipality_data,
    safe_numeric, find_municipality_columns, extract_municipality_name,
    REGION_MAP, MUNICIPALITY_KEYWORDS, ALL_KNOWN_AGENCIES,
)
from unified_data import (
    load_unified_data, get_agency_master, save_agency,
    toggle_agency_active, get_agency_summary, get_data_source_info,
    get_db_data, seed_agencies_from_sheets, get_active_agencies,
    import_safety_check_from_sheets,
)
from data_input import DATA_TYPES, process_pasted_data, detect_data_type
from local_db import (
    save_safety_check, save_generic, get_safety_check_data,
    get_all_dates, get_data_stats, init_db,
    save_agency, deactivate_agency, activate_agency, delete_agency,
    get_all_agencies, get_agency_summary,
    get_note, save_note,
)
from sheets_data import (
    get_sheet_note, save_sheet_note,
    get_safe_status_from_sheet, save_safe_status_to_sheet,
)

# ============================================================
# 페이지 설정
# ============================================================
st.set_page_config(
    page_title="와플랫 공공 지표 대시보드",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ============================================================
# 🔐 비밀번호 보호
# ============================================================
def _check_password() -> bool:
    """비밀번호 확인 — 통과 시 True 반환
    REQUIRE_AUTH=true 일 때만 비밀번호 요구 (로컬에서는 자동 통과)
    """
    # 로컬 실행 시 비밀번호 스킵 (secrets에 REQUIRE_AUTH=true 없으면 통과)
    if str(st.secrets.get("REQUIRE_AUTH", "false")).lower() != "true":
        return True

    if st.session_state.get("authenticated"):
        return True

    # 중앙 정렬된 로그인 카드
    st.markdown("""
    <style>
    #login-wrap {
        max-width: 400px; margin: 80px auto; padding: 2.5rem 2rem;
        background: white; border-radius: 20px;
        box-shadow: 0 8px 32px rgba(102,126,234,0.18);
        text-align: center;
    }
    #login-wrap h2 { color: #667eea; margin-bottom: 0.2rem; }
    #login-wrap p  { color: #888; font-size: 0.9rem; margin-bottom: 1.5rem; }
    </style>
    <div id="login-wrap">
      <h2>📊 와플랫 공공 대시보드</h2>
      <p>접근 권한이 필요합니다</p>
    </div>
    """, unsafe_allow_html=True)

    col = st.columns([1, 2, 1])[1]
    with col:
        pw = st.text_input("비밀번호", type="password", label_visibility="collapsed",
                           placeholder="비밀번호를 입력하세요")
        if st.button("로그인", use_container_width=True, type="primary"):
            correct = st.secrets.get("PASSWORD", "waplat2025!")
            if pw == correct:
                st.session_state["authenticated"] = True
                st.rerun()
            else:
                st.error("비밀번호가 틀렸습니다.")
    return False

if not _check_password():
    st.stop()

# ============================================================
# 커스텀 CSS
# ============================================================
st.markdown("""
<style>
@import url('https://cdn.jsdelivr.net/gh/orioncactus/pretendard@v1.3.9/dist/web/static/pretendard.min.css');
@import url('https://fonts.googleapis.com/css2?family=Noto+Sans+KR:wght@300;400;500;600;700;900&display=swap');

/* ══════════════════════════════════════
   기본 폰트 & 배경
══════════════════════════════════════ */
html, body, [class*="css"] {
    font-family: 'Pretendard', 'Noto Sans KR', -apple-system, BlinkMacSystemFont, sans-serif;
    -webkit-font-smoothing: antialiased;
    letter-spacing: -0.01em;
}

/* 커스텀 스크롤바 */
::-webkit-scrollbar { width: 5px; height: 5px; }
::-webkit-scrollbar-track { background: transparent; }
::-webkit-scrollbar-thumb { background: #CBD5E1; border-radius: 99px; }
::-webkit-scrollbar-thumb:hover { background: #94A3B8; }

/* 배경 — 부드러운 블루그레이 */
.stApp {
    background: linear-gradient(160deg, #F0F4FF 0%, #F7F8FC 50%, #EFF3FB 100%);
    background-attachment: fixed;
}

/* 메인 콘텐츠 여백 */
section.main > div { padding-top: 1.2rem; }
.block-container { padding-left: 2rem !important; padding-right: 2rem !important; }

/* ══════════════════════════════════════
   KPI 카드 — 글래스모피즘 + 그라디언트
══════════════════════════════════════ */
.metric-card, .metric-card-green, .metric-card-red, .metric-card-orange {
    position: relative;
    padding: 1.5rem 1.3rem 1.2rem;
    border-radius: 22px;
    color: white;
    text-align: center;
    margin-bottom: 0.6rem;
    overflow: hidden;
    transition: transform 0.22s cubic-bezier(.34,1.56,.64,1), box-shadow 0.22s ease;
    border: 1px solid rgba(255,255,255,0.25);
}
.metric-card:hover,
.metric-card-green:hover,
.metric-card-red:hover,
.metric-card-orange:hover {
    transform: translateY(-5px) scale(1.01);
    box-shadow: 0 20px 48px rgba(0,0,0,0.20) !important;
}

/* 카드 빛번짐 효과 — 상단 */
.metric-card::before, .metric-card-green::before,
.metric-card-red::before, .metric-card-orange::before {
    content: '';
    position: absolute;
    top: -40%; right: -15%;
    width: 180px; height: 180px;
    background: rgba(255,255,255,0.13);
    border-radius: 50%;
    pointer-events: none;
}
/* 카드 빛번짐 효과 — 하단 */
.metric-card::after, .metric-card-green::after,
.metric-card-red::after, .metric-card-orange::after {
    content: '';
    position: absolute;
    bottom: -40%; left: -10%;
    width: 120px; height: 120px;
    background: rgba(0,0,0,0.07);
    border-radius: 50%;
    pointer-events: none;
}

/* 라벨 */
.metric-card h3, .metric-card-green h3,
.metric-card-red h3, .metric-card-orange h3 {
    margin: 0 0 0.45rem;
    font-size: 0.68rem;
    font-weight: 600;
    letter-spacing: 0.10em;
    text-transform: uppercase;
    opacity: 0.78;
}

/* 숫자 값 */
.metric-card h1, .metric-card-green h1,
.metric-card-red h1, .metric-card-orange h1 {
    margin: 0 0 0.45rem;
    font-size: 2.7rem;
    font-weight: 900;
    letter-spacing: -0.03em;
    line-height: 1;
    text-shadow: 0 2px 8px rgba(0,0,0,0.15);
}

/* 델타 */
.metric-card p, .metric-card-green p,
.metric-card-red p, .metric-card-orange p {
    margin: 0;
    font-size: 0.76rem;
    opacity: 0.88;
}
.metric-card .up, .metric-card-green .up,
.metric-card-red .up, .metric-card-orange .up   { color: #A7F3D0; font-weight: 700; }
.metric-card .down, .metric-card-green .down,
.metric-card-red .down, .metric-card-orange .down { color: #FCA5A5; font-weight: 700; }
.metric-card .flat, .metric-card-green .flat,
.metric-card-red .flat, .metric-card-orange .flat { color: rgba(255,255,255,0.55); }

/* 카드별 색상 */
.metric-card {
    background: linear-gradient(140deg, #6366F1 0%, #7C3AED 100%);
    box-shadow: 0 8px 24px rgba(99,102,241,0.40);
}
.metric-card-green {
    background: linear-gradient(140deg, #10B981 0%, #047857 100%);
    box-shadow: 0 8px 24px rgba(16,185,129,0.38);
}
.metric-card-red {
    background: linear-gradient(140deg, #F87171 0%, #B91C1C 100%);
    box-shadow: 0 8px 24px rgba(248,113,113,0.38);
}
.metric-card-orange {
    background: linear-gradient(140deg, #F97316 0%, #C2410C 100%);
    box-shadow: 0 8px 24px rgba(249,115,22,0.38);
}

/* ══════════════════════════════════════
   섹션 헤더 — 그라디언트 라인 + 깊이감
══════════════════════════════════════ */
.section-header {
    display: flex;
    align-items: center;
    gap: 0.7rem;
    padding: 0.75rem 1.3rem;
    margin: 2rem 0 1.2rem;
    background: white;
    border-radius: 14px;
    border-left: 5px solid transparent;
    background-clip: padding-box;
    background-image: linear-gradient(white, white),
                      linear-gradient(135deg, #6366F1, #10B981);
    background-origin: border-box;
    font-weight: 700;
    font-size: 1rem;
    color: #0F172A;
    box-shadow: 0 4px 16px rgba(0,0,0,0.07), 0 1px 3px rgba(0,0,0,0.05);
    letter-spacing: -0.01em;
}

/* ══════════════════════════════════════
   인사이트 박스 — 부드럽고 세련되게
══════════════════════════════════════ */
.insight-box {
    background: linear-gradient(135deg, #EEF2FF 0%, #E0E7FF 100%);
    border-left: 4px solid #6366F1;
    padding: 1rem 1.3rem;
    margin: 0.5rem 0;
    border-radius: 0 14px 14px 0;
    font-size: 0.875rem;
    line-height: 1.7;
    color: #1E293B;
    box-shadow: 0 2px 8px rgba(99,102,241,0.08);
}
.insight-box-danger {
    background: linear-gradient(135deg, #FFF1F2 0%, #FFE4E6 100%);
    border-left: 4px solid #F43F5E;
    padding: 1rem 1.3rem;
    margin: 0.5rem 0;
    border-radius: 0 14px 14px 0;
    font-size: 0.875rem;
    line-height: 1.7;
    color: #1E293B;
    box-shadow: 0 2px 8px rgba(244,63,94,0.08);
}
.insight-box-success {
    background: linear-gradient(135deg, #F0FDF4 0%, #DCFCE7 100%);
    border-left: 4px solid #10B981;
    padding: 1rem 1.3rem;
    margin: 0.5rem 0;
    border-radius: 0 14px 14px 0;
    font-size: 0.875rem;
    line-height: 1.7;
    color: #1E293B;
    box-shadow: 0 2px 8px rgba(16,185,129,0.08);
}

/* ══════════════════════════════════════
   상태 뱃지 — 약간 더 선명하게
══════════════════════════════════════ */
.status-danger   { background:#FEE2E2; color:#9F1239; padding:4px 12px; border-radius:999px; font-weight:700; font-size:0.76rem; display:inline-block; box-shadow:0 1px 4px rgba(159,18,57,0.15); }
.status-caution  { background:#FEF3C7; color:#92400E; padding:4px 12px; border-radius:999px; font-weight:700; font-size:0.76rem; display:inline-block; box-shadow:0 1px 4px rgba(146,64,14,0.12); }
.status-normal   { background:#F1F5F9; color:#475569; padding:4px 12px; border-radius:999px; font-size:0.76rem; display:inline-block; box-shadow:0 1px 4px rgba(0,0,0,0.06); }
.status-excellent{ background:#DCFCE7; color:#14532D; padding:4px 12px; border-radius:999px; font-weight:700; font-size:0.76rem; display:inline-block; box-shadow:0 1px 4px rgba(20,83,45,0.15); }

/* ══════════════════════════════════════
   탭 스타일 — 더 부드럽고 선명하게
══════════════════════════════════════ */
.stTabs [data-baseweb="tab-list"] {
    gap: 3px;
    background: rgba(226,232,240,0.7);
    border-radius: 14px;
    padding: 5px;
    backdrop-filter: blur(4px);
}
.stTabs [data-baseweb="tab"] {
    padding: 8px 20px;
    border-radius: 10px;
    font-weight: 500;
    font-size: 0.84rem;
    color: #64748B;
    background: transparent;
    border: none;
    transition: all 0.15s ease;
}
.stTabs [data-baseweb="tab"]:hover {
    color: #6366F1 !important;
    background: rgba(255,255,255,0.6) !important;
}
.stTabs [aria-selected="true"] {
    background: white !important;
    color: #6366F1 !important;
    font-weight: 700 !important;
    box-shadow: 0 2px 12px rgba(99,102,241,0.15), 0 1px 3px rgba(0,0,0,0.06) !important;
}

/* ══════════════════════════════════════
   사이드바 — 다크 + 세련된 메뉴
══════════════════════════════════════ */
section[data-testid="stSidebar"] {
    background: linear-gradient(180deg, #0F172A 0%, #1E293B 100%);
    border-right: 1px solid rgba(255,255,255,0.04);
}
section[data-testid="stSidebar"] * {
    color: #94A3B8 !important;
}
/* 라디오 원 숨기기 */
section[data-testid="stSidebar"] .stRadio [data-baseweb="radio"] > div:first-child {
    display: none !important;
}
section[data-testid="stSidebar"] .stRadio label {
    font-size: 0.85rem !important;
    padding: 8px 12px !important;
    border-radius: 9px;
    transition: all 0.15s ease;
    cursor: pointer;
    display: flex !important;
    align-items: center !important;
    gap: 6px;
    margin: 2px 0;
    color: #94A3B8 !important;
}
section[data-testid="stSidebar"] .stRadio label:hover {
    background: rgba(99,102,241,0.14) !important;
    color: #C7D2FE !important;
}
section[data-testid="stSidebar"] .stRadio [aria-checked="true"] ~ div label,
section[data-testid="stSidebar"] .stRadio label[data-checked="true"] {
    background: rgba(99,102,241,0.18) !important;
    color: #A5B4FC !important;
    font-weight: 600 !important;
}
section[data-testid="stSidebar"] [data-testid="stMarkdownContainer"] p {
    color: #475569 !important;
    font-size: 0.76rem !important;
}
section[data-testid="stSidebar"] h1,
section[data-testid="stSidebar"] h2,
section[data-testid="stSidebar"] h3 {
    color: #E2E8F0 !important;
}
section[data-testid="stSidebar"] .stDivider {
    border-color: rgba(255,255,255,0.06) !important;
}

/* ── 버튼 스타일 ── */
.stButton > button {
    border-radius: 10px !important;
    font-weight: 600 !important;
    font-size: 0.85rem !important;
    transition: all 0.18s ease !important;
    border: none !important;
}
.stButton > button:hover {
    transform: translateY(-1px) !important;
    box-shadow: 0 6px 16px rgba(0,0,0,0.12) !important;
}

/* ── 차트 컨테이너 카드화 ── */
div[data-testid="stPlotlyChart"] {
    background: white;
    border-radius: 18px;
    padding: 0.5rem;
    box-shadow: 0 4px 20px rgba(0,0,0,0.06), 0 1px 4px rgba(0,0,0,0.04);
    margin-bottom: 0.5rem;
}

/* ── expander 스타일 ── */
.stExpander {
    border: 1px solid rgba(0,0,0,0.06) !important;
    border-radius: 14px !important;
    box-shadow: 0 2px 8px rgba(0,0,0,0.04) !important;
    overflow: hidden;
}

/* ── 데이터프레임 ── */
div[data-testid="stMetricValue"] { font-size: 1.4rem; font-weight: 700; }

/* ── info / warning / error 박스 ── */
div[data-testid="stAlert"] {
    border-radius: 12px !important;
    border: none !important;
    box-shadow: 0 2px 8px rgba(0,0,0,0.06) !important;
}

/* ── 선택박스 / 슬라이더 ── */
div[data-baseweb="select"] > div {
    border-radius: 10px !important;
    border-color: #E2E8F0 !important;
    box-shadow: 0 1px 4px rgba(0,0,0,0.05) !important;
}
div[data-baseweb="select"] > div:hover {
    border-color: #6366F1 !important;
}
</style>
""", unsafe_allow_html=True)

# ============================================================
# 데이터 로드 — DB 우선, Google Sheets는 새로고침 시에만
# ============================================================
@st.cache_data(ttl=14400, show_spinner="Google Sheets 데이터 로딩 중... (최초 1회, 이후 4시간 캐시)")
def load_all_data():
    """전체 Google Sheets 로드 (4시간 캐시)"""
    sheets = fetch_all_sheets()
    data = build_dashboard_data(sheets)
    return sheets, data


@st.cache_data(ttl=14400, show_spinner=False)
def cached_heatmap(_data: dict, week: str) -> "pd.DataFrame":
    """지자체 히트맵 — 주차별 캐시 (페이지 재진입 시 즉시 반환)"""
    return build_municipality_heatmap_data(_data, week)


def cached_week_summary(sheets: dict, data: dict, week: str) -> dict:
    return get_week_summary(sheets, data, week)

try:
    sheets, data = load_all_data()
    DATA_LOADED = True

    _G_CR: dict = {}  # 미사용 (KPI 카드는 get_week_summary의 안부체크율 사용)
    # 최초 1회: Google Sheets에서 지자체 자동 등록 + 안부확인 raw 데이터 임포트
    if "agency_seeded" not in st.session_state:
        seeded = seed_agencies_from_sheets(sheets)
        if seeded > 0:
            st.toast(f"Google Sheets에서 {seeded}개 지자체 자동 등록됨")
        # 안부확인 raw 데이터 임포트
        imported = import_safety_check_from_sheets(sheets)
        if imported > 0:
            st.toast(f"안부확인 raw 데이터 {imported}건 임포트됨")
        st.session_state["agency_seeded"] = True
except Exception as e:
    st.error(f"데이터 로딩 실패: {e}")
    DATA_LOADED = False
    sheets, data = {}, {}

# 시트 H열(사업구분)이 있으면 BUSINESS_TYPE_MAP을 동적으로 덮어씀
# → 계약 변경 시 시트만 수정하면 자동 반영
_reg_biz = data.get("registration", pd.DataFrame())
if not _reg_biz.empty and "사업구분" in _reg_biz.columns:
    BUSINESS_TYPE_MAP = {
        str(row["지자체명"]): str(row["사업구분"])
        for _, row in _reg_biz.iterrows()
        if str(row.get("사업구분", "")).strip() not in ("", "nan")
    }

# 협약인원 > 0인 기관만 기준으로 실제 활성 사업구분 계산
# (이 집합만 biz_selector 라디오 버튼에 표시)
def _compute_active_biz_types():
    reg = data.get("registration", pd.DataFrame())
    if reg.empty:
        return set(BUSINESS_TYPE_MAP.values())
    active = set()
    for _, row in reg.iterrows():
        try:
            if float(row.get("협약인원", 0) or 0) > 0:
                nm = str(row.get("지자체명", "")).strip()
                biz = BUSINESS_TYPE_MAP.get(nm)
                if not biz:
                    nm_n = nm.replace(" ", "")
                    for k, v in BUSINESS_TYPE_MAP.items():
                        if k.replace(" ", "") == nm_n:
                            biz = v
                            break
                if biz:
                    active.add(biz)
        except Exception:
            pass
    return active if active else set(BUSINESS_TYPE_MAP.values())

# ============================================================
# 상태 색상 / 배지
# ============================================================
STATUS_COLORS = {"집중관리": "#FF4B4B", "주의관리": "#FFA500", "정상": "#9E9E9E", "우수사례": "#00C853"}
DETAIL_STATUS_COLORS = {"위험": "#FF4B4B", "주의": "#FFA500", "보통": "#9E9E9E", "우수": "#00C853"}

def status_badge(status):
    cls = {"집중관리": "danger", "주의관리": "caution", "정상": "normal", "우수사례": "excellent"}
    return f'<span class="status-{cls.get(status, "normal")}">{status}</span>'

def delta_html(val, suffix="", invert=False, prev_val=0):
    """증감 표시 HTML (invert=True면 음수가 좋은 것)
    val: 현재값과 이전값의 차이 (delta)
    prev_val: 이전 값 (비율 계산용)
    """
    if val > 0:
        cls = "down" if invert else "up"
        arrow = "▲"
    elif val < 0:
        cls = "up" if invert else "down"
        arrow = "▼"
    else:
        cls, arrow = "flat", "→"

    # 비율 계산
    if prev_val and prev_val != 0:
        pct = abs(val) / abs(prev_val) * 100
        if suffix == "명":
            return f'<span class="{cls}">{arrow} {abs(val):,.0f}{suffix} ({pct:.1f}%)</span>'
        else:
            return f'<span class="{cls}">{arrow} {abs(val):.1f}{suffix} ({pct:.1f}%)</span>'
    else:
        if suffix == "명":
            return f'<span class="{cls}">{arrow} {abs(val):,.0f}{suffix}</span>'
        else:
            return f'<span class="{cls}">{arrow} {abs(val):.1f}{suffix}</span>'

# ============================================================
# 사이드바
# ============================================================
with st.sidebar:
    # ── 로고 ──────────────────────────────────────────────────
    import os, base64
    _logo_path = os.path.join(os.path.dirname(__file__), "assets", "logo.png")
    if os.path.exists(_logo_path):
        with open(_logo_path, "rb") as _f:
            _logo_b64 = base64.b64encode(_f.read()).decode()
        st.markdown(
            f"""<div style="background:white;border-radius:12px;
                            padding:10px 16px 8px;margin:4px 0 12px;
                            text-align:center;
                            box-shadow:0 2px 8px rgba(0,0,0,0.25)">
                  <img src="data:image/png;base64,{_logo_b64}"
                       style="max-width:150px;height:auto;display:block;margin:0 auto">
                  <div style="font-size:0.65rem;color:#64748B;margin-top:4px;
                              font-family:'Pretendard','Noto Sans KR',sans-serif">
                    공공 서비스 지표 대시보드
                  </div>
                </div>""",
            unsafe_allow_html=True,
        )
    else:
        st.markdown(
            """<div style="padding:0.6rem 0 0.4rem;text-align:left">
              <span style="font-size:1.6rem;font-weight:900;color:white;
                           font-family:'Pretendard','Noto Sans KR',sans-serif;
                           letter-spacing:-0.03em">waplat</span>
              <span style="font-size:0.7rem;color:rgba(255,255,255,0.55);
                           display:block;margin-top:2px">공공 서비스 지표 대시보드</span>
            </div>""",
            unsafe_allow_html=True,
        )

    weeks = data.get("주차목록", [])
    if weeks:
        selected_week = st.selectbox(
            "📅 주차 선택",
            options=list(reversed(weeks)),
            index=0,
            help="데이터를 확인할 주차를 선택하세요"
        )
    else:
        selected_week = None
        st.warning("주차 데이터 없음")

    st.divider()

    if st.button("🔄 데이터 새로고침", use_container_width=True, help="캐시를 초기화하고 Google Sheets에서 최신 데이터를 불러옵니다"):
        st.cache_data.clear()
        st.rerun()

    st.divider()
    st.caption("v2026-07-20")

    # 페이지 선택
    page = st.radio(
        "페이지 선택",
        [
            "📋 Summary",
            "🧭 사업구분별 현황",
            "🛡 세이프·베이직 현황",
            "👥 1.회원가입 & 이탈",
            "🖐 2.안부확인",
            "📊 3.안부체크율",
            "🔄 4.안부체크 변경(베이직)",
            "🛡 5.안부체크 변경(세이프)",
            "❤ 6.심혈관체크",
            "😰 7.스트레스체크",
            "💊 8.복약관리",
            "🩺 9.건강상담",
            "💬 10.생활상담",
            "🃏 11.맞고(와플랫)",
            "🎮 12.맞고(와플랫+게스트)",
            "👤 13.맞고(게스트)",
            "🤖 AI 생활지원사",
            "🚶 걸음수",
        ],
        index=0,
    )

    st.divider()
    if st.button("🔄 데이터 새로고침"):
        st.cache_data.clear()
        st.rerun()
    st.caption("4시간마다 자동 새로고침")

    if DATA_LOADED:
        # 데이터 소스 상태 — 항상 최신 데이터 사용 (summary와 동기화)
        agency_stats = get_agency_summary(sheets)
        if agency_stats.get("total", 0) > 0:
            st.caption(f"🏛 지자체 {agency_stats.get('active', 0)}개 운영 중")
            st.caption(f"🛡 세이프 {agency_stats.get('safe_count', 0)} | 📋 베이직 {agency_stats.get('basic_count', 0)}")
        st.caption("✅ 데이터 연결됨")
    else:
        st.caption("⚠️ 데이터 로딩 실패")

# ============================================================
# Helper 함수
# ============================================================
def weekly_total(df, value_col="값", agg="sum"):
    """주차별 합계를 시작일 포함하여 계산"""
    if "시작일" in df.columns:
        # 주차+시작일 쌍의 첫 번째 시작일 유지
        date_map = df.groupby("주차")["시작일"].first()
        total = df.groupby("주차")[value_col].agg(agg).reset_index()
        total["시작일"] = total["주차"].map(date_map)
    else:
        total = df.groupby("주차")[value_col].agg(agg).reset_index()
    return total

def get_prev_week(week):
    """이전 주차 반환"""
    if week in weeks:
        idx = weeks.index(week)
        return weeks[idx - 1] if idx > 0 else None
    return None

def shorten_date(date_str):
    """날짜 문자열에서 '20' 접두사 제거: 2026-03-21 → 26-03-21"""
    s = str(date_str).strip()
    if s.startswith("20") and len(s) >= 10:
        return s[2:]
    return s

def shorten_dates_in_df(df, col):
    """DataFrame의 날짜 컬럼을 짧은 형식으로 변환"""
    df = df.copy()
    df[col] = df[col].apply(shorten_date)
    return df

def date_to_week_label(date_str):
    """날짜 문자열을 ISO 주차 형식으로 변환: 2026-04-05 → 26-15 (다른 주차 컬럼과 통일)"""
    from datetime import datetime
    s = str(date_str).strip()
    try:
        if len(s) >= 10 and s[4:5] == "-":
            dt = datetime.strptime(s[:10], "%Y-%m-%d")
            yr, wk, _ = dt.isocalendar()
            return f"{str(yr)[2:]}-{wk:02d}"
    except Exception:
        pass
    return s

def week_label_df(df, col):
    """DataFrame의 날짜 컬럼을 ISO 주차 레이블로 변환 (호출 전 시간순 정렬 권장)"""
    df = df.copy()
    df[col] = df[col].apply(date_to_week_label)
    return df

# 공통 범례 설정 (X축 겹침 방지)
LEGEND_BELOW = dict(orientation="h", yanchor="top", y=-0.22, xanchor="center", x=0.5, font=dict(size=9))
LEGEND_BELOW_LARGE = dict(orientation="h", yanchor="top", y=-0.28, xanchor="center", x=0.5, font=dict(size=8))

# 가입완료 20명 미만 소규모 지자체 — 범례 이름 앞에 ○ 표시
# 수동 fallback (데이터 없을 때)
_SMALL_AGENCIES = {"강원사회서비스원", "희망나래장애인복지관", "희망나래", "양양군청",
                   "다살림재가노인지원서비스센터", "영월군청"}

# 별표 표시 지자체 (신규 계약)
_STAR_AGENCIES = {"광명시청", "양평군청", "정선군청", "제주시청", "서귀포시청", "고성군청",
                  "용인시청", "광주동구청", "용인시청통합돌봄",
                  "계양구청", "연수구청", "다살림재가노인지원서비스센터", "영월군청"}

# 자동 ○ 표시용 가입인원 캐시 (데이터 로드 후 갱신)
_joined_count_cache: dict = {}

def _update_joined_cache(reg_df) -> None:
    """등록 현황 DataFrame으로 지자체별 가입인원 캐시 갱신 (20명 미만 → ○ 자동 표시)"""
    global _joined_count_cache
    if reg_df is None or reg_df.empty:
        return
    name_col  = "지자체명" if "지자체명" in reg_df.columns else None
    count_col = "가입완료" if "가입완료" in reg_df.columns else None
    if name_col and count_col:
        for _, row in reg_df.iterrows():
            n = str(row[name_col]).strip()
            c = safe_numeric(row.get(count_col, 0))
            if n:
                _joined_count_cache[n] = int(c)

def _mun_label(name: str) -> str:
    # ① 자동: 실제 가입인원 20명 미만 → ○
    joined = _joined_count_cache.get(name)
    if joined is not None and joined < 20:
        return f"○ {name}"
    # ② fallback: 수동 세트
    if name in _SMALL_AGENCIES:
        return f"○ {name}"
    # ③ 신규 지자체 → ★
    if name in _STAR_AGENCIES:
        return f"★ {name}"
    return name

# ○ 자동 표시: 데이터 로드 완료 후 가입인원 캐시 갱신
try:
    _update_joined_cache(data.get("registration", pd.DataFrame()))
except Exception:
    pass

# 권역 분류 및 색상 (전역 — 여러 페이지에서 공유)
DETAIL_REGION = {
    "서초구청": "서울권", "강북구청": "서울권", "마포구청": "서울권", "광진구청": "서울권",
    "경기도청": "경기권", "용인시청": "경기권", "용인시청통합돌봄": "경기권",
    "포천시청": "경기권", "광명시청": "경기권", "양평군청": "경기권",
    "청주시청": "충청권", "진천군청": "충청권", "음성군청": "충청권", "괴산군청": "충청권",
    "증평군청": "충청권", "충북사회서비스원": "충청권", "충남사회서비스원": "충청권",
    "금정구청": "영남권", "경남사회서비스원": "영남권",
    "강릉시청": "강원권", "강원사회서비스원": "강원권", "홍천군청": "강원권",
    "삼척시청": "강원권", "양양군청": "강원권", "정선군청": "강원권", "고성군청": "강원권",
    "광주동구청": "호남권",
    "독거노인지원종합센터": "기타", "희망나래장애인복지관": "기타",
    "제주시청": "제주권", "서귀포시청": "제주권",
    "계양구청": "인천권", "연수구청": "인천권",
    "영월군청": "강원권",
    "다살림재가노인지원서비스센터": "기타",
}
REGION_COLORS = {
    "서울권": "#2F5496", "경기권": "#00897B", "충청권": "#E65100",
    "영남권": "#6A1B9A", "강원권": "#00838F", "호남권": "#558B2F",
    "제주권": "#AD1457", "인천권": "#1A237E", "기타": "#757575",
}

# 사업구분 매핑 (지자체명 → 사업구분)
BUSINESS_TYPE_MAP = {
    # 통합돌봄 (15개 — 7월 기준 라이브)
    "진천군청": "통합돌봄", "음성군청": "통합돌봄", "증평군청": "통합돌봄",
    "충북사회서비스원": "통합돌봄", "양양군청": "통합돌봄", "양평군청": "통합돌봄",
    "정선군청": "통합돌봄", "고성군청": "통합돌봄",
    "용인시청통합돌봄": "통합돌봄", "용인시청": "통합돌봄", "광주동구청": "통합돌봄", "연수구청": "통합돌봄",
    "계양구청": "통합돌봄", "영월군청": "통합돌봄",
    "동해시청": "통합돌봄", "세종사회서비스원": "통합돌봄",
    # 노인맞춤돌봄 (6개 — 7월 기준 라이브; 독거노인종합지원센터·강북구청 계약 종료)
    "서초구청": "노인맞춤돌봄", "포천시청": "노인맞춤돌봄",
    "홍천군청": "노인맞춤돌봄", "경기도청": "노인맞춤돌봄",
    "경남사회서비스원": "노인맞춤돌봄",
    "다살림재가노인지원서비스센터": "노인맞춤돌봄",
    # 고독사예방 (6개)
    "강릉시청": "고독사예방", "금정구청": "고독사예방", "부산금정구청": "고독사예방",
    "삼척시청": "고독사예방", "광명시청": "고독사예방",
    "제주시청": "고독사예방", "서귀포시청": "고독사예방",
    # 취약지지원 (1개)
    "충남사회서비스원": "취약지지원",
    # 장애인지원 (1개)
    "희망나래": "장애인지원", "희망나래장애인복지관": "장애인지원",
    # 퇴원환자지원 (1개 — 7월 신규)
    "인천사회서비스원": "퇴원환자지원",
    # 기타 (1개)
    "강원사회서비스원": "기타",
}
BUSINESS_TYPE_ORDER = ["통합돌봄", "노인맞춤돌봄", "고독사예방", "취약지지원", "장애인지원", "퇴원환자지원", "기타"]
BUSINESS_TYPE_COLORS = {
    "통합돌봄": "#1565C0", "노인맞춤돌봄": "#2E7D32", "고독사예방": "#C62828",
    "취약지지원": "#E65100", "장애인지원": "#6A1B9A", "퇴원환자지원": "#00695C", "기타": "#757575",
}

# BUSINESS_TYPE_MAP이 최종 확정된 뒤에 계산해야 함 (그렇지 않으면 상단의 시트 기반
# 임시 매핑 — 사업구분 표기가 시트 원본과 다를 수 있음 — 으로 오염됨)
_ACTIVE_BIZ_TYPES: set = _compute_active_biz_types()

# 지자체별 고유 색상 (20개+ 명확히 구분되는 색 — 권역 내에서도 차별화)
MUNICIPALITY_COLORS = {
    # 서울권 — 파랑 계열 (진→연)
    "서초구청":             "#0D47A1",  # 짙은 남색
    "강북구청":             "#1976D2",  # 파랑
    "마포구청":             "#42A5F5",  # 하늘파랑
    "광진구청":             "#90CAF9",  # 연파랑
    # 경기권 — 초록 계열 (진→연)
    "경기도청":             "#1B5E20",  # 짙은 초록
    "용인시청":             "#388E3C",  # 초록
    "용인시청통합돌봄":     "#66BB6A",  # 밝은 초록
    "포천시청":             "#00695C",  # 청록초록
    "광명시청":             "#26A69A",  # 청록
    "양평군청":             "#80CBC4",  # 연청록
    # 충청권 — 주황/갈색 계열 (진→연)
    "청주시청":             "#BF360C",  # 짙은 벽돌
    "진천군청":             "#E64A19",  # 진주황
    "음성군청":             "#FF7043",  # 주황
    "괴산군청":             "#9E6B00",  # 황갈색
    "증평군청":             "#F9A825",  # 황금
    "충북사회서비스원":     "#FF8F00",  # 앰버
    "충남사회서비스원":     "#795548",  # 갈색
    # 영남권 — 보라 계열
    "금정구청":             "#4A148C",  # 짙은 보라
    "경남사회서비스원":     "#8E24AA",  # 보라
    # 강원권 — 청록/시안 계열 (진→연, 개별 구분)
    "강릉시청":             "#006064",  # 짙은 청록
    "강원사회서비스원":     "#00838F",  # 청록
    "홍천군청":             "#00ACC1",  # 밝은 청록
    "삼척시청":             "#0288D1",  # 하늘파랑 (강원이지만 파랑 계열로 분리)
    "양양군청":             "#01579B",  # 짙은 하늘
    "정선군청":             "#00BFA5",  # 청록민트
    "고성군청":             "#26C6DA",  # 연사이안
    # 호남권
    "광주동구청":           "#558B2F",  # 올리브그린
    # 제주권 — 빨강/핑크 계열
    "제주시청":             "#B71C1C",  # 짙은 빨강
    "서귀포시청":           "#E91E63",  # 핫핑크 (제주와 구분)
    # 기타
    "독거노인지원종합센터": "#546E7A",  # 청회색
    "독거노인":             "#546E7A",
    "희망나래장애인복지관": "#90A4AE",  # 연청회색
    "희망나래":             "#90A4AE",
    "다살림재가노인지원서비스센터": "#8D6E63",  # 브라운그레이
    # 인천권 — 인디고 계열
    "계양구청":             "#283593",  # 짙은 인디고
    "연수구청":             "#3949AB",  # 인디고
    # 강원권 추가
    "영월군청":             "#004D40",  # 짙은 청록
}

def plot_weekly_series(df, x_col, y_col, title, color="#2F5496", height=300):
    """주간 시계열 차트 — 최신 포인트 강조 + 전주 대비 변화 표시"""
    df = shorten_dates_in_df(df, x_col)
    # 최신 13주만 표시 (한 화면에 맞춤)
    df = df.tail(13).reset_index(drop=True)
    fig = px.area(df, x=x_col, y=y_col, color_discrete_sequence=[color])

    # 최신 데이터 포인트 강조 + 전주 대비 어노테이션
    if len(df) >= 2:
        latest = df.iloc[-1]
        prev = df.iloc[-2]
        latest_val = float(latest[y_col]) if not pd.isna(latest[y_col]) else 0
        prev_val = float(prev[y_col]) if not pd.isna(prev[y_col]) else 0
        delta = latest_val - prev_val
        pct = (delta / prev_val * 100) if prev_val != 0 else 0
        arrow = "▲" if delta > 0 else "▼" if delta < 0 else "→"
        color_d = "#00C853" if delta >= 0 else "#FF4B4B"

        fig.add_trace(go.Scatter(
            x=[latest[x_col]], y=[latest_val],
            mode="markers+text",
            marker=dict(size=13, color=color, line=dict(width=2, color="white")),
            text=[f"{arrow} {abs(delta):,.0f} ({abs(pct):.1f}%)"],
            textposition="top center",
            textfont=dict(size=13, color=color_d, family="Pretendard, Noto Sans KR"),
            showlegend=False,
            hoverinfo="skip",
        ))

    n = len(df)
    fig.update_layout(
        title=title, height=height,
        margin=dict(t=40, b=60, l=40, r=60),
        hovermode="x unified",
        xaxis=dict(type="category", title="", range=[-0.5, n - 0.5]),
        yaxis_title="",
    )
    fig.update_traces(
        hovertemplate="<b>%{x}</b><br>값: %{y:,.0f}<extra></extra>",
        line=dict(width=2),
        fillcolor=f"rgba({int(color[1:3],16)},{int(color[3:5],16)},{int(color[5:7],16)},0.15)",
        selector=dict(type="scatter", fill="tozeroy"),
    )
    st.plotly_chart(fig, use_container_width=True)

def plot_bar_rate_dual(df, x_col, bar_col, bar_label, bar_color,
                       line_col, line_label, line_color,
                       title, bar_unit="명", line_unit="%", height=430):
    """이용자수(막대) + 이용률(꺾은선) 통합 듀얼 Y축 차트 — 모든 서비스 페이지 공통"""
    fig = make_subplots(specs=[[{"secondary_y": True}]])
    valid = df[df[x_col].astype(str).str.strip() != "nan"].copy()
    bar_vals = valid[bar_col].apply(safe_numeric)
    # 막대: 반투명으로 배경 처리 → 꺾은선(이용비중)이 더 잘 보임
    fig.add_trace(go.Bar(
        x=valid[x_col], y=bar_vals, name=bar_label,
        marker_color=bar_color, opacity=0.55,
        text=bar_vals.apply(lambda v: f"{int(v):,}" if v == int(v) else f"{v:.1f}"),
        textposition="outside", textfont=dict(size=14, color="#222", family="Noto Sans KR"),
        hovertemplate=f"<b>%{{x}}</b><br>{bar_label}: %{{y:,}}{bar_unit}<extra></extra>"
    ), secondary_y=False)
    if line_col and line_col in valid.columns:
        line_vals = valid[line_col].apply(safe_numeric)
        # 꺾은선: 굵고 선명하게 — 이용비중이 핵심 지표
        fig.add_trace(go.Scatter(
            x=valid[x_col], y=line_vals, name=line_label,
            mode="lines+markers+text",
            line=dict(color=line_color, width=3),
            marker=dict(size=10, color=line_color,
                        line=dict(color="white", width=2)),
            text=line_vals.apply(lambda v: f"<b>{v:.1f}{line_unit}</b>"),
            textposition="top center",
            textfont=dict(size=15, color=line_color, family="Noto Sans KR"),
            hovertemplate=f"<b>%{{x}}</b><br>{line_label}: %{{y:.1f}}{line_unit}<extra></extra>"
        ), secondary_y=True)
    fig.update_layout(
        title=title, height=height, hovermode="x unified",
        xaxis=dict(type="category"),
        legend=LEGEND_BELOW, margin=dict(t=40, b=70), bargap=0.3,
    )
    fig.update_yaxes(title_text=bar_unit, secondary_y=False)
    fig.update_yaxes(title_text=line_unit, secondary_y=True, showgrid=False)
    st.plotly_chart(fig, use_container_width=True)


def extract_mun_ratio_trend(raw_df: pd.DataFrame) -> pd.DataFrame:
    """시트 원본에서 이용자비중 컬럼(AI~BK)을 (주차, 지자체명, 값) long format으로 변환"""
    if raw_df.empty:
        return pd.DataFrame()
    week_col = next((c for c in raw_df.columns if "주차" in str(c)), None)
    if week_col is None:
        return pd.DataFrame()
    ratio_cols = [c for c in raw_df.columns if "이용자비중" in str(c)]
    if not ratio_cols:
        return pd.DataFrame()

    sub = raw_df[[week_col] + ratio_cols].copy()
    sub = sub[~sub[week_col].astype(str).str.strip().isin(["", "nan"])]
    sub[week_col] = sub[week_col].astype(str).str.strip()
    for col in ratio_cols:
        sub[col] = sub[col].apply(safe_numeric)

    long = sub.melt(id_vars=[week_col], value_vars=ratio_cols,
                    var_name="_col", value_name="값")

    def _mun_name(s):
        s = str(s).strip()
        # "경남사회서비스원 이용자비중\n이용자비중" 등 중복 제거
        s = s.replace("\n이용자비중", "").replace(" 이용자비중", "")
        return s.strip()

    long["지자체명"] = long["_col"].apply(_mun_name)
    long = long.rename(columns={week_col: "주차"})
    return long[["주차", "지자체명", "값"]].reset_index(drop=True)


def plot_municipality_bar(df, value_col, title, color_map=None, height=400):
    """지자체별 바 차트 (내림차순 정렬)"""
    df_sorted = df.sort_values(value_col, ascending=True).copy()
    df_sorted["지자체명"] = df_sorted["지자체명"].apply(_mun_label)
    fig = px.bar(df_sorted, y="지자체명", x=value_col, orientation="h",
                 color="권역" if "권역" in df_sorted.columns else None,
                 color_discrete_map={"수도권": "#2F5496", "비수도권": "#FF6F00", "기관": "#7B1FA2", "기타": "#9E9E9E"},
                 height=max(height, len(df_sorted) * 28))
    fig.update_layout(
        title=title, margin=dict(t=40, b=10, l=10, r=60),
        xaxis_title="", yaxis_title="",
        legend=dict(orientation="h", yanchor="top", y=-0.15, xanchor="center", x=0.5),
    )
    fig.update_traces(
        hovertemplate="<b>%{y}</b><br>%{x:,.1f}<extra></extra>",
    )
    st.plotly_chart(fig, use_container_width=True)

def page_week_range_selector(key_prefix: str, all_weeks: list):
    """페이지 내 주차 범위 선택기 (시작~끝 주차)
    기본 시작: 25-52 주차 (없으면 최근 12주)
    Returns: (start_week, end_week) or (None, None)
    """
    if not all_weeks:
        return None, None

    # 기본 시작 인덱스: 25-52 주차
    default_start = "25-52"
    if default_start in all_weeks:
        default_idx = all_weeks.index(default_start)
    else:
        # 25-52가 없으면 26-01 시도
        if "26-01" in all_weeks:
            default_idx = all_weeks.index("26-01")
        else:
            default_idx = max(0, len(all_weeks) - 12)

    with st.expander("📅 기간 설정 (펼쳐서 변경)", expanded=False):
        col1, col2 = st.columns(2)
        with col1:
            start_week = st.selectbox(
                "시작 주차", all_weeks,
                index=default_idx,
                key=f"{key_prefix}_start"
            )
        with col2:
            start_idx = all_weeks.index(start_week) if start_week in all_weeks else 0
            end_options = all_weeks[start_idx:]
            end_week = st.selectbox(
                "종료 주차", end_options,
                index=len(end_options) - 1,  # 기본: 마지막 주차
                key=f"{key_prefix}_end"
            )
        st.caption(f"선택 기간: {start_week} ~ {end_week}")
    return start_week, end_week


def filter_by_week_range(df, week_col, start_week, end_week, all_weeks):
    """DataFrame을 주차 범위로 필터링"""
    if df.empty or not start_week or not end_week:
        return df
    if week_col not in df.columns:
        return df
    start_idx = all_weeks.index(start_week) if start_week in all_weeks else 0
    end_idx = all_weeks.index(end_week) if end_week in all_weeks else len(all_weeks) - 1
    valid_weeks = set(all_weeks[start_idx:end_idx + 1])
    return df[df[week_col].astype(str).str.strip().isin(valid_weeks)]


def get_active_agencies_for_week(week: str) -> list:
    """특정 주차 기준으로 계약 중인 지자체 반환"""
    agencies = get_agency_master()
    if agencies.empty:
        return []

    # 주차 → 시작일 매핑
    wu = data.get("weekly_users", pd.DataFrame())
    target_date = None
    if not wu.empty and "주차" in wu.columns and "시작일" in wu.columns:
        match = wu[wu["주차"].astype(str).str.strip() == week]
        if not match.empty:
            target_date = str(match.iloc[0].get("시작일", "")).strip()

    active = []
    for _, row in agencies.iterrows():
        if row.get("is_active", 0) != 1:
            continue
        # 계약 기간 체크
        if target_date:
            contract_start = str(row.get("contract_start", "")).strip()
            contract_end = str(row.get("contract_end", "")).strip()
            # 시작일 이전이면 제외
            if contract_start and target_date < contract_start:
                continue
            # 종료일 이후면 제외
            if contract_end and contract_end != "" and target_date > contract_end:
                continue
        active.append(row["agency_name"])
    return active


def plot_municipality_lines(df_long, title, height=350, metric_label="값", show_avg=True):
    """지자체별 주간 추이 라인 차트 — 10개 이상이면 Top/Bottom 5 포커스"""
    if df_long.empty:
        st.info("데이터 없음")
        return

    x_col = "주차"
    df_long = shorten_dates_in_df(df_long, x_col)
    mun_count = df_long["지자체명"].nunique()

    # 지자체가 10개 이상이면 포커스 모드
    if mun_count >= 10:
        latest_week = df_long[x_col].max()
        latest = df_long[df_long[x_col] == latest_week].copy()
        top5 = latest.nlargest(5, "값")["지자체명"].tolist()
        bot5 = latest.nsmallest(5, "값")["지자체명"].tolist()

        view_mode = st.radio(
            "표시 모드", ["전체", "Top 5", "Bottom 5", "Top 5 + Bottom 5"],
            index=3, horizontal=True, key=f"view_{hash(title) % 100000}"
        )
        if view_mode == "Top 5":
            df_long = df_long[df_long["지자체명"].isin(top5)]
        elif view_mode == "Bottom 5":
            df_long = df_long[df_long["지자체명"].isin(bot5)]
        elif view_mode == "Top 5 + Bottom 5":
            df_long = df_long[df_long["지자체명"].isin(top5 + bot5)]

    df_plot = df_long.copy()
    df_plot["지자체명"] = df_plot["지자체명"].apply(_mun_label)
    fig = px.line(df_plot, x=x_col, y="값", color="지자체명", markers=True)

    # 전체 평균 참조선 (점선)
    if show_avg and not df_long.empty:
        avg_by_week = df_long.groupby(x_col)["값"].mean().reset_index()
        fig.add_trace(go.Scatter(
            x=avg_by_week[x_col], y=avg_by_week["값"],
            mode="lines", name="── 전체 평균",
            line=dict(color="#333", width=3, dash="dash"),
            hovertemplate="평균: %{y:,.1f}<extra></extra>"
        ))

    fig.update_layout(
        title=title, height=height,
        margin=dict(t=40, b=60, l=40, r=60),
        hovermode="x unified",
        legend=LEGEND_BELOW,
        xaxis=dict(type="category", title=""),
        yaxis_title=metric_label,
    )
    fig.update_traces(hovertemplate="%{y:,.0f}<extra>%{fullData.name}</extra>")
    st.plotly_chart(fig, use_container_width=True)


# ============================================================
# 📋 Summary 페이지
# ============================================================
# ── 사업구분 필터 헬퍼 ──────────────────────────────────────
_BIZ_ORDER_ALL = ["전체", "통합돌봄", "노인맞춤돌봄", "고독사예방", "취약지지원", "장애인지원", "퇴원환자지원", "기타"]

def _active_biz_opts():
    """협약인원 > 0인 기관이 존재하는 사업구분만 순서 유지해 반환 (전체 포함)"""
    return [b for b in _BIZ_ORDER_ALL if b == "전체" or b in _ACTIVE_BIZ_TYPES]

def biz_filter_df(df, biz, col="지자체명"):
    """페이지 내 사업구분 필터 적용"""
    if biz == "전체" or df.empty or col not in df.columns:
        return df
    def _match(name):
        name_n = str(name).replace(" ", "")
        for k, v in BUSINESS_TYPE_MAP.items():
            if v == biz:
                k_n = k.replace(" ", "")
                if k_n == name_n or k_n in name_n or name_n in k_n:
                    return True
        return False
    return df[df[col].apply(_match)]

def biz_selector(key):
    """페이지 내 사업구분 선택 위젯 — 실제 데이터에 존재하는 사업구분만 표시"""
    opts = _active_biz_opts()
    return st.radio("사업구분 선택", opts, horizontal=True,
                    label_visibility="collapsed", key=f"biz_{key}")

def biz_filter_wide_cols(cols, biz):
    """Wide-format DataFrame의 컬럼 목록에서 선택된 사업구분 소속 지자체 컬럼만 반환.
    biz == '전체'면 원본 그대로 반환."""
    if biz == "전체":
        return cols
    result = []
    for col in cols:
        col_n = str(col).replace("\n", " ").strip().replace(" ", "")
        for k, v in BUSINESS_TYPE_MAP.items():
            if v == biz:
                k_n = k.replace(" ", "")
                if k_n == col_n or k_n in col_n or col_n in k_n:
                    result.append(col)
                    break
    return result

def biz_agg_raw(df, biz, week_col):
    """사업구분 소속 지자체 이용자수 합산 + 이용비중 계산 → (count_Series, ratio_Series) or (None, None)"""
    if biz == "전체" or df.empty:
        return None, None

    def _norm(s, strip_ratio=False):
        s = str(s).replace("\n", " ").strip()
        if strip_ratio:
            s = s.replace("이용자비중", "").strip()
        return s.replace(" ", "")

    def _biz_match(n):
        for k, v in BUSINESS_TYPE_MAP.items():
            if v == biz:
                k_n = k.replace(" ", "")
                if k_n == n or k_n in n or n in k_n:
                    return True
        return False

    skip = {"이용자비중", "합계", "비중", "전체", "_"}
    count_cols = [c for c in df.columns
                  if c != week_col
                  and not any(kw in str(c) for kw in skip)]
    ratio_cols  = [c for c in df.columns
                   if "이용자비중" in str(c) and "전체이용비중" not in str(c)]

    matched_count = [c for c in count_cols if _biz_match(_norm(c))]
    matched_ratio = [c for c in ratio_cols  if _biz_match(_norm(c, True))]

    if not matched_count:
        return None, None

    biz_count = df[matched_count].apply(lambda s: s.apply(safe_numeric)).sum(axis=1)

    # 이용비중 = 사업구분 이용자수 / 사업구분 전체등록자수 × 100
    # 전체등록자수 = 이용자수 / (이용자비중/100) 로 역산
    biz_ratio = None
    if matched_ratio:
        ratio_norm_map = {_norm(c, True): c for c in matched_ratio}
        total_enrolled = pd.Series(0.0, index=df.index)
        n_matched = 0
        for cc in matched_count:
            cn = _norm(cc)
            rc = next((ratio_norm_map[rn] for rn in ratio_norm_map
                       if cn == rn or cn in rn or rn in cn), None)
            if rc is not None:
                cnt = df[cc].apply(safe_numeric)
                rat = df[rc].apply(safe_numeric)
                total_enrolled += cnt.div(rat / 100).where(rat > 0, 0)
                n_matched += 1
        if n_matched > 0:
            biz_ratio = biz_count.div(total_enrolled).where(total_enrolled > 0, 0) * 100

    return biz_count, biz_ratio

if page == "📋 Summary":
    if selected_week:
        summary = cached_week_summary(sheets, data, selected_week)
        prev_week = get_prev_week(selected_week)
        prev_summary = cached_week_summary(sheets, data, prev_week) if prev_week else {}


        st.markdown(f'<div class="section-header">📅 {selected_week}주차 ({summary.get("시작일", "")}) 운영 현황</div>', unsafe_allow_html=True)

        # 회원가입 현황 — 이용자주간 선택 주차 기준 (스냅샷 시트보다 최신)
        reg = data.get("registration", pd.DataFrame())
        total_contract = 0
        total_registered = 0
        total_incomplete = 0

        wu = data.get("weekly_users", pd.DataFrame())
        total_reg_rate = 0
        if not wu.empty and "주차" in wu.columns:
            _wu_sel = wu[wu["주차"].astype(str).str.strip() == selected_week]
            if _wu_sel.empty:
                _wu_sel = wu.iloc[[-1]]
            if not _wu_sel.empty:
                _row = _wu_sel.iloc[0]
                # 협약인원 (대상자수)
                if "대상자수" in wu.columns:
                    _v = safe_numeric(_row.get("대상자수", 0))
                    if _v and _v > 0:
                        total_contract = int(_v)
                # 가입완료합계 — 이용자주간 해당 주차 값 우선 사용
                if "가입완료합계" in wu.columns:
                    _r = safe_numeric(_row.get("가입완료합계", 0))
                    if _r and _r > 0:
                        total_registered = int(_r)
        # fallback: 이용자현황 스냅샷
        if total_contract == 0 and not reg.empty and "협약인원" in reg.columns:
            total_contract = int(reg["협약인원"].sum())
        if total_registered == 0 and not reg.empty and "가입완료" in reg.columns:
            total_registered = int(reg["가입완료"].sum())
        if not reg.empty and "가입미완료" in reg.columns:
            total_incomplete = total_contract - total_registered
        # 가입률: 가입완료/대상자수 직접 계산
        if total_contract > 0:
            total_reg_rate = round(total_registered / total_contract * 100, 1)

        # 주간 데이터에서 전주 가입완료 수 가져오기
        cur_registered = total_registered
        prev_registered = total_registered  # 기본값
        if not wu.empty and "주차" in wu.columns and "가입완료합계" in wu.columns:
            wu_cur = wu[wu["주차"].astype(str).str.strip() == selected_week]
            if not wu_cur.empty:
                cur_registered = safe_numeric(wu_cur.iloc[0].get("가입완료합계", total_registered))
            if prev_week:
                wu_prev = wu[wu["주차"].astype(str).str.strip() == prev_week]
                if not wu_prev.empty:
                    prev_registered = safe_numeric(wu_prev.iloc[0].get("가입완료합계", total_registered))

        delta_registered = cur_registered - prev_registered

        # KPI 카드
        cols = st.columns(4)
        kpi_data = [
            ("총 협약인원", total_contract, 0, "명", "metric-card", False, ",.0f"),
            ("총 가입완료", cur_registered, prev_registered, "명", "metric-card-green", False, ",.0f"),
            ("전체 가입률", total_reg_rate, 0, "%", "metric-card", False, ".1f"),
            ("안부확인율", round(float(summary.get("안부확인율", 0)), 1), round(float(prev_summary.get("안부확인율", 0)), 1), "%", "metric-card-orange", False, ".1f"),
        ]
        for col, (label, val, prev_val, suffix, card_cls, invert, fmt) in zip(cols, kpi_data):
            delta = float(val) - float(prev_val) if prev_val else 0
            val_str = format(val, fmt)
            with col:
                st.markdown(f"""
                <div class="{card_cls}">
                    <h3>{label}</h3>
                    <h1>{val_str}{suffix}</h1>
                    <p>{delta_html(delta, suffix, invert, prev_val=float(prev_val) if prev_val else 0)}</p>
                </div>
                """, unsafe_allow_html=True)

        st.markdown("")

        # ── 사업구분별 지자체 / 이용자 수 ─────────────────────────────
        st.markdown('<div class="section-header">사업구분별 지자체 / 이용자 수</div>', unsafe_allow_html=True)
        _reg_df_s = data.get("registration", pd.DataFrame())

        def _biz_classify(name):
            name_n = str(name).replace(" ", "")
            for k, v in BUSINESS_TYPE_MAP.items():
                if name_n == k.replace(" ", "") or k.replace(" ", "") in name_n or name_n in k.replace(" ", ""):
                    return v
            return "기타"

        _all_biz_s = {b: {"count": 0, "users": 0, "names": []} for b in BUSINESS_TYPE_ORDER}
        if not _reg_df_s.empty and "지자체명" in _reg_df_s.columns:
            for _, _row in _reg_df_s.iterrows():
                # 협약인원이 0인 패딩 행 제외 (계약 없는 placeholder)
                if safe_numeric(_row.get("협약인원", 0)) <= 0:
                    continue
                _nm = str(_row["지자체명"]).strip()
                _b  = _biz_classify(_nm)
                if _b not in _all_biz_s:
                    _all_biz_s[_b] = {"count": 0, "users": 0, "names": []}
                if _nm not in _all_biz_s[_b]["names"]:
                    _all_biz_s[_b]["count"] += 1
                    _all_biz_s[_b]["names"].append(_nm)
                _all_biz_s[_b]["users"] += int(safe_numeric(_row.get("협약인원", 0)))

        _ordered = [b for b in BUSINESS_TYPE_ORDER if b in _all_biz_s and _all_biz_s[b]["count"] > 0]
        _bcols = st.columns(len(_ordered))
        for _col, _b in zip(_bcols, _ordered):
            _info  = _all_biz_s[_b]
            _color = BUSINESS_TYPE_COLORS.get(_b, "#666")
            _col.markdown(
                f"""<div style="background:#fff;border:2px solid {_color};border-radius:14px;
                padding:14px 12px;text-align:center">
                <div style="font-size:12px;font-weight:700;color:{_color}">{_b}</div>
                <div style="font-size:24px;font-weight:800;color:#1e2533">{_info['count']}</div>
                <div style="font-size:11px;color:#6b7488">지자체</div>
                <div style="font-size:16px;font-weight:700;color:{_color}">{_info['users']:,}</div>
                <div style="font-size:11px;color:#6b7488">이용자(협약)</div>
                </div>""",
                unsafe_allow_html=True,
            )

        st.markdown("")
        with st.expander("📋 사업구분별 지자체 목록", expanded=False):
            for _b in _ordered:
                _names = _all_biz_s[_b]["names"]
                _color = BUSINESS_TYPE_COLORS.get(_b, "#666")
                st.markdown(
                    f"**<span style='color:{_color}'>{_b}</span>** ({len(_names)}개): "
                    + " · ".join(_names),
                    unsafe_allow_html=True,
                )
        st.markdown("---")

        # 지자체 계약 현황 (베이직/세이프) + 계약 시작/종료 알림
        agency_sum = get_agency_summary(sheets)
        if agency_sum["total"] > 0:
            st.markdown('<div class="section-header">지자체 계약 현황</div>', unsafe_allow_html=True)

            # 상단 3개 카드: 활성 / 세이프 / 베이직
            acols = st.columns(3)
            with acols[0]:
                st.markdown(f"""
                <div style="text-align:center; padding:10px; border-radius:12px;
                            background:#E3F2FD; border:2px solid #1565C0;">
                    <span style="font-size:1.8rem; font-weight:700; color:#1565C0;">{agency_sum['active']}</span>
                    <br><span style="font-size:0.85rem; color:#555;">활성 지자체</span>
                </div>
                """, unsafe_allow_html=True)
            with acols[1]:
                st.markdown(f"""
                <div style="text-align:center; padding:10px; border-radius:12px;
                            background:#E8F5E9; border:2px solid #2E7D32;">
                    <span style="font-size:1.8rem; font-weight:700; color:#2E7D32;">{agency_sum['safe']}</span>
                    <br><span style="font-size:0.85rem; color:#555;">세이프 (관제)</span>
                </div>
                """, unsafe_allow_html=True)
            with acols[2]:
                st.markdown(f"""
                <div style="text-align:center; padding:10px; border-radius:12px;
                            background:#FFF3E0; border:2px solid #E65100;">
                    <span style="font-size:1.8rem; font-weight:700; color:#E65100;">{agency_sum['basic']}</span>
                    <br><span style="font-size:0.85rem; color:#555;">베이직</span>
                </div>
                """, unsafe_allow_html=True)

            st.markdown("")

            # 계약 시작/종료 알림 게시판 (최근 4주)
            from datetime import datetime, timedelta
            today = datetime.now().strftime("%Y-%m-%d")
            four_weeks_later = (datetime.now() + timedelta(days=28)).strftime("%Y-%m-%d")
            four_weeks_ago = (datetime.now() - timedelta(days=28)).strftime("%Y-%m-%d")

            MODEL_KR = {"safe": "세이프", "safe_plus": "세이프 플러스", "basic": "베이직", "basic_plus": "베이직 플러스"}

            agencies_df = get_agency_master()
            if not agencies_df.empty:
                # 4주 내 계약 시작 예정 (오늘 이후만 — 오늘 시작은 recently_started에 포함)
                starting_soon = agencies_df[
                    (agencies_df["contract_start"] > today) &
                    (agencies_df["contract_start"] <= four_weeks_later) &
                    (agencies_df["contract_start"] != "")
                ].copy().sort_values("contract_start")

                # 4주 내 계약 종료 예정 (오늘 이후만 — 오늘 종료는 recently_ended에 포함)
                ending_soon = agencies_df[
                    (agencies_df["contract_end"] > today) &
                    (agencies_df["contract_end"] <= four_weeks_later) &
                    (agencies_df["contract_end"] != "")
                ].copy().sort_values("contract_end")

                # 최근 4주 내 계약 시작됨
                recently_started = agencies_df[
                    (agencies_df["contract_start"] >= four_weeks_ago) &
                    (agencies_df["contract_start"] <= today) &
                    (agencies_df["contract_start"] != "")
                ].copy().sort_values("contract_start")

                # 최근 4주 내 계약 종료됨
                recently_ended = agencies_df[
                    (agencies_df["contract_end"] >= four_weeks_ago) &
                    (agencies_df["contract_end"] <= today) &
                    (agencies_df["contract_end"] != "")
                ].copy().sort_values("contract_end")

                alert_col1, alert_col2 = st.columns(2)

                with alert_col1:
                    st.markdown("**🟢 계약 시작 (최근 4주)**")
                    if not recently_started.empty:
                        for _, r in recently_started.iterrows():
                            model = MODEL_KR.get(r.get("service_model", ""), r.get("service_model", ""))
                            users = int(r.get("target_users", 0))
                            users_str = f" ({users}명)" if users > 0 else " (규모 미정)"
                            memo = str(r.get("memo", "")).strip()
                            memo_tag = f' &nbsp;<span style="color:#f59e0b;font-size:0.8em">⚠ {memo}</span>' if memo else ""
                            st.markdown(f'<div class="insight-box-success">{r["contract_start"]} {r["agency_name"]} {model} 계약 시작{users_str}{memo_tag}</div>', unsafe_allow_html=True)
                    if not starting_soon.empty:
                        for _, r in starting_soon.iterrows():
                            model = MODEL_KR.get(r.get("service_model", ""), r.get("service_model", ""))
                            users = int(r.get("target_users", 0))
                            users_str = f" ({users}명)" if users > 0 else " (규모 미정)"
                            memo = str(r.get("memo", "")).strip()
                            memo_tag = f' &nbsp;<span style="color:#f59e0b;font-size:0.8em">⚠ {memo}</span>' if memo else ""
                            st.markdown(f'<div class="insight-box">{r["contract_start"]} {r["agency_name"]} {model} 계약 시작 예정{users_str}{memo_tag}</div>', unsafe_allow_html=True)
                    if recently_started.empty and starting_soon.empty:
                        st.markdown('<div class="insight-box">최근 4주 내 계약 시작 없음</div>', unsafe_allow_html=True)

                with alert_col2:
                    st.markdown("**🔴 계약 종료 (최근 4주)**")
                    if not ending_soon.empty:
                        for _, r in ending_soon.iterrows():
                            try:
                                days_left = (datetime.strptime(r["contract_end"], "%Y-%m-%d") - datetime.now()).days
                            except:
                                days_left = 0
                            model = MODEL_KR.get(r.get("service_model", ""), r.get("service_model", ""))
                            st.markdown(f'<div class="insight-box-danger">{r["contract_end"]} {r["agency_name"]} {model} 계약 종료 예정 ({days_left}일 남음)</div>', unsafe_allow_html=True)
                    if not recently_ended.empty:
                        for _, r in recently_ended.iterrows():
                            model = MODEL_KR.get(r.get("service_model", ""), r.get("service_model", ""))
                            st.markdown(f'<div class="insight-box-danger">{r["contract_end"]} {r["agency_name"]} {model} 계약 종료</div>', unsafe_allow_html=True)
                    if ending_soon.empty and recently_ended.empty:
                        st.markdown('<div class="insight-box">최근 4주 내 계약 종료 없음</div>', unsafe_allow_html=True)

                # 4주 이후 계약 예정 (memo에 "계약 예정" 포함 or contract_start > four_weeks_later)
                upcoming_planned = agencies_df[
                    (agencies_df["contract_start"] > four_weeks_later) &
                    (agencies_df["contract_start"] != "")
                ].copy().sort_values("contract_start")
                if not upcoming_planned.empty:
                    st.markdown("**🗓 계약 예정 (4주 이후)**")
                    for _, r in upcoming_planned.iterrows():
                        model = MODEL_KR.get(r.get("service_model", ""), r.get("service_model", ""))
                        users = int(r.get("target_users", 0))
                        users_str = f" · {users}명" if users > 0 else ""
                        memo = str(r.get("memo", ""))
                        memo_str = f" ({memo})" if memo else ""
                        st.markdown(
                            f'<div class="insight-box" style="border-left:4px solid #7C3AED;">'
                            f'📌 {r["contract_start"]} &nbsp;{r["agency_name"]} &nbsp;{model}{users_str}{memo_str}'
                            f'</div>',
                            unsafe_allow_html=True
                        )

            st.markdown("")

            # 세이프 대상 지자체 현황 — Google Sheet 우선, SQLite fallback
            safe_status = get_safe_status_from_sheet()
            if safe_status.empty:
                try:
                    from local_db import get_connection as _gc
                    _conn = _gc()
                    safe_status = pd.read_sql_query("SELECT * FROM safe_agency_status ORDER BY monitoring_start_date", _conn)
                    _conn.close()
                except Exception:
                    safe_status = pd.DataFrame()

            st.markdown('<div class="section-header">🛡 세이프 대상 지자체 현황</div>', unsafe_allow_html=True)

            if not safe_status.empty:
                # 요약 카드
                total_safe_target = int(safe_status["contract_users"].sum())
                total_monitoring = int(safe_status["registered_users"].sum())
                total_call = int(safe_status["joined_users"].sum())
                avg_m_rate = round(total_monitoring / total_safe_target * 100, 1) if total_safe_target > 0 else 0
                avg_c_rate = round(total_call / total_safe_target * 100, 1) if total_safe_target > 0 else 0

                sc1, sc2, sc3, sc4 = st.columns(4)
                with sc1:
                    st.markdown(f'<div style="text-align:center; padding:6px; border-radius:10px; background:#E8F5E9; border:1px solid #2E7D32;"><b style="font-size:1.2rem; color:#2E7D32;">{total_safe_target:,}명</b><br><span style="font-size:0.75rem;">세이프 대상 합계</span></div>', unsafe_allow_html=True)
                with sc2:
                    st.markdown(f'<div style="text-align:center; padding:6px; border-radius:10px; background:#E3F2FD; border:1px solid #1565C0;"><b style="font-size:1.2rem; color:#1565C0;">{total_monitoring:,}명</b><br><span style="font-size:0.75rem;">등록 이용자</span></div>', unsafe_allow_html=True)
                with sc3:
                    st.markdown(f'<div style="text-align:center; padding:6px; border-radius:10px; background:#FFF3E0; border:1px solid #E65100;"><b style="font-size:1.2rem; color:#E65100;">{avg_m_rate}%</b><br><span style="font-size:0.75rem;">등록 이용률</span></div>', unsafe_allow_html=True)
                with sc4:
                    st.markdown(f'<div style="text-align:center; padding:6px; border-radius:10px; background:#FCE4EC; border:1px solid #C62828;"><b style="font-size:1.2rem; color:#C62828;">{avg_c_rate}%</b><br><span style="font-size:0.75rem;">가입 이용률</span></div>', unsafe_allow_html=True)

                st.markdown("")

                # 데이터 테이블
                safe_display = safe_status[["monitoring_start_date", "memo", "agency_name", "contract_users",
                                             "registered_users", "joined_users", "registered_rate", "joined_rate"]].copy()
                safe_display.columns = ["관제시작일", "비고", "지자체명", "계약인원", "등록이용자", "가입이용자", "등록이용률(%)", "가입이용률(%)"]
                st.dataframe(safe_display, use_container_width=True, hide_index=True)

            # 계약 시작 / 명단 미등록 지자체 메모
            st.markdown("**📋 계약 시작 / 명단 미등록 지자체**")
            # Google Sheet '메모' 탭 우선 → 없으면 로컬 SQLite fallback
            _pending_note = get_sheet_note("pending_agencies") or get_note("pending_agencies", "영월군청")
            _note_col, _btn_col = st.columns([5, 1])
            with _note_col:
                _new_note = st.text_area(
                    "계약은 체결됐지만 아직 명단을 제출하지 않은 지자체를 메모해두세요.",
                    value=_pending_note,
                    height=80,
                    key="pending_agencies_note",
                    label_visibility="collapsed",
                )
            with _btn_col:
                if st.button("저장", key="save_pending_note", use_container_width=True):
                    _saved = save_sheet_note("pending_agencies", _new_note)
                    if _saved:
                        st.success("저장됨 ✓")
                    else:
                        save_note("pending_agencies", _new_note)
                        st.warning("시트 저장 실패 — 로컬에만 저장됨")

            # 파일 업로드 + 로컬 경로 + 수동 편집
            _safe_expander_open = st.session_state.get("safe_expander_open", False)
            with st.expander("📤 세이프 현황 업데이트 (엑셀 업로드 또는 수동 편집)", expanded=_safe_expander_open):
                upload_tab, path_tab, manual_tab = st.tabs(["📁 엑셀 파일 업로드", "📂 로컬 파일 경로 (VPN 우회)", "✏️ 수동 편집"])

                with path_tab:
                    st.caption("VPN으로 파일 업로드가 막힐 때 사용하세요. 서버가 파일을 직접 읽습니다.")
                    _local_path = st.text_input(
                        "엑셀 파일 경로 입력",
                        placeholder=r"예) C:\Users\NHN\Downloads\계산용.xlsx",
                        key="safe_local_path"
                    )
                    if st.button("📂 경로로 파일 읽기", key="safe_local_load"):
                        import os
                        if _local_path and os.path.exists(_local_path):
                            with open(_local_path, "rb") as _f:
                                st.session_state["safe_upload_bytes"] = _f.read()
                                st.session_state["safe_upload_name"] = os.path.basename(_local_path)
                                st.session_state["safe_expander_open"] = True
                            st.success(f"✅ 파일 읽기 성공: {os.path.basename(_local_path)}")
                            st.rerun()
                        elif _local_path:
                            st.error(f"❌ 파일을 찾을 수 없습니다: {_local_path}")
                        else:
                            st.warning("파일 경로를 입력해주세요.")

                with upload_tab:
                    st.caption("계산용.xlsx 같은 엑셀 파일을 업로드하면 자동으로 반영됩니다.")
                    st.caption("필수 컬럼: 관제 등록 날짜, 구분, 지자체명, 계약 인원, 전체 등록 이용자, 전체 가입한 이용자")
                    uploaded_file = st.file_uploader("엑셀 파일 업로드 (.xlsx)", type=["xlsx"], key="safe_upload")

                    # 파일이 업로드되면 bytes를 session_state에 보관 (버튼 클릭 후 rerun돼도 유지)
                    if uploaded_file is not None:
                        st.session_state["safe_upload_bytes"] = uploaded_file.read()
                        st.session_state["safe_upload_name"] = uploaded_file.name
                        st.session_state["safe_expander_open"] = True

                    _upload_bytes = st.session_state.get("safe_upload_bytes")
                    if _upload_bytes is not None:
                        try:
                            import openpyxl, io
                            wb = openpyxl.load_workbook(io.BytesIO(_upload_bytes), data_only=True)
                            # Sheet2 (정리된 데이터) 우선, 없으면 Sheet1
                            ws = wb["Sheet2"] if "Sheet2" in wb.sheetnames else wb[wb.sheetnames[0]]

                            # 헤더 찾기 (첫 번째 비어있지 않은 행)
                            header_row = None
                            for row in ws.iter_rows(min_row=1, max_row=5, values_only=False):
                                vals = [cell.value for cell in row]
                                if any(v is not None and str(v).strip() for v in vals):
                                    header_row = row
                                    break

                            if header_row:
                                headers = [str(cell.value).strip() if cell.value else f"col_{i}" for i, cell in enumerate(header_row)]
                                start_row = header_row[0].row + 1

                                rows_data = []
                                for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row, values_only=True):
                                    vals = list(row)
                                    if len(vals) >= len(headers):
                                        vals = vals[:len(headers)]
                                    else:
                                        vals.extend([None] * (len(headers) - len(vals)))
                                    row_dict = dict(zip(headers, vals))
                                    # 지자체명이 있는 행만
                                    agency = None
                                    for k, v in row_dict.items():
                                        if v and str(v).strip() and any(kw in str(k) for kw in ["지자체", "기관"]):
                                            agency = str(v).strip()
                                            break
                                    if agency and agency != "합계" and agency != "총합":
                                        rows_data.append(row_dict)

                                if rows_data:
                                    preview_df = pd.DataFrame(rows_data)
                                    st.success(f"✅ {len(rows_data)}개 지자체 데이터 감지!")
                                    st.dataframe(preview_df, use_container_width=True, hide_index=True)

                                    if st.button("💾 이 데이터로 세이프 현황 업데이트", key="upload_safe_save"):
                                        try:
                                            # 컬럼명 유연 매칭 → 정규화된 행 목록 생성
                                            normalized_rows = []
                                            for rd in rows_data:
                                                agency_name = ""
                                                start_date = ""
                                                memo = ""
                                                contract = 0
                                                registered = 0
                                                joined = 0

                                                for k, v in rd.items():
                                                    kl = str(k).replace(" ", "")
                                                    if "지자체" in kl or "기관" in kl:
                                                        agency_name = str(v).strip() if v else ""
                                                    elif "날짜" in kl or "시작" in kl or "등록날짜" in kl:
                                                        start_date = str(v).strip() if v else ""
                                                    elif "구분" in kl or "비고" in kl:
                                                        memo = str(v).strip() if v else ""
                                                    elif "계약" in kl and "인원" in kl:
                                                        contract = int(float(v)) if v else 0
                                                    elif "등록" in kl and "이용" in kl:
                                                        registered = int(float(v)) if v else 0
                                                    elif "가입" in kl and "이용" in kl:
                                                        joined = int(float(v)) if v else 0

                                                if not agency_name:
                                                    continue

                                                r_rate = round(registered / contract * 100, 1) if contract > 0 else 0
                                                j_rate = round(joined / contract * 100, 1) if contract > 0 else 0
                                                normalized_rows.append({
                                                    "monitoring_start_date": start_date,
                                                    "memo": memo,
                                                    "agency_name": agency_name,
                                                    "contract_users": contract,
                                                    "registered_users": registered,
                                                    "joined_users": joined,
                                                    "registered_rate": r_rate,
                                                    "joined_rate": j_rate,
                                                })

                                            # 1) Google Sheets에 저장 (영구)
                                            _sheet_saved = save_safe_status_to_sheet(normalized_rows)

                                            # 2) SQLite에도 저장 (세션 내 즉시 반영용)
                                            from local_db import get_connection as _gc2
                                            _conn2 = _gc2()
                                            _conn2.execute("DELETE FROM safe_agency_status")
                                            for nr in normalized_rows:
                                                _conn2.execute("""
                                                    INSERT INTO safe_agency_status
                                                    (monitoring_start_date, memo, agency_name, contract_users, registered_users, joined_users, registered_rate, joined_rate)
                                                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                                                """, (
                                                    nr["monitoring_start_date"], nr["memo"], nr["agency_name"],
                                                    nr["contract_users"], nr["registered_users"], nr["joined_users"],
                                                    nr["registered_rate"], nr["joined_rate"],
                                                ))
                                            _conn2.commit()
                                            _conn2.close()

                                            if _sheet_saved:
                                                st.success("✅ 세이프 현황이 업데이트되었습니다! (시트에 영구 저장됨)")
                                            else:
                                                st.success("✅ 세이프 현황이 업데이트되었습니다!")
                                                st.warning("⚠️ 시트 저장 실패 — 재시작 시 초기화될 수 있습니다. 서비스 계정을 설정해주세요.")

                                            st.session_state.pop("safe_upload_bytes", None)
                                            st.session_state.pop("safe_upload_name", None)
                                            st.session_state["safe_expander_open"] = True
                                            st.rerun()
                                        except Exception as e:
                                            st.error(f"저장 실패: {e}")
                                else:
                                    st.warning("지자체 데이터를 찾을 수 없습니다. 컬럼명을 확인해주세요.")
                        except Exception as e:
                            st.error(f"파일 읽기 실패: {e}")

                with manual_tab:
                    if not safe_status.empty:
                        safe_edit = safe_status[["monitoring_start_date", "memo", "agency_name", "contract_users",
                                                  "registered_users", "joined_users"]].copy()
                        safe_edit.columns = ["관제시작일", "비고", "지자체명", "계약인원", "등록이용자", "가입이용자"]
                    else:
                        safe_edit = pd.DataFrame(columns=["관제시작일", "비고", "지자체명", "계약인원", "등록이용자", "가입이용자"])

                    edited_safe = st.data_editor(
                        safe_edit,
                        use_container_width=True,
                        num_rows="dynamic",
                        key="safe_editor",
                    )

                    if st.button("💾 세이프 현황 저장", key="save_safe_status"):
                        try:
                            normalized_rows_manual = []
                            for _, r in edited_safe.iterrows():
                                agency = str(r.get("지자체명", "")).strip()
                                if not agency:
                                    continue
                                target = int(r.get("계약인원", 0))
                                mon = int(r.get("등록이용자", 0))
                                call = int(r.get("가입이용자", 0))
                                m_rate = round(mon / target * 100, 1) if target > 0 else 0
                                c_rate = round(call / target * 100, 1) if target > 0 else 0
                                normalized_rows_manual.append({
                                    "monitoring_start_date": str(r.get("관제시작일", "")),
                                    "memo": str(r.get("비고", "")),
                                    "agency_name": agency,
                                    "contract_users": target,
                                    "registered_users": mon,
                                    "joined_users": call,
                                    "registered_rate": m_rate,
                                    "joined_rate": c_rate,
                                })

                            _sheet_saved_m = save_safe_status_to_sheet(normalized_rows_manual)

                            from local_db import get_connection as _gc2
                            _conn2 = _gc2()
                            _conn2.execute("DELETE FROM safe_agency_status")
                            for nr in normalized_rows_manual:
                                _conn2.execute("""
                                    INSERT INTO safe_agency_status
                                    (monitoring_start_date, memo, agency_name, contract_users, registered_users, joined_users, registered_rate, joined_rate)
                                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                                """, (nr["monitoring_start_date"], nr["memo"], nr["agency_name"],
                                      nr["contract_users"], nr["registered_users"], nr["joined_users"],
                                      nr["registered_rate"], nr["joined_rate"]))
                            _conn2.commit()
                            _conn2.close()

                            if _sheet_saved_m:
                                st.success("✅ 세이프 현황이 저장되었습니다! (시트에 영구 저장됨)")
                            else:
                                st.success("세이프 현황이 저장되었습니다!")
                                st.warning("⚠️ 시트 저장 실패 — 재시작 시 초기화될 수 있습니다. 서비스 계정을 설정해주세요.")
                            st.rerun()
                        except Exception as e:
                            st.error(f"저장 실패: {e}")

            st.markdown("")

        # 히트맵
        heatmap_df = cached_heatmap(data, selected_week)
        # 해당 주차 기준 계약 중인 지자체만 필터링 (부분 매칭 지원)
        active_list = get_active_agencies_for_week(selected_week)
        if active_list and not heatmap_df.empty:
            def _is_active_fuzzy(name):
                if name in active_list:
                    return True
                for a in active_list:
                    if name in a or a in name:
                        return True
                return False
            heatmap_df = heatmap_df[heatmap_df["지자체명"].apply(_is_active_fuzzy)]
    else:
        st.info("사이드바에서 주차를 선택해주세요.")


# ============================================================
# 🧭 사업구분별 현황
# ============================================================
elif page == "🧭 사업구분별 현황":
    st.markdown('<div class="section-header">🧭 사업구분별 현황</div>', unsafe_allow_html=True)
    st.divider()

    _biz_list = [b for b in BUSINESS_TYPE_ORDER if b in _ACTIVE_BIZ_TYPES]

    # 날짜 → 주차 매핑 (일별 데이터를 주차로 묶기 위함, weekly_users 시작일 기준 7일 전개)
    _wu_bt = data.get("weekly_users", pd.DataFrame())
    _daymap = {}
    if not _wu_bt.empty and "주차" in _wu_bt.columns and "시작일" in _wu_bt.columns:
        for _, _r in _wu_bt.iterrows():
            _rs = pd.to_datetime(str(_r["시작일"]), errors="coerce")
            if pd.isna(_rs):
                continue
            _wk = str(_r["주차"]).strip()
            for _i in range(7):
                _daymap[(_rs + pd.Timedelta(days=_i)).strftime("%Y-%m-%d")] = _wk

    _snap_week = selected_week if selected_week else (weeks[-1] if weeks else None)
    st.markdown(f'<div class="section-header">📅 {_snap_week}주차 사업구분별 현황</div>' if _snap_week
                else '<div class="section-header">사업구분별 현황</div>', unsafe_allow_html=True)
    st.caption("사이드바에서 주차를 바꾸면 이 화면도 해당 주차 기준으로 바뀝니다.")

    _reg_all = data.get("registration", pd.DataFrame())
    _wr_all = data.get("weekly_registered_by_mun", pd.DataFrame())
    _cw_all = data.get("checkin_mun_weekly", pd.DataFrame())
    _checkin_rate_all = data.get("checkin_municipality_rate", pd.DataFrame())
    _hc_all = data.get("건강상담지자체", pd.DataFrame())
    _steps_raw = sheets.get("걸음수현황", pd.DataFrame())
    _STEPS_EXCLUDE = {"WAPLAT", "ai생활지원사테스트", "한전MCS"}
    _steps_all = (_steps_raw[~_steps_raw["agencyName"].isin(_STEPS_EXCLUDE)].copy()
                  if not _steps_raw.empty and "agencyName" in _steps_raw.columns else pd.DataFrame())

    def _wide_week_row(sheet_key, wk):
        raw = sheets.get(sheet_key, pd.DataFrame())
        if raw.empty:
            return pd.DataFrame(), None
        _wc = next((c for c in raw.columns if "주차" in str(c).replace("\n", "").strip()), None)
        if _wc is None:
            return pd.DataFrame(), None
        return raw[raw[_wc].astype(str).str.strip() == str(wk).strip()], _wc

    _cardio_row, _cardio_wc = _wide_week_row("심혈관이용자", _snap_week)
    _stress_row, _stress_wc = _wide_week_row("스트레스이용자", _snap_week)

    _rows = []
    for _biz in _biz_list:
        _row = {"사업구분": _biz}

        _reg_b = biz_filter_df(_reg_all, _biz)
        _active_reg_b = _reg_b.loc[_reg_b["협약인원"].apply(safe_numeric) > 0] if not _reg_b.empty else _reg_b
        _contract = _reg_b["협약인원"].apply(safe_numeric).sum() if not _reg_b.empty else 0
        _reg_completed = _reg_b["가입완료"].apply(safe_numeric).sum() if not _reg_b.empty and "가입완료" in _reg_b.columns else 0
        _row["지자체수"] = _active_reg_b["지자체명"].nunique() if not _active_reg_b.empty else 0
        _row["지자체명목록"] = ", ".join(sorted(_active_reg_b["지자체명"].astype(str).str.strip().unique())) if not _active_reg_b.empty else ""
        _row["이용자(협약)"] = int(_contract)

        # 가입률 — 해당 주차 가입완료 ÷ 현재 협약인원(고정)
        _wr_b = biz_filter_df(_wr_all, _biz)
        _wr_wk = _wr_b[_wr_b["주차"].astype(str).str.strip() == str(_snap_week).strip()] if not _wr_b.empty else pd.DataFrame()
        if not _wr_wk.empty and _contract > 0:
            _reg_cnt = _wr_wk["가입완료"].apply(safe_numeric).sum()
            _row["가입수(명)"] = int(_reg_cnt)
            _row["가입률(%)"] = round(_reg_cnt / _contract * 100, 1)
        else:
            _row["가입수(명)"] = None
            _row["가입률(%)"] = None

        # 안부확인율 — 해당 주차 분자/분모 합산
        _cw_b = biz_filter_df(_cw_all, _biz)
        if not _cw_b.empty:
            _cw_b2 = _cw_b.copy()
            _cw_b2["_wk"] = _cw_b2["시작일"].astype(str).str.strip().map(_daymap)
            _cw_wk = _cw_b2[_cw_b2["_wk"] == _snap_week]
            _den = _cw_wk["분모"].apply(safe_numeric).sum()
            _num = _cw_wk["분자"].apply(safe_numeric).sum()
            _row["안부확인율(%)"] = round(_num / _den * 100, 1) if _den > 0 else None
        else:
            _row["안부확인율(%)"] = None

        # 안부체크율 — 해당 주차 발송/응답 합산
        _chk_b = biz_filter_df(_checkin_rate_all, _biz)
        if not _chk_b.empty and "안부체크발송" in _chk_b.columns:
            _chk_b2 = _chk_b.copy()
            _chk_b2["_wk"] = _chk_b2["시작일"].astype(str).str.strip().map(_daymap)
            _chk_wk = _chk_b2[_chk_b2["_wk"] == _snap_week]
            _send = _chk_wk["안부체크발송"].apply(safe_numeric).sum()
            _resp = _chk_wk["안부체크응답"].apply(safe_numeric).sum()
            # off대상자는 지자체별 고정값(일별로 반복 저장됨) — 지자체당 1회만 집계
            _off = (_chk_wk.drop_duplicates(subset=["지자체명"])["off대상자"].apply(safe_numeric).sum()
                    if "off대상자" in _chk_wk.columns else 0)
            _denom = _send - _off
            _row["안부체크율(%)"] = round(_resp / _denom * 100, 1) if _denom > 0 else None
        else:
            _row["안부체크율(%)"] = None

        # 심혈관/스트레스 이용비중 — 시트에 내장된 비율 대신 직접 계산 (이용자수 ÷ 가입완료)
        if not _cardio_row.empty:
            _c_cnt, _ = biz_agg_raw(_cardio_row, _biz, _cardio_wc)
            _row["심혈관 이용자(명)"] = int(_c_cnt.iloc[0]) if _c_cnt is not None else None
            _row["심혈관 이용비중(%)"] = (round(_c_cnt.iloc[0] / _reg_completed * 100, 1)
                                    if _c_cnt is not None and _reg_completed > 0 else None)
        else:
            _row["심혈관 이용자(명)"] = None
            _row["심혈관 이용비중(%)"] = None

        if not _stress_row.empty:
            _s_cnt, _ = biz_agg_raw(_stress_row, _biz, _stress_wc)
            _row["스트레스 이용자(명)"] = int(_s_cnt.iloc[0]) if _s_cnt is not None else None
            _row["스트레스 이용비중(%)"] = (round(_s_cnt.iloc[0] / _reg_completed * 100, 1)
                                    if _s_cnt is not None and _reg_completed > 0 else None)
        else:
            _row["스트레스 이용자(명)"] = None
            _row["스트레스 이용비중(%)"] = None

        # 건강상담 — 해당 주차 일평균
        _hc_b = biz_filter_df(_hc_all, _biz, col="지자체")
        if not _hc_b.empty and "합계" in _hc_b.columns:
            _hc_b2 = _hc_b.copy()
            _hc_b2["_wk"] = _hc_b2["날짜"].astype(str).str.strip().map(_daymap)
            _hc_wk = _hc_b2[_hc_b2["_wk"] == _snap_week]
            _row["건강상담 건수(일평균)"] = round(_hc_wk["합계"].apply(safe_numeric).mean(), 1) if not _hc_wk.empty else None
        else:
            _row["건강상담 건수(일평균)"] = None

        # 걸음수 — 해당 주차 일평균
        _steps_b = biz_filter_df(_steps_all, _biz, col="agencyName")
        if not _steps_b.empty and "date" in _steps_b.columns:
            _steps_b2 = _steps_b.copy()
            _steps_b2["_dstr"] = pd.to_datetime(_steps_b2["date"], errors="coerce").dt.strftime("%Y-%m-%d")
            _steps_b2["_wk"] = _steps_b2["_dstr"].map(_daymap)
            _steps_wk = _steps_b2[_steps_b2["_wk"] == _snap_week]
            _row["걸음수 참여자(명)"] = round(_steps_wk["memberCnt"].apply(safe_numeric).mean(), 1) if not _steps_wk.empty else None
        else:
            _row["걸음수 참여자(명)"] = None

        _rows.append(_row)

    biz_status_df = pd.DataFrame(_rows)

    def _fmt(v, suffix=""):
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return "-"
        if isinstance(v, float) and v == int(v):
            v = int(v)
        return f"{v:,}{suffix}"

    for _, r in biz_status_df.iterrows():
        _color = BUSINESS_TYPE_COLORS.get(r["사업구분"], "#666")
        st.markdown(f"""<div style="border-left:6px solid {_color};background:#fff;border-radius:10px;
            padding:14px 18px;margin-bottom:10px;box-shadow:0 1px 3px rgba(0,0,0,0.06)">
            <div style="font-size:16px;font-weight:800;color:{_color};margin-bottom:8px">{r['사업구분']}
                <span style="font-size:12px;font-weight:500;color:#6b7488"> · 지자체 {r['지자체수']}개 · 이용자(협약) {r['이용자(협약)']:,}명</span>
            </div>
            <div style="display:flex;flex-wrap:wrap;gap:22px;font-size:13px;color:#33394a">
                <div>가입률 <b>{_fmt(r['가입률(%)'], '%')}</b> <span style="color:#8a94a8">({_fmt(r['가입수(명)'], '명')})</span></div>
                <div>안부확인율 <b>{_fmt(r['안부확인율(%)'], '%')}</b></div>
                <div>안부체크율 <b>{_fmt(r['안부체크율(%)'], '%')}</b></div>
                <div>심혈관 이용자 <b>{_fmt(r['심혈관 이용자(명)'], '명')}</b> <span style="color:#8a94a8">({_fmt(r['심혈관 이용비중(%)'], '%')})</span></div>
                <div>스트레스 이용자 <b>{_fmt(r['스트레스 이용자(명)'], '명')}</b> <span style="color:#8a94a8">({_fmt(r['스트레스 이용비중(%)'], '%')})</span></div>
                <div>건강상담(일평균) <b>{_fmt(r['건강상담 건수(일평균)'], '건')}</b></div>
                <div>걸음수 참여자 <b>{_fmt(r['걸음수 참여자(명)'], '명')}</b></div>
            </div>
            <div style="margin-top:8px;font-size:12px;color:#8a94a8">{r['지자체명목록']}</div>
        </div>""", unsafe_allow_html=True)

    st.markdown("")
    with st.expander("📋 표로 보기 (상세 데이터)", expanded=False):
        st.dataframe(biz_status_df, use_container_width=True)

    st.caption("⚠ 복약관리·생활상담은 지자체별 원본 데이터가 없어 사업구분별 집계가 불가능해 표에서 제외했습니다. "
               "맞고(와플랫+게스트)·맞고(게스트)도 동일한 이유로 제외됩니다.")

    st.markdown("")
    st.markdown('<div class="section-header">📈 사업구분별 주차별 추이</div>', unsafe_allow_html=True)
    _bt_start, _bt_end = page_week_range_selector("bizstatus_trend", weeks)
    st.divider()

    def _plot_biz_weekly(long_df, title, y_title, is_pct=True):
        if long_df.empty:
            st.info(f"{title} — 주차별 데이터가 없습니다.")
            return
        fig = go.Figure()
        for _biz in _biz_list:
            _bdf = long_df[long_df["사업구분"] == _biz]
            if _bdf.empty:
                continue
            fig.add_trace(go.Scatter(
                x=_bdf["주차"], y=_bdf["값"], mode="lines+markers", name=_biz,
                line=dict(color=BUSINESS_TYPE_COLORS.get(_biz, "#666"), width=2.5),
                marker=dict(size=6),
                hovertemplate=f"<b>%{{x}}</b><br>{_biz}: %{{y:.1f}}{'%' if is_pct else ''}<extra></extra>",
            ))
        fig.update_layout(
            title=title, height=420, hovermode="x unified",
            xaxis=dict(type="category", tickangle=-45),
            yaxis=dict(title=y_title),
            legend=dict(orientation="h", yanchor="top", y=-0.3, xanchor="left", x=0, font=dict(size=11)),
            margin=dict(t=40, b=100),
        )
        st.plotly_chart(fig, use_container_width=True)

    _tab_gr, _tab_cr, _tab_cc, _tab_cd, _tab_st, _tab_hc, _tab_step = st.tabs(
        ["가입률", "안부확인율", "안부체크율", "심혈관", "스트레스", "건강상담", "걸음수"]
    )

    with _tab_gr:
        _wr = data.get("weekly_registered_by_mun", pd.DataFrame())
        if not _wr.empty:
            _wr_f = filter_by_week_range(_wr, "주차", _bt_start, _bt_end, weeks)
            _wr_f = shorten_dates_in_df(_wr_f, "주차")
            _rows_gr = []
            for _biz in _biz_list:
                _contract_total = biz_filter_df(_reg_all, _biz)["협약인원"].apply(safe_numeric).sum()
                if _contract_total <= 0:
                    continue
                _b = biz_filter_df(_wr_f, _biz)
                if _b.empty:
                    continue
                for _wk in pd.unique(_b["주차"]):
                    _g = _b[_b["주차"] == _wk]
                    _rows_gr.append({"주차": _wk, "사업구분": _biz,
                                      "값": round(_g["가입완료"].apply(safe_numeric).sum() / _contract_total * 100, 1)})
            _plot_biz_weekly(pd.DataFrame(_rows_gr), "사업구분별 주차별 가입률 추이", "가입률 (%)")
            st.caption("※ 협약인원은 현재 시점 기준 고정값으로 계산했습니다 (주차별 협약인원 이력 데이터 없음).")
        else:
            st.info("주차별 가입 데이터가 없습니다.")

    with _tab_cr:
        _cw = data.get("checkin_mun_weekly", pd.DataFrame())
        if not _cw.empty and "분자" in _cw.columns and "분모" in _cw.columns:
            _cw2 = _cw.copy()
            # 시작일이 일별 날짜이므로(주차 시작일이 아님) daymap으로 주차 매핑
            _cw2["주차"] = _cw2["시작일"].astype(str).str.strip().map(_daymap)
            _cw2 = _cw2.dropna(subset=["주차"])
            _cw2 = filter_by_week_range(_cw2, "주차", _bt_start, _bt_end, weeks)
            _cw2 = shorten_dates_in_df(_cw2, "주차")
            _rows_cr = []
            for _biz in _biz_list:
                _b = biz_filter_df(_cw2, _biz)
                if _b.empty:
                    continue
                for _wk in pd.unique(_b["주차"]):
                    _g = _b[_b["주차"] == _wk]
                    _den = _g["분모"].apply(safe_numeric).sum()
                    _num = _g["분자"].apply(safe_numeric).sum()
                    if _den > 0:
                        _rows_cr.append({"주차": _wk, "사업구분": _biz, "값": round(_num / _den * 100, 1)})
            _plot_biz_weekly(pd.DataFrame(_rows_cr), "사업구분별 주차별 안부확인율 추이", "안부확인율 (%)")
        else:
            st.info("주차별 안부확인율 데이터가 없습니다.")

    with _tab_cc:
        _ccr = _checkin_rate_all.copy() if not _checkin_rate_all.empty else pd.DataFrame()
        if not _ccr.empty and "안부체크발송" in _ccr.columns and "안부체크응답" in _ccr.columns and "시작일" in _ccr.columns:
            # 시작일이 일별 날짜이므로(주차 시작일이 아님) daymap으로 주차 매핑
            _ccr["주차"] = _ccr["시작일"].astype(str).str.strip().map(_daymap)
            _ccr = _ccr.dropna(subset=["주차"])
            _ccr = filter_by_week_range(_ccr, "주차", _bt_start, _bt_end, weeks)
            _ccr = shorten_dates_in_df(_ccr, "주차")
            _rows_cc = []
            for _biz in _biz_list:
                _b = biz_filter_df(_ccr, _biz)
                if _b.empty:
                    continue
                for _wk in pd.unique(_b["주차"]):
                    _g = _b[_b["주차"] == _wk]
                    _send = _g["안부체크발송"].apply(safe_numeric).sum()
                    _resp = _g["안부체크응답"].apply(safe_numeric).sum()
                    # off대상자는 지자체별 고정값(일별로 반복 저장됨) — 지자체당 1회만 집계
                    _off = (_g.drop_duplicates(subset=["지자체명"])["off대상자"].apply(safe_numeric).sum()
                            if "off대상자" in _g.columns else 0)
                    _denom = _send - _off
                    if _denom > 0:
                        _rows_cc.append({"주차": _wk, "사업구분": _biz, "값": round(_resp / _denom * 100, 1)})
            _plot_biz_weekly(pd.DataFrame(_rows_cc), "사업구분별 주차별 안부체크율 추이", "안부체크율 (%)")
        else:
            st.info("주차별 안부체크율 데이터가 없습니다.")

    with _tab_cd:
        _cardio_full = sheets.get("심혈관이용자", pd.DataFrame())
        _c_wc = next((c for c in _cardio_full.columns if "주차" in str(c).replace("\n", "").strip()), None) if not _cardio_full.empty else None
        if _c_wc:
            _cf = filter_by_week_range(_cardio_full, _c_wc, _bt_start, _bt_end, weeks)
            _cf = shorten_dates_in_df(_cf, _c_wc)
            _rows_cd, _rows_cd_r = [], []
            for _biz in _biz_list:
                _cnt, _ = biz_agg_raw(_cf, _biz, _c_wc)
                if _cnt is None:
                    continue
                for _wk, _v in zip(_cf[_c_wc], _cnt):
                    _rows_cd.append({"주차": _wk, "사업구분": _biz, "값": round(float(_v), 1)})
                # 시트 내장 이용비중 대신 이용자수 ÷ 가입완료(registration)로 직접 계산
                _biz_completed = biz_filter_df(_reg_all, _biz)["가입완료"].apply(safe_numeric).sum()
                if _biz_completed > 0:
                    for _wk, _v in zip(_cf[_c_wc], _cnt):
                        _rows_cd_r.append({"주차": _wk, "사업구분": _biz, "값": round(float(_v) / _biz_completed * 100, 1)})
            _plot_biz_weekly(pd.DataFrame(_rows_cd), "사업구분별 주차별 심혈관 이용자수 추이", "이용자수 (명)", is_pct=False)
            _plot_biz_weekly(pd.DataFrame(_rows_cd_r), "사업구분별 주차별 심혈관 이용비중 추이", "이용비중 (%)")
            st.caption("※ 이용비중 = 심혈관 이용자수 ÷ 가입완료 인원(registration, 현재 시점 고정값) × 100")
        else:
            st.info("심혈관 주차별 데이터가 없습니다.")

    with _tab_st:
        _stress_full = sheets.get("스트레스이용자", pd.DataFrame())
        _s_wc = next((c for c in _stress_full.columns if "주차" in str(c).replace("\n", "").strip()), None) if not _stress_full.empty else None
        if _s_wc:
            _sf = filter_by_week_range(_stress_full, _s_wc, _bt_start, _bt_end, weeks)
            _sf = shorten_dates_in_df(_sf, _s_wc)
            _rows_st, _rows_st_r = [], []
            for _biz in _biz_list:
                _cnt, _ = biz_agg_raw(_sf, _biz, _s_wc)
                if _cnt is None:
                    continue
                for _wk, _v in zip(_sf[_s_wc], _cnt):
                    _rows_st.append({"주차": _wk, "사업구분": _biz, "값": round(float(_v), 1)})
                _biz_completed = biz_filter_df(_reg_all, _biz)["가입완료"].apply(safe_numeric).sum()
                if _biz_completed > 0:
                    for _wk, _v in zip(_sf[_s_wc], _cnt):
                        _rows_st_r.append({"주차": _wk, "사업구분": _biz, "값": round(float(_v) / _biz_completed * 100, 1)})
            _plot_biz_weekly(pd.DataFrame(_rows_st), "사업구분별 주차별 스트레스 이용자수 추이", "이용자수 (명)", is_pct=False)
            _plot_biz_weekly(pd.DataFrame(_rows_st_r), "사업구분별 주차별 스트레스 이용비중 추이", "이용비중 (%)")
            st.caption("※ 이용비중 = 스트레스 이용자수 ÷ 가입완료 인원(registration, 현재 시점 고정값) × 100")
        else:
            st.info("스트레스 주차별 데이터가 없습니다.")

    with _tab_hc:
        if not _hc_all.empty and "합계" in _hc_all.columns and "날짜" in _hc_all.columns:
            _hc2 = _hc_all.copy()
            _hc2["주차"] = _hc2["날짜"].astype(str).str.strip().map(_daymap)
            _hc2 = _hc2.dropna(subset=["주차"])
            _hc2 = filter_by_week_range(_hc2, "주차", _bt_start, _bt_end, weeks)
            _hc2 = shorten_dates_in_df(_hc2, "주차")
            _rows_hc = []
            for _biz in _biz_list:
                _b = biz_filter_df(_hc2, _biz, col="지자체")
                if _b.empty:
                    continue
                for _wk in pd.unique(_b["주차"]):
                    _g = _b[_b["주차"] == _wk]
                    _rows_hc.append({"주차": _wk, "사업구분": _biz, "값": round(_g["합계"].apply(safe_numeric).mean(), 1)})
            _plot_biz_weekly(pd.DataFrame(_rows_hc), "사업구분별 주차별 건강상담 건수 추이 (일평균)", "건강상담 건수 (일평균)", is_pct=False)
            st.caption("※ 건강상담은 일별 데이터라 각 주차 내 일평균값으로 계산했습니다.")
        else:
            st.info("건강상담 주차별 데이터가 없습니다.")

    with _tab_step:
        if not _steps_all.empty and "date" in _steps_all.columns and "memberCnt" in _steps_all.columns:
            _st2 = _steps_all.copy()
            _st2["_dstr"] = pd.to_datetime(_st2["date"], errors="coerce").dt.strftime("%Y-%m-%d")
            _st2["주차"] = _st2["_dstr"].map(_daymap)
            _st2 = _st2.dropna(subset=["주차"])
            _st2 = filter_by_week_range(_st2, "주차", _bt_start, _bt_end, weeks)
            _st2 = shorten_dates_in_df(_st2, "주차")
            _rows_step = []
            for _biz in _biz_list:
                _b = biz_filter_df(_st2, _biz, col="agencyName")
                if _b.empty:
                    continue
                for _wk in pd.unique(_b["주차"]):
                    _g = _b[_b["주차"] == _wk]
                    _rows_step.append({"주차": _wk, "사업구분": _biz, "값": round(_g["memberCnt"].apply(safe_numeric).mean(), 1)})
            _plot_biz_weekly(pd.DataFrame(_rows_step), "사업구분별 주차별 걸음수 참여자 추이 (일평균)", "참여자수 (일평균, 명)", is_pct=False)
            st.caption("※ 걸음수는 일별 데이터라 각 주차 내 일평균값으로 계산했습니다.")
        else:
            st.info("걸음수 주차별 데이터가 없습니다.")


# ============================================================
# 🛡 세이프·베이직 현황
# ============================================================
elif page == "🛡 세이프·베이직 현황":
    st.markdown('<div class="section-header">🛡 세이프·베이직 현황</div>', unsafe_allow_html=True)
    st.caption("세이프(관제)·베이직 구분별 핵심 지표와 KT 관제 현황을 확인합니다.")
    st.divider()

    _TIER_COLORS = {"세이프": "#1565C0", "세이프 플러스": "#6A1B9A", "베이직": "#757575"}
    _reg_tier_all = data.get("registration", pd.DataFrame())
    _tier_list = []
    if not _reg_tier_all.empty and "구분" in _reg_tier_all.columns:
        _active_reg = _reg_tier_all[_reg_tier_all["협약인원"].apply(safe_numeric) > 0]
        _tier_counts = _active_reg["구분"].astype(str).str.strip().value_counts()
        _tier_list = [t for t in ["세이프", "세이프 플러스", "베이직"] if t in _tier_counts.index]

    def _tier_filter_df(df, tier, col="지자체명"):
        if df.empty or _reg_tier_all.empty or "구분" not in _reg_tier_all.columns or col not in df.columns:
            return df
        munis = set(_reg_tier_all.loc[_reg_tier_all["구분"].astype(str).str.strip() == tier, "지자체명"]
                    .astype(str).str.strip())
        munis_n = {m.replace(" ", "") for m in munis}
        def _match(name):
            n = str(name).replace(" ", "")
            return any(m == n or m in n or n in m for m in munis_n)
        return df[df[col].apply(_match)]

    _cr_direct_tier = data.get("checkin_mun_rate_direct", pd.DataFrame())
    _rows_tier = []
    for _tier in _tier_list:
        _row = {"구분": _tier}
        _rb = _tier_filter_df(_reg_tier_all, _tier)
        _contract = _rb["협약인원"].apply(safe_numeric).sum() if not _rb.empty else 0
        _registered = _rb["가입완료"].apply(safe_numeric).sum() if not _rb.empty and "가입완료" in _rb.columns else 0
        _row["지자체수"] = _rb.loc[_rb["협약인원"].apply(safe_numeric) > 0, "지자체명"].nunique() if not _rb.empty else 0
        _row["이용자(협약)"] = int(_contract)
        _row["가입률(%)"] = round(_registered / _contract * 100, 1) if _contract > 0 else 0.0

        _crb = _tier_filter_df(_cr_direct_tier, _tier)
        if not _crb.empty and "분자" in _crb.columns and "분모" in _crb.columns:
            _num = _crb["분자"].apply(safe_numeric).sum()
            _den = _crb["분모"].apply(safe_numeric).sum()
            _row["안부확인율(%)"] = round(_num / _den * 100, 1) if _den > 0 else 0.0
        else:
            _row["안부확인율(%)"] = None
        _rows_tier.append(_row)

    tier_status_df = pd.DataFrame(_rows_tier)

    def _fmt_tier(v, suffix=""):
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return "-"
        if isinstance(v, float) and v == int(v):
            v = int(v)
        return f"{v:,}{suffix}"

    for _, r in tier_status_df.iterrows():
        _color = _TIER_COLORS.get(r["구분"], "#666")
        st.markdown(f"""<div style="border-left:6px solid {_color};background:#fff;border-radius:10px;
            padding:14px 18px;margin-bottom:10px;box-shadow:0 1px 3px rgba(0,0,0,0.06)">
            <div style="font-size:16px;font-weight:800;color:{_color};margin-bottom:8px">{r['구분']}
                <span style="font-size:12px;font-weight:500;color:#6b7488"> · 지자체 {r['지자체수']}개 · 이용자(협약) {r['이용자(협약)']:,}명</span>
            </div>
            <div style="display:flex;flex-wrap:wrap;gap:22px;font-size:13px;color:#33394a">
                <div>가입률 <b>{r['가입률(%)']}%</b></div>
                <div>안부확인율 <b>{_fmt_tier(r['안부확인율(%)'], '%')}</b></div>
            </div>
        </div>""", unsafe_allow_html=True)

    st.markdown("")
    st.markdown('<div class="section-header">📡 KT 관제 현황 (세이프 전용)</div>', unsafe_allow_html=True)
    st.caption("KT 관제·출동은 세이프 계약 지자체에만 적용되는 서비스라 전체(세이프) 기준으로 표시합니다.")

    _safe_raw_tier = sheets.get("안부체크횟수", pd.DataFrame())
    if not _safe_raw_tier.empty:
        _sft = _safe_raw_tier.copy()
        _sf_date_col_t = None
        for _c in _sft.columns:
            if "시작일" in str(_c).replace("\n", "").strip():
                _sf_date_col_t = _c
                break
        if _sf_date_col_t is None:
            _sf_date_col_t = _sft.columns[0]

        _all_safe_dates_t = sorted([str(d).strip() for d in _sft[_sf_date_col_t].dropna().unique()
                                     if str(d).strip() and str(d).strip() != "nan"])

        _safe_default_t = "2026-03-01"
        _safe_idx_t = next((i for i, d in enumerate(_all_safe_dates_t) if d >= _safe_default_t), 0)
        with st.expander("📅 기간 설정 (펼쳐서 변경)", expanded=False):
            _stc1, _stc2 = st.columns(2)
            with _stc1:
                _safe_start_t = st.selectbox("시작일", _all_safe_dates_t, index=_safe_idx_t, key="tier_kt_start")
            with _stc2:
                _ss_idx_t = _all_safe_dates_t.index(_safe_start_t) if _safe_start_t in _all_safe_dates_t else 0
                _safe_end_opts_t = _all_safe_dates_t[_ss_idx_t:]
                _safe_end_t = st.selectbox("종료일", _safe_end_opts_t, index=len(_safe_end_opts_t) - 1, key="tier_kt_end")
            st.caption(f"기간: {_safe_start_t} ~ {_safe_end_t}")

        _df_safe_t = _sft[(_sft[_sf_date_col_t].astype(str) >= _safe_start_t) &
                          (_sft[_sf_date_col_t].astype(str) <= _safe_end_t)].copy()
        _df_safe_t = shorten_dates_in_df(_df_safe_t, _sf_date_col_t)

        _kt_total_col_t = _kt_send_col_t = _kt_rate_col_t = _kt_mgmt_rate_col_t = _kt_disp_rate_col_t = None
        for _c in _df_safe_t.columns:
            _cl_raw_t = str(_c).replace("\n", "").strip()
            _cl_low_t = _cl_raw_t.replace(" ", "").lower()
            if _cl_raw_t == "KT 관제 수":
                _kt_total_col_t = _c
            elif _cl_raw_t == "전체 발송수":
                _kt_send_col_t = _c
            elif _cl_raw_t == "KT 관제 대응률":
                _kt_rate_col_t = _c
            if "kt관제율" in _cl_low_t:
                _kt_mgmt_rate_col_t = _c
            elif "kt출동율" in _cl_low_t or "kt출동률" in _cl_low_t:
                _kt_disp_rate_col_t = _c

        if _kt_mgmt_rate_col_t or _kt_disp_rate_col_t:
            _fig_kt_rate_t = go.Figure()
            if _kt_mgmt_rate_col_t:
                _df_safe_t[_kt_mgmt_rate_col_t] = _df_safe_t[_kt_mgmt_rate_col_t].apply(safe_numeric)
                _kt_mgmt_t = _df_safe_t[_df_safe_t[_kt_mgmt_rate_col_t] > 0]
                _fig_kt_rate_t.add_trace(go.Scatter(
                    x=_kt_mgmt_t[_sf_date_col_t], y=_kt_mgmt_t[_kt_mgmt_rate_col_t],
                    name="KT 관제율", mode="lines+markers",
                    line=dict(color="#1565C0", width=2.5), marker=dict(size=7),
                    hovertemplate="<b>%{x}</b><br>KT 관제율: <b>%{y:.1f}%</b><extra></extra>"
                ))
            if _kt_disp_rate_col_t:
                _df_safe_t[_kt_disp_rate_col_t] = _df_safe_t[_kt_disp_rate_col_t].apply(safe_numeric)
                _kt_disp_t = _df_safe_t[_df_safe_t[_kt_disp_rate_col_t] > 0]
                if not _kt_disp_t.empty:
                    _fig_kt_rate_t.add_trace(go.Scatter(
                        x=_kt_disp_t[_sf_date_col_t], y=_kt_disp_t[_kt_disp_rate_col_t],
                        name="KT 출동율", mode="lines+markers",
                        line=dict(color="#E65100", width=2.5, dash="dot"), marker=dict(size=7, symbol="diamond"),
                        hovertemplate="<b>%{x}</b><br>KT 출동율: <b>%{y:.1f}%</b><extra></extra>"
                    ))
            _fig_kt_rate_t.update_layout(
                title="KT 관제율 · 출동율 추이", height=380,
                xaxis=dict(type="category", title=""), yaxis=dict(title="%", ticksuffix="%"),
                hovermode="x unified", margin=dict(t=40, b=60, l=50, r=20),
                legend=dict(orientation="h", yanchor="top", y=-0.18, xanchor="center", x=0.5),
            )
            st.plotly_chart(_fig_kt_rate_t, use_container_width=True)
        else:
            st.info("KT 관제율·출동율 컬럼을 찾을 수 없습니다.")

        if _kt_total_col_t and _kt_send_col_t:
            _df_safe_t[_kt_total_col_t] = _df_safe_t[_kt_total_col_t].apply(safe_numeric)
            _df_safe_t[_kt_send_col_t] = _df_safe_t[_kt_send_col_t].apply(safe_numeric)
            _fig3_t = make_subplots(specs=[[{"secondary_y": True}]])
            _fig3_t.add_trace(go.Bar(x=_df_safe_t[_sf_date_col_t], y=_df_safe_t[_kt_send_col_t], name="전체 발송수",
                                     marker_color="#B0BEC5", opacity=0.6,
                                     hovertemplate="%{y:,.0f}건<extra>전체 발송수</extra>"), secondary_y=False)
            _fig3_t.add_trace(go.Bar(x=_df_safe_t[_sf_date_col_t], y=_df_safe_t[_kt_total_col_t], name="KT 관제 수",
                                     marker_color="#1565C0",
                                     hovertemplate="%{y:,.0f}건<extra>KT 관제 수</extra>"), secondary_y=False)
            if _kt_rate_col_t:
                _df_safe_t[_kt_rate_col_t] = _df_safe_t[_kt_rate_col_t].apply(safe_numeric)
                _fig3_t.add_trace(go.Scatter(x=_df_safe_t[_sf_date_col_t], y=_df_safe_t[_kt_rate_col_t], name="KT 관제 대응률",
                                             mode="lines+markers", line=dict(color="#D32F2F", width=2),
                                             hovertemplate="%{y:.1f}%<extra>KT 관제 대응률</extra>"), secondary_y=True)
            _fig3_t.update_layout(title="세이프 - KT 관제 현황", height=400,
                                  xaxis=dict(type="category"), hovermode="x unified", barmode="group",
                                  margin=dict(t=40, b=60, l=40, r=40),
                                  legend=dict(orientation="h", yanchor="top", y=-0.15, xanchor="center", x=0.5))
            _fig3_t.update_yaxes(title_text="건수", secondary_y=False)
            _fig3_t.update_yaxes(title_text="%", secondary_y=True)
            st.plotly_chart(_fig3_t, use_container_width=True)
    else:
        st.info("KT 관제 데이터가 없습니다.")


# ============================================================
# 👥 회원가입 & 이탈
# ============================================================
elif page == "👥 1.회원가입 & 이탈":
    st.markdown('<div class="section-header">👥 회원가입 및 앱 삭제자 현황</div>', unsafe_allow_html=True)
    p_start, p_end = page_week_range_selector("member", weeks)
    selected_biz = biz_selector("회원가입이탈")
    st.divider()

    # 지자체별 회원가입 현황
    reg = biz_filter_df(data.get("registration", pd.DataFrame()).copy(), selected_biz)
    wu = data.get("weekly_users", pd.DataFrame())

    if not reg.empty and "지자체명" in reg.columns:
        # 사업구분 선택 시 집계 시트(weekly_users)는 분리 불가 → reg 스냅샷 기준
        total_contract = 0
        total_registered = 0
        if selected_biz == "전체":
            wu_kpi = data.get("weekly_users", pd.DataFrame())
            if not wu_kpi.empty and "주차" in wu_kpi.columns:
                _sel = wu_kpi[wu_kpi["주차"].astype(str).str.strip() == selected_week]
                if _sel.empty:
                    _sel = wu_kpi.iloc[[-1]]
                if not _sel.empty:
                    _r = _sel.iloc[0]
                    if "대상자수" in wu_kpi.columns:
                        _v = safe_numeric(_r.get("대상자수", 0))
                        if _v > 0: total_contract = int(_v)
                    if "가입완료합계" in wu_kpi.columns:
                        _v = safe_numeric(_r.get("가입완료합계", 0))
                        if _v > 0: total_registered = int(_v)
        # fallback: 이용자현황 스냅샷 (사업구분 필터 적용)
        if total_contract == 0:
            total_contract = int(reg["협약인원"].apply(safe_numeric).sum()) if "협약인원" in reg.columns else 0
        if total_registered == 0:
            total_registered = int(reg["가입완료"].apply(safe_numeric).sum()) if "가입완료" in reg.columns else 0
        total_rate = round(total_registered / total_contract * 100, 1) if total_contract > 0 else 0
        total_incomplete = total_contract - total_registered

        kcols = st.columns(4)
        with kcols[0]:
            st.markdown(f'<div class="metric-card"><h3>총 협약인원</h3><h1>{total_contract:,}명</h1></div>', unsafe_allow_html=True)
        with kcols[1]:
            st.markdown(f'<div class="metric-card-green"><h3>가입완료</h3><h1>{total_registered:,}명</h1></div>', unsafe_allow_html=True)
        with kcols[2]:
            st.markdown(f'<div class="metric-card-red"><h3>미완료</h3><h1>{total_incomplete:,}명</h1></div>', unsafe_allow_html=True)
        with kcols[3]:
            st.markdown(f'<div class="metric-card-orange"><h3>전체 가입률</h3><h1>{total_rate}%</h1></div>', unsafe_allow_html=True)

        st.markdown("")

        # 주차별 대상자 수 대비 가입완료 비중 추이 (전체일 때만 표시 — 지자체별 분리 데이터 없음)
        wu = data.get("weekly_users", pd.DataFrame())
        if selected_biz == "전체" and not wu.empty and "주차" in wu.columns:
            wu_chart = wu.copy()
            # 기간 필터 적용
            if p_start:
                wu_chart = filter_by_week_range(wu_chart, "주차", p_start, p_end, weeks)
            else:
                wu_chart = wu_chart[wu_chart["주차"].astype(str).str.strip() >= "25-52"]
            # 필요한 컬럼 찾기 및 숫자 변환
            target_col = None
            reg_col = None
            for c in wu_chart.columns:
                cl = str(c).replace("\n", "").strip()
                if cl == "대상자수" or ("대상자" in cl and "수" in cl):
                    target_col = c
                elif c == "가입완료합계" or ("회원가입" in cl and "완료" in cl and "주간" not in cl and "비중" not in cl and "율" not in cl):
                    reg_col = c

            if target_col and reg_col:
                wu_chart[target_col] = wu_chart[target_col].apply(safe_numeric)
                wu_chart[reg_col] = wu_chart[reg_col].apply(safe_numeric)
                # 전체가입률: 가입완료/대상자수 직접 계산 (신규 지자체 포함 정확한 분모 반영)
                if False:
                    pass
                elif "대상자수" in wu_chart.columns:
                    wu_chart["대상자수"] = wu_chart["대상자수"].apply(safe_numeric)
                    wu_chart["_rate"] = (wu_chart[reg_col] / wu_chart["대상자수"].replace(0, float("nan")) * 100).round(1).fillna(0)
                else:
                    wu_chart["_rate"] = (wu_chart[reg_col] / wu_chart[target_col].replace(0, float("nan")) * 100).round(1).fillna(0)

                fig = make_subplots(specs=[[{"secondary_y": True}]])
                fig.add_trace(
                    go.Bar(x=wu_chart["주차"], y=wu_chart[target_col], name="대상자 수",
                           marker_color="#B0BEC5",
                           hovertemplate="<b>%{x}</b><br>대상자: %{y:,}명<extra></extra>"),
                    secondary_y=False,
                )
                fig.add_trace(
                    go.Bar(x=wu_chart["주차"], y=wu_chart[reg_col], name="가입완료",
                           marker_color="#00C853",
                           hovertemplate="<b>%{x}</b><br>가입완료: %{y:,}명<extra></extra>"),
                    secondary_y=False,
                )
                fig.add_trace(
                    go.Scatter(x=wu_chart["주차"], y=wu_chart["_rate"], name="가입률(%)",
                               mode="lines+markers", line=dict(color="#FF6F00", width=3),
                               marker=dict(size=5),
                               hovertemplate="<b>%{x}</b><br>가입률: %{y:.1f}%<extra></extra>"),
                    secondary_y=True,
                )

                fig.update_layout(
                    title="주차별 대상자 수 대비 회원가입 완료 비중",
                    height=420, margin=dict(t=40, b=30),
                    hovermode="x unified", barmode="group",
                    legend=dict(orientation="h", yanchor="top", y=-0.15, xanchor="center", x=0.5),
                    xaxis=dict(type="category"),  # 주차를 카테고리로 (날짜 자동해석 방지)
                )
                fig.update_yaxes(title_text="명", secondary_y=False)
                fig.update_yaxes(title_text="가입률 (%)", secondary_y=True, range=[0, 110])
                st.plotly_chart(fig, use_container_width=True)

        elif selected_biz != "전체" and not wu.empty and "주차" in wu.columns:
            wu_chart = wu.copy()
            # 기간 필터 적용
            if p_start:
                wu_chart = filter_by_week_range(wu_chart, "주차", p_start, p_end, weeks)
            else:
                wu_chart = wu_chart[wu_chart["주차"].astype(str).str.strip() >= "25-52"]
            wu_chart = wu_chart.dropna(subset=["주차"])

            mun_reg_df = data.get("weekly_registered_by_mun", pd.DataFrame())
            reg_snapshot = data.get("registration", pd.DataFrame())
            mun_biz = biz_filter_df(mun_reg_df.copy(), selected_biz, col="지자체명") if not mun_reg_df.empty else pd.DataFrame()

            if not mun_biz.empty and "주차" in mun_biz.columns:
                weeks_set = set(wu_chart["주차"].astype(str).str.strip().tolist())
                mun_biz = mun_biz[mun_biz["주차"].astype(str).str.strip().isin(weeks_set)]
                agg_w = mun_biz.groupby("주차")["가입완료"].sum().reset_index()
                agg_w["주차"] = agg_w["주차"].astype(str).str.strip()

                # 대상자 수: registration 스냅샷에서 biz 소속 지자체 협약인원 합산 (주차별 고정값)
                if not reg_snapshot.empty and "협약인원" in reg_snapshot.columns:
                    reg_biz = biz_filter_df(reg_snapshot.copy(), selected_biz)
                    biz_target = int(reg_biz["협약인원"].apply(safe_numeric).sum())
                else:
                    biz_target = 0

                agg_w["대상자수"] = biz_target
                agg_w["_rate"] = (
                    agg_w["가입완료"] / biz_target * 100
                ).round(1) if biz_target > 0 else 0.0

                fig_biz = make_subplots(specs=[[{"secondary_y": True}]])
                fig_biz.add_trace(
                    go.Bar(x=agg_w["주차"], y=agg_w["대상자수"], name="대상자 수",
                           marker_color="#B0BEC5",
                           hovertemplate="<b>%{x}</b><br>대상자: %{y:,}명<extra></extra>"),
                    secondary_y=False,
                )
                fig_biz.add_trace(
                    go.Bar(x=agg_w["주차"], y=agg_w["가입완료"], name="가입완료",
                           marker_color="#00C853",
                           hovertemplate="<b>%{x}</b><br>가입완료: %{y:,}명<extra></extra>"),
                    secondary_y=False,
                )
                fig_biz.add_trace(
                    go.Scatter(x=agg_w["주차"], y=agg_w["_rate"], name="가입률(%)",
                               mode="lines+markers", line=dict(color="#FF6F00", width=3),
                               marker=dict(size=5),
                               hovertemplate="<b>%{x}</b><br>가입률: %{y:.1f}%<extra></extra>"),
                    secondary_y=True,
                )
                fig_biz.update_layout(
                    title=f"주차별 {selected_biz} 대상자 수 대비 회원가입 완료 비중",
                    height=420, margin=dict(t=40, b=30),
                    hovermode="x unified", barmode="group",
                    legend=dict(orientation="h", yanchor="top", y=-0.15, xanchor="center", x=0.5),
                    xaxis=dict(type="category"),
                )
                fig_biz.update_yaxes(title_text="명", secondary_y=False)
                fig_biz.update_yaxes(title_text="가입률 (%)", secondary_y=True, range=[0, 110])
                st.plotly_chart(fig_biz, use_container_width=True)
            else:
                st.info(f"{selected_biz} 해당 지자체의 주차별 가입완료 데이터가 없습니다.")

        tab1, tab2 = st.tabs(["지자체별 비중", "지자체별 가입률"])

        with tab1:
            # 지자체별 전체 회원 중 비중 (%)
            if "가입완료" in reg.columns:
                reg["전체비중"] = (reg["가입완료"] / total_registered * 100).round(1) if total_registered > 0 else 0
                reg["권역"] = reg["지자체명"].map(REGION_MAP).fillna("기타")

                reg_pie = reg.copy()
                reg_pie["지자체명"] = reg_pie["지자체명"].apply(_mun_label)
                fig = px.pie(reg_pie, values="가입완료", names="지자체명",
                             title="지자체별 회원 비중",
                             color_discrete_sequence=px.colors.qualitative.Set3)
                fig.update_traces(
                    textposition="inside",
                    textinfo="label+percent",
                    hovertemplate="<b>%{label}</b><br>가입완료: %{value:,}명<br>비중: %{percent}<extra></extra>"
                )
                fig.update_layout(height=450, margin=dict(t=40, b=10))
                st.plotly_chart(fig, use_container_width=True)

        with tab2:
            # 지자체별 가입 완료율 바 차트
            if "완료율" in reg.columns:
                reg["권역"] = reg["지자체명"].map(REGION_MAP).fillna("기타")
                plot_municipality_bar(reg, "완료율", "지자체별 회원가입 완료율 (협약 대비 %)")

    else:
        st.info("이용자 현황 데이터가 없습니다.")

    # 주차별 앱 삭제의심자 + 전체삭제비중 추이
    deletion_raw = sheets.get("심혈관현황", pd.DataFrame())  # gid=981210016 (앱삭제 데이터 포함)
    if not deletion_raw.empty:
        st.markdown('<div class="section-header">앱 삭제 현황</div>', unsafe_allow_html=True)

        del_trend = deletion_raw.copy()
        week_col_d = None
        delete_col = None
        ratio_col = None
        for c in del_trend.columns:
            cl = str(c).replace("\n", "").strip()
            if "주차" in cl:
                week_col_d = c
            elif "앱" in cl and "삭제" in cl and "의심" in cl:
                delete_col = c
            elif "전체삭제비중" in cl or "전체" in cl and "삭제" in cl and "비중" in cl:
                ratio_col = c

        if week_col_d and (delete_col or ratio_col):
            # 기간 필터 적용
            if p_start:
                del_trend = filter_by_week_range(del_trend, week_col_d, p_start, p_end, weeks)
            if delete_col:
                del_trend[delete_col] = del_trend[delete_col].apply(safe_numeric)
            if ratio_col:
                del_trend[ratio_col] = del_trend[ratio_col].apply(safe_numeric)

            del_chart = shorten_dates_in_df(del_trend, week_col_d)

            fig = make_subplots(specs=[[{"secondary_y": True}]])
            if delete_col:
                fig.add_trace(go.Bar(
                    x=del_chart[week_col_d], y=del_chart[delete_col],
                    name="앱 삭제의심자", marker_color="#EF5350",
                    hovertemplate="%{y:,}명<extra>삭제의심자</extra>"
                ), secondary_y=False)
            if ratio_col:
                fig.add_trace(go.Scatter(
                    x=del_chart[week_col_d], y=del_chart[ratio_col],
                    name="전체삭제비중(%)", mode="lines+markers",
                    line=dict(color="#D32F2F", width=2),
                    hovertemplate="%{y:.1f}%<extra>삭제비중</extra>"
                ), secondary_y=True)
            fig.update_layout(
                title="주차별 앱 삭제의심자 및 전체삭제비중",
                height=380, hovermode="x unified",
                xaxis=dict(type="category"),
                legend=LEGEND_BELOW, margin=dict(t=40, b=70),
            )
            fig.update_yaxes(title_text="삭제의심자 (명)", secondary_y=False)
            fig.update_yaxes(title_text="전체삭제비중 (%)", secondary_y=True, range=[0, max(20, del_chart[ratio_col].max() * 1.3) if ratio_col else 20])
            st.plotly_chart(fig, use_container_width=True)

    # 지자체별 앱삭제율
    del_df = data.get("app_deletion", pd.DataFrame())
    if not del_df.empty and selected_week:
        st.markdown('<div class="section-header">지자체별 앱 삭제율</div>', unsafe_allow_html=True)
        week_del = biz_filter_df(del_df[del_df["주차"].astype(str).str.strip() == selected_week].copy(), selected_biz)
        if not week_del.empty:
            # 앱삭제율이 0이고 다른 지표도 없는 지자체 제외 (계약 종료)
            # 히트맵에 있는 활성 지자체만 표시
            active_list = get_active_agencies()
            heatmap_muns = set()
            hm_df = cached_heatmap(data, selected_week)
            if not hm_df.empty:
                heatmap_muns = set(hm_df["지자체명"].tolist())
            if heatmap_muns:
                week_del = week_del[week_del["지자체명"].isin(heatmap_muns)]
            week_del["권역"] = week_del["지자체명"].map(REGION_MAP).fillna("기타")
            if not week_del.empty:
                plot_municipality_bar(week_del, "앱삭제율", f"{selected_week} 주차 지자체별 앱삭제율 (%)")


# ============================================================
# 🖐 안부확인
# ============================================================
elif page == "🖐 2.안부확인":
    st.markdown('<div class="section-header">🖐 안부확인 현황</div>', unsafe_allow_html=True)
    selected_biz = biz_selector("안부확인")
    st.divider()

    # DB에서 안부확인 raw 데이터 조회
    safety_db = data.get("dashboard_data", {}).get("db_safety_check", pd.DataFrame()) if isinstance(data.get("dashboard_data"), dict) else pd.DataFrame()
    if safety_db.empty:
        try:
            safety_db = get_db_data("raw_safety_check")
        except:
            safety_db = pd.DataFrame()

    # DB 데이터가 있으면 일별 지표 계산
    if not safety_db.empty and "date" in safety_db.columns:
        safety_db = safety_db.copy()

        # 📅 날짜 기간 선택기
        all_dates = sorted(safety_db["date"].unique())
        # 기본 시작: 2025-12-22 (25-52주차 시작일 근사) 또는 최근 60일
        default_start_date = "2026-01-01"
        if default_start_date in all_dates:
            default_start_idx = all_dates.index(default_start_date)
        else:
            # 가장 가까운 날짜 찾기
            default_start_idx = 0
            for i, d in enumerate(all_dates):
                if d >= default_start_date:
                    default_start_idx = i
                    break

        with st.expander("📅 기간 설정 (펼쳐서 변경)", expanded=False):
            dc1, dc2 = st.columns(2)
            with dc1:
                date_start = st.selectbox("시작일", all_dates, index=default_start_idx, key="safety_date_start")
            with dc2:
                start_idx = all_dates.index(date_start) if date_start in all_dates else 0
                end_options = all_dates[start_idx:]
                date_end = st.selectbox("종료일", end_options, index=len(end_options)-1, key="safety_date_end")
            st.caption(f"선택 기간: {date_start} ~ {date_end}")

        # 선택 기간으로 필터링
        safety_db = safety_db[(safety_db["date"] >= date_start) & (safety_db["date"] <= date_end)]

        # 사업구분 필터: agency_name 기준으로 소속 지자체만 유지
        if selected_biz != "전체" and "agency_name" in safety_db.columns:
            safety_db = biz_filter_df(safety_db, selected_biz, col="agency_name")

        # 일별 전체 합산
        agg_cols = {
            "alarm_send_count": "sum", "confirm_count": "sum",
            "target_user_count": "sum", "complete_user_count": "sum",
            "impossible_user_count": "sum", "detect_motion_count": "sum",
            "ai_care_generate_count": "sum", "ai_care_response_count": "sum",
            "call_generate_count": "sum", "call_response_count": "sum",
            "uncheck_48hr_user_count": "sum", "uncheck_48hr_target_count": "sum",
        }
        valid_agg = {k: v for k, v in agg_cols.items() if k in safety_db.columns}
        daily = safety_db.groupby("date").agg(valid_agg).reset_index().sort_values("date")

        # 주차별 집계 (사업구분 필터 반영 — biz 선택 시 해당 지자체만 합산됨)
        _sf_weekly = safety_db.copy()
        _sf_weekly["_week"] = _sf_weekly["date"].apply(date_to_week_label)
        weekly = _sf_weekly.groupby("_week").agg({k: v for k, v in valid_agg.items()}).reset_index().sort_values("_week")
        weekly.rename(columns={"_week": "week"}, inplace=True)
        _t_w = weekly["target_user_count"].replace(0, float("nan")) if "target_user_count" in weekly.columns else None
        _s_w = weekly["alarm_send_count"].replace(0, float("nan")) if "alarm_send_count" in weekly.columns else None
        if "confirm_count" in weekly.columns and _s_w is not None:
            weekly["안부체크응답률"] = (weekly["confirm_count"] / _s_w * 100).round(1).fillna(0)
        if "impossible_user_count" in weekly.columns and _t_w is not None:
            weekly["안부미확인률"] = (weekly["impossible_user_count"] / _t_w * 100).round(1).fillna(0)
        if "uncheck_48hr_user_count" in weekly.columns and _t_w is not None:
            weekly["48시간미확인률"] = (weekly["uncheck_48hr_user_count"] / _t_w * 100).round(1).fillna(0)

        # X축 날짜 짧게
        daily = shorten_dates_in_df(daily, "date")

        # 비율 지표: Google Sheets 수식 결과를 우선 사용 (DB 자체 계산보다 정확)
        cd_sheets = data.get("checkin_daily", pd.DataFrame())
        rate_cols = ["안부미확인률", "48시간미확인률", "안부체크응답률", "콜응답률", "AI케어응답률",
                     "안부체크비중", "동작감지비중", "AI케어비중", "안부확인콜비중", "안부체크율"]

        if not cd_sheets.empty and "날짜" in cd_sheets.columns:
            # Google Sheets 날짜를 짧은 형식으로 변환해서 매칭
            cd_match = cd_sheets.copy()
            cd_match = shorten_dates_in_df(cd_match, "날짜")
            # daily의 date와 매칭
            for rc in rate_cols:
                if rc in cd_match.columns:
                    rate_map = dict(zip(cd_match["날짜"], cd_match[rc]))
                    mapped_col = rc if rc != "콜응답률" else "안부확인콜응답률"
                    daily[mapped_col] = daily["date"].map(rate_map).fillna(0)

        # Google Sheets에 없는 컬럼은 DB에서 fallback 계산
        t = daily["target_user_count"].replace(0, float("nan"))
        if "안부미확인률" not in daily.columns or daily["안부미확인률"].sum() == 0:
            daily["안부미확인률"] = (daily.get("impossible_user_count", 0) / t * 100).round(1).fillna(0)
        if "48시간미확인률" not in daily.columns or daily["48시간미확인률"].sum() == 0:
            daily["48시간미확인률"] = (daily.get("uncheck_48hr_user_count", 0) / t * 100).round(1).fillna(0)
        if "안부체크비중" not in daily.columns or daily["안부체크비중"].sum() == 0:
            daily["안부체크비중"] = (daily["confirm_count"] / daily["complete_user_count"].replace(0, float("nan")) * 100).round(1).fillna(0)
        if "동작감지비중" not in daily.columns or daily["동작감지비중"].sum() == 0:
            daily["동작감지비중"] = (daily.get("detect_motion_count", 0) / daily["complete_user_count"].replace(0, float("nan")) * 100).round(1).fillna(0)
        if "AI케어비중" not in daily.columns or daily["AI케어비중"].sum() == 0:
            daily["AI케어비중"] = (daily["ai_care_response_count"] / daily["complete_user_count"].replace(0, float("nan")) * 100).round(1).fillna(0)
        if "안부확인콜비중" not in daily.columns or daily["안부확인콜비중"].sum() == 0:
            daily["안부확인콜비중"] = (daily["call_response_count"] / daily["complete_user_count"].replace(0, float("nan")) * 100).round(1).fillna(0)
        # 안부체크응답률: off 대상자 반영 (2025-11-15 이후)
        # off 대상자 = 안부체크 발송을 끈 이용자 (분모에서 제외)
        from sheets_data import get_check_off_users as _get_off
        _off_total = sum(_get_off(sheets).values()) if sheets else 0
        if "안부체크응답률" not in daily.columns or daily["안부체크응답률"].sum() == 0:
            _send = daily["alarm_send_count"].copy()
            # 2025-11-15 이후 데이터만 off 대상자 차감
            _off_mask = daily["date"] >= "2025-11-15"
            _adjusted_send = _send.copy()
            _adjusted_send[_off_mask] = _send[_off_mask] - _off_total
            _adjusted_send = _adjusted_send.replace(0, float("nan")).clip(lower=1)
            daily["안부체크응답률"] = (daily["confirm_count"] / _adjusted_send * 100).round(1).fillna(0)
        else:
            # Google Sheets 값이 있어도 off 반영으로 재계산
            _send = daily.get("alarm_send_count", None)
            if _send is not None and _send.sum() > 0:
                _off_mask = daily["date"] >= "2025-11-15"
                _adjusted_send = _send.copy()
                _adjusted_send[_off_mask] = _send[_off_mask] - _off_total
                _adjusted_send = _adjusted_send.replace(0, float("nan")).clip(lower=1)
                daily["안부체크응답률"] = (daily["confirm_count"] / _adjusted_send * 100).round(1).fillna(0)
        if "AI케어응답률" not in daily.columns or daily["AI케어응답률"].sum() == 0:
            daily["AI케어응답률"] = (daily["ai_care_response_count"] / daily["ai_care_generate_count"].replace(0, float("nan")) * 100).round(1).fillna(0)
        if "안부확인콜응답률" not in daily.columns or daily["안부확인콜응답률"].sum() == 0:
            daily["안부확인콜응답률"] = (daily["call_response_count"] / daily["call_generate_count"].replace(0, float("nan")) * 100).round(1).fillna(0)

        # 4/19, 4/20 데이터 제외 (미완료 데이터)
        daily = daily[~daily["date"].isin(["26-04-19", "26-04-20"])]

        st.caption(f"데이터 기간: {daily['date'].min()} ~ {daily['date'].max()}")

        tab1, tab2, tab3, tab4, tab5, tab6, tab7, tab8 = st.tabs([
            "완료현황", "안부확인 비중", "안부확인율",
            "AI케어 알람", "안부확인콜", "📅 주차별 추이",
            "📊 일자별 데이터", "📊 지자체별 데이터"
        ])

        # ── Tab 1: 일별 안부확인 완료현황 (안부미확인률, 48시간미확인률)
        with tab1:
            fig = make_subplots(specs=[[{"secondary_y": True}]])
            fig.add_trace(go.Bar(x=daily["date"], y=daily["complete_user_count"], name="완료자",
                                 marker_color="#00C853",
                                 hovertemplate="%{y:,}명<extra>완료자</extra>"), secondary_y=False)
            fig.add_trace(go.Bar(x=daily["date"], y=daily.get("impossible_user_count", 0), name="미확인자",
                                 marker_color="#FF8A80",
                                 hovertemplate="%{y:,}명<extra>미확인자</extra>"), secondary_y=False)
            fig.add_trace(go.Scatter(x=daily["date"], y=daily["안부미확인률"], name="안부미확인률(%)",
                                     mode="lines", line=dict(color="#FF4B4B", width=2, dash="dot"),
                                     hovertemplate="%{y:.1f}%<extra>안부미확인률</extra>"), secondary_y=True)
            fig.add_trace(go.Scatter(x=daily["date"], y=daily["48시간미확인률"], name="48시간미확인률(%)",
                                     mode="lines", line=dict(color="#D32F2F", width=2),
                                     hovertemplate="%{y:.1f}%<extra>48시간미확인률</extra>"), secondary_y=True)
            fig.update_layout(title="일별 안부확인 완료현황", height=420, hovermode="x unified",
                              barmode="stack", xaxis=dict(type="category"),
                              legend=LEGEND_BELOW)
            fig.update_yaxes(title_text="명", secondary_y=False)
            fig.update_yaxes(title_text="%", secondary_y=True, range=[0, max(30, daily["안부미확인률"].max() * 1.3)])
            st.plotly_chart(fig, use_container_width=True)

        # ── Tab 2: 일별 안부확인 비중 (4개 비중 Area chart)
        with tab2:
            fig = go.Figure()
            fig.add_trace(go.Scatter(x=daily["date"], y=daily["안부체크비중"], name="안부체크",
                                     fill="tozeroy", mode="lines", line=dict(color="#2F5496"),
                                     hovertemplate="%{y:.1f}%<extra>안부체크</extra>"))
            fig.add_trace(go.Scatter(x=daily["date"], y=daily["동작감지비중"], name="동작감지/서비스이용",
                                     fill="tozeroy", mode="lines", line=dict(color="#00897B"),
                                     hovertemplate="%{y:.1f}%<extra>동작감지</extra>"))
            fig.add_trace(go.Scatter(x=daily["date"], y=daily["AI케어비중"], name="AI케어 알람",
                                     fill="tozeroy", mode="lines", line=dict(color="#7B1FA2"),
                                     hovertemplate="%{y:.1f}%<extra>AI케어</extra>"))
            fig.add_trace(go.Scatter(x=daily["date"], y=daily["안부확인콜비중"], name="안부확인콜",
                                     fill="tozeroy", mode="lines", line=dict(color="#E91E63"),
                                     hovertemplate="%{y:.1f}%<extra>안부확인콜</extra>"))
            fig.update_layout(title="일별 안부확인 비중 (완료자 대비 %)", height=400, hovermode="x unified",
                              xaxis=dict(type="category"),
                              legend=LEGEND_BELOW)
            st.plotly_chart(fig, use_container_width=True)

        # ── Tab 3: 일별 안부확인율 + 지자체별 안부확인율 (주간)
        with tab3:
            # 일별 안부확인율 = 안부확인완료자(C열) / 전체회원(B열) × 100 (Sheets 기준)
            _cd_cr = data.get("checkin_daily", pd.DataFrame())
            _comp_c = next((c for c in _cd_cr.columns if "완료자" in str(c) and "안부확인" in str(c)), None)
            _total_c = next((c for c in _cd_cr.columns if str(c).replace("\n","").strip() in ("전체회원", "전체 회원")), None)
            _date_c = next((c for c in _cd_cr.columns if "날짜" in str(c) or str(c).lower() == "date"), None)
            if _comp_c and _total_c and _date_c and not _cd_cr.empty:
                _cd_cr2 = _cd_cr[[_date_c, _comp_c, _total_c]].copy()
                _cd_cr2 = shorten_dates_in_df(_cd_cr2, _date_c)
                _cd_cr2["안부확인율"] = (
                    _cd_cr2[_comp_c].apply(safe_numeric)
                    / _cd_cr2[_total_c].apply(safe_numeric).replace(0, float("nan")) * 100
                ).round(1).fillna(0)
                _cr_map = dict(zip(_cd_cr2[_date_c], _cd_cr2["안부확인율"]))
                daily["안부확인율"] = daily["date"].map(_cr_map).fillna(0)
            elif "complete_user_count" in daily.columns and "target_user_count" in daily.columns:
                daily["안부확인율"] = (
                    daily["complete_user_count"]
                    / daily["target_user_count"].replace(0, float("nan")) * 100
                ).round(1).fillna(0)
            if "안부확인율" in daily.columns and daily["안부확인율"].sum() > 0:
                # 최근 90일만 표시 (최신 날짜가 오른쪽에 꽉 차도록)
                _cr_daily = daily.sort_values("date").tail(90).reset_index(drop=True)
                fig_cr = go.Figure()
                fig_cr.add_trace(go.Scatter(
                    x=_cr_daily["date"], y=_cr_daily["안부확인율"],
                    mode="lines+markers", name="안부확인율",
                    line=dict(color="#2F5496", width=2.5),
                    fill="tozeroy", fillcolor="rgba(47,84,150,0.08)",
                    hovertemplate="<b>%{x}</b><br>안부확인율: %{y:.1f}%<extra></extra>"
                ))
                # x축 틱 수 제한 (최대 18개)
                _cr_dates = _cr_daily["date"].tolist()
                _step = max(1, len(_cr_dates) // 18)
                _tick_vals = _cr_dates[::_step]
                _n_cr = len(_cr_dates)
                fig_cr.update_layout(
                    title=f"일별 안부확인율 (안부확인완료자 / 전체회원) — 최근 {_n_cr}일",
                    height=350, hovermode="x unified",
                    xaxis=dict(
                        type="category", title="",
                        tickmode="array", tickvals=_tick_vals,
                        tickangle=-45, tickfont=dict(size=11),
                        range=[-0.5, _n_cr - 0.5],
                        automargin=True,
                    ),
                    yaxis=dict(title="안부확인율 (%)", range=[0, 100]),
                    margin=dict(t=40, b=80, r=80),
                )
                st.plotly_chart(fig_cr, use_container_width=True)

            # 지자체별 안부확인율 — C:AK(분모)/AL:BT(분자) 직접 계산 바 차트
            cr_direct = biz_filter_df(data.get("checkin_mun_rate_direct", pd.DataFrame()), selected_biz)
            # 0%도 포함해서 전체 지자체(29개) 목록 표시 — registration 기준으로 빠진 지자체 추가
            _all_muns_reg = biz_filter_df(data.get("registration", pd.DataFrame()), selected_biz)
            if not cr_direct.empty and not _all_muns_reg.empty and "지자체명" in _all_muns_reg.columns:
                _latest_d = cr_direct["시작일"].iloc[0] if "시작일" in cr_direct.columns else ""
                _existing = set(cr_direct["지자체명"].tolist())
                _missing = [m for m in _all_muns_reg["지자체명"].tolist() if m not in _existing]
                if _missing:
                    _zero_rows = pd.DataFrame({
                        "지자체명": _missing,
                        "안부확인율": 0.0,
                        "분자": 0,
                        "분모": 0,
                        "시작일": _latest_d,
                    })
                    cr_direct = pd.concat([cr_direct, _zero_rows], ignore_index=True)
            if not cr_direct.empty and "안부확인율" in cr_direct.columns:
                latest_date = cr_direct["시작일"].iloc[0] if not cr_direct.empty else str(pd.Timestamp.now().date())
                latest_cr = cr_direct.copy()
                latest_cr["권역"] = latest_cr["지자체명"].map(DETAIL_REGION).fillna("기타")
                latest_cr = latest_cr.sort_values("안부확인율", ascending=True).copy()
                latest_cr["지자체명_표시"] = latest_cr["지자체명"].apply(_mun_label)

                st.markdown(f"**{latest_date} 기준**  |  분모: C~AK열, 분자: AL~BT열")
                fig_mun = px.bar(
                    latest_cr, y="지자체명_표시", x="안부확인율", orientation="h",
                    color="권역", color_discrete_map=REGION_COLORS,
                    custom_data=["분자", "분모"],
                    height=min(800, max(480, len(latest_cr) * 26)),
                )
                fig_mun.update_layout(
                    title=f"지자체별 안부확인율 ({latest_date})",
                    legend=LEGEND_BELOW, margin=dict(t=40, b=70),
                    xaxis=dict(range=[0, 105], title="안부확인율 (%)"),
                    yaxis=dict(tickfont=dict(size=11)),
                )
                fig_mun.update_traces(
                    texttemplate="%{x:.1f}%", textposition="outside",
                    textfont=dict(size=12),
                    hovertemplate=(
                        "<b>%{y}</b><br>"
                        "안부확인율: %{x:.1f}%<br>"
                        "분자(안부확인): %{customdata[0]:,}명<br>"
                        "분모(전체발송): %{customdata[1]:,}명"
                        "<extra></extra>"
                    ),
                )
                st.plotly_chart(fig_mun, use_container_width=True)
            else:
                st.info("지자체별 안부확인율 데이터가 없습니다. (안부확인지자체 시트 확인 필요)")

            # ── 사업구분별 주차별 안부확인율 추이
            st.divider()
            _mw_all = data.get("checkin_mun_weekly", pd.DataFrame())
            if not _mw_all.empty and "시작일" in _mw_all.columns:
                if selected_biz != "전체":
                    # 선택된 사업구분의 지자체별 + 집계 라인
                    _mw_biz = biz_filter_df(_mw_all, selected_biz, col="지자체명")
                    if not _mw_biz.empty:
                        _mw_biz = _mw_biz.copy()
                        _mw_biz["주차"] = _mw_biz["시작일"].apply(date_to_week_label)
                        _mw_biz = _mw_biz[(_mw_biz["시작일"] >= date_start) & (_mw_biz["시작일"] <= date_end)]
                        _mw_biz = _mw_biz.sort_values("주차")
                        if not _mw_biz.empty:
                            _agg_cr = _mw_biz.groupby("주차").agg({"분모": "sum", "분자": "sum"}).reset_index()
                            _agg_cr["안부확인율"] = (_agg_cr["분자"] / _agg_cr["분모"].replace(0, float("nan")) * 100).round(1).fillna(0)
                            _biz_colors2 = ["#42A5F5","#66BB6A","#AB47BC","#EC407A","#26C6DA","#FFA726","#8D6E63","#78909C"]
                            fig_cr_biz = go.Figure()
                            for _i, (_mn, _mdf) in enumerate(_mw_biz.groupby("지자체명")):
                                _mdf_w = _mdf.groupby("주차").agg({"분모": "sum", "분자": "sum"}).reset_index()
                                _mdf_w["rate"] = (_mdf_w["분자"] / _mdf_w["분모"].replace(0, float("nan")) * 100).round(1).fillna(0)
                                fig_cr_biz.add_trace(go.Scatter(
                                    x=_mdf_w["주차"], y=_mdf_w["rate"],
                                    mode="lines+markers", name=_mn,
                                    line=dict(width=1.5, color=_biz_colors2[_i % len(_biz_colors2)]),
                                    marker=dict(size=4), opacity=0.8,
                                    hovertemplate=f"<b>%{{x}}</b><br>{_mn}: %{{y:.1f}}%<extra></extra>",
                                ))
                            fig_cr_biz.add_trace(go.Scatter(
                                x=_agg_cr["주차"], y=_agg_cr["안부확인율"],
                                mode="lines+markers", name=f"{selected_biz} 평균",
                                line=dict(color="#FF6F00", width=3),
                                marker=dict(size=7, symbol="diamond"),
                                hovertemplate=f"<b>%{{x}}</b><br>{selected_biz} 평균: %{{y:.1f}}%<extra></extra>",
                            ))
                            fig_cr_biz.update_layout(
                                title=f"주차별 {selected_biz} 지자체별 안부확인율 추이",
                                height=460, hovermode="x unified",
                                xaxis=dict(type="category", tickangle=-45),
                                yaxis=dict(title="안부확인율 (%)", range=[0, 110]),
                                margin=dict(t=40, b=90),
                                legend=dict(orientation="h", yanchor="top", y=-0.2, xanchor="center", x=0.5),
                            )
                            st.plotly_chart(fig_cr_biz, use_container_width=True)
                else:
                    # 전체: 사업구분별 집계 라인 비교
                    _mw_all2 = _mw_all.copy()
                    _mw_all2["주차"] = _mw_all2["시작일"].apply(date_to_week_label)
                    _mw_all2 = _mw_all2[(_mw_all2["시작일"] >= date_start) & (_mw_all2["시작일"] <= date_end)]
                    _mw_all2 = _mw_all2.sort_values("주차")
                    _mw_all2["사업구분"] = _mw_all2["지자체명"].apply(
                        lambda n: next((v for k, v in BUSINESS_TYPE_MAP.items()
                                        if k.replace(" ", "") == n.replace(" ", "")
                                        or k.replace(" ", "") in n.replace(" ", "")
                                        or n.replace(" ", "") in k.replace(" ", "")), "기타")
                    )
                    _biz_line_colors = {
                        "통합돌봄": "#42A5F5", "노인맞춤돌봄": "#66BB6A",
                        "고독사예방": "#AB47BC", "장애인지원": "#EC407A", "기타": "#78909C",
                    }
                    fig_cr_all = go.Figure()
                    for _bz, _bdf in _mw_all2.groupby("사업구분"):
                        _bagg = _bdf.groupby("주차").agg({"분모": "sum", "분자": "sum"}).reset_index()
                        _bagg["안부확인율"] = (_bagg["분자"] / _bagg["분모"].replace(0, float("nan")) * 100).round(1).fillna(0)
                        fig_cr_all.add_trace(go.Scatter(
                            x=_bagg["주차"], y=_bagg["안부확인율"],
                            mode="lines+markers", name=_bz,
                            line=dict(width=2, color=_biz_line_colors.get(_bz, "#90A4AE")),
                            marker=dict(size=5),
                            hovertemplate=f"<b>%{{x}}</b><br>{_bz}: %{{y:.1f}}%<extra></extra>",
                        ))
                    fig_cr_all.update_layout(
                        title="사업구분별 주차별 안부확인율 추이",
                        height=420, hovermode="x unified",
                        xaxis=dict(type="category", tickangle=-45),
                        yaxis=dict(title="안부확인율 (%)", range=[0, 110]),
                        margin=dict(t=40, b=90),
                        legend=dict(orientation="h", yanchor="top", y=-0.2, xanchor="center", x=0.5),
                    )
                    st.plotly_chart(fig_cr_all, use_container_width=True)

        # ── Tab 4: 일별 AI케어 알람 응답 (발송수, 응답자수, 응답률)
        with tab4:
            fig = make_subplots(specs=[[{"secondary_y": True}]])
            fig.add_trace(go.Bar(x=daily["date"], y=daily["ai_care_generate_count"], name="AI케어 발송수",
                                 marker_color="#CE93D8",
                                 hovertemplate="%{y:,}건<extra>발송수</extra>"), secondary_y=False)
            fig.add_trace(go.Bar(x=daily["date"], y=daily["ai_care_response_count"], name="AI케어 응답자수",
                                 marker_color="#7B1FA2",
                                 hovertemplate="%{y:,}명<extra>응답자</extra>"), secondary_y=False)
            fig.add_trace(go.Scatter(x=daily["date"], y=daily["AI케어응답률"], name="AI케어 응답률(%)",
                                     mode="lines+markers", line=dict(color="#FF6F00", width=2),
                                     hovertemplate="%{y:.1f}%<extra>응답률</extra>"), secondary_y=True)
            fig.update_layout(title="일별 AI케어 알람 응답", height=420, hovermode="x unified",
                              barmode="group", xaxis=dict(type="category"),
                              legend=LEGEND_BELOW)
            fig.update_yaxes(title_text="건/명", secondary_y=False)
            fig.update_yaxes(title_text="응답률(%)", secondary_y=True, range=[0, 110])
            st.plotly_chart(fig, use_container_width=True)

        # ── Tab 5: 일별 안부확인콜 (발송수, 응답자수, 응답률)
        with tab5:
            fig = make_subplots(specs=[[{"secondary_y": True}]])
            fig.add_trace(go.Bar(x=daily["date"], y=daily["call_generate_count"], name="콜 발송수",
                                 marker_color="#F48FB1",
                                 hovertemplate="%{y:,}건<extra>발송수</extra>"), secondary_y=False)
            fig.add_trace(go.Bar(x=daily["date"], y=daily["call_response_count"], name="콜 응답자수",
                                 marker_color="#E91E63",
                                 hovertemplate="%{y:,}명<extra>응답자</extra>"), secondary_y=False)
            fig.add_trace(go.Scatter(x=daily["date"], y=daily["안부확인콜응답률"], name="콜 응답률(%)",
                                     mode="lines+markers", line=dict(color="#FF6F00", width=2),
                                     hovertemplate="%{y:.1f}%<extra>응답률</extra>"), secondary_y=True)
            fig.update_layout(title="일별 안부확인콜", height=420, hovermode="x unified",
                              barmode="group", xaxis=dict(type="category"),
                              legend=LEGEND_BELOW)
            fig.update_yaxes(title_text="건/명", secondary_y=False)
            fig.update_yaxes(title_text="응답률(%)", secondary_y=True, range=[0, 110])
            st.plotly_chart(fig, use_container_width=True)

        # ── Tab 6: 주차별 추이 (사업구분 필터 반영)
        with tab6:
            if not weekly.empty and "week" in weekly.columns:
                biz_label = f" ({selected_biz})" if selected_biz != "전체" else ""
                st.markdown(f"**주차별 안부확인 추이{biz_label}** — {weekly['week'].min()} ~ {weekly['week'].max()}")

                fig_w = make_subplots(specs=[[{"secondary_y": True}]])
                if "confirm_count" in weekly.columns:
                    fig_w.add_trace(go.Bar(x=weekly["week"], y=weekly["confirm_count"], name="안부체크 완료",
                                           marker_color="#00C853",
                                           hovertemplate="%{y:,}건<extra>완료</extra>"), secondary_y=False)
                if "alarm_send_count" in weekly.columns:
                    fig_w.add_trace(go.Bar(x=weekly["week"], y=weekly["alarm_send_count"], name="발송수",
                                           marker_color="#B0BEC5",
                                           hovertemplate="%{y:,}건<extra>발송</extra>"), secondary_y=False)
                if "안부체크응답률" in weekly.columns:
                    fig_w.add_trace(go.Scatter(x=weekly["week"], y=weekly["안부체크응답률"], name="안부체크응답률(%)",
                                               mode="lines+markers", line=dict(color="#2196F3", width=2.5),
                                               marker=dict(size=6),
                                               hovertemplate="%{y:.1f}%<extra>응답률</extra>"), secondary_y=True)
                if "안부미확인률" in weekly.columns:
                    fig_w.add_trace(go.Scatter(x=weekly["week"], y=weekly["안부미확인률"], name="안부미확인률(%)",
                                               mode="lines+markers", line=dict(color="#FF4B4B", width=2, dash="dot"),
                                               marker=dict(size=5),
                                               hovertemplate="%{y:.1f}%<extra>미확인률</extra>"), secondary_y=True)
                fig_w.update_layout(
                    title=f"주차별 안부확인 현황{biz_label}", height=420, hovermode="x unified",
                    barmode="group", xaxis=dict(type="category", tickangle=-45),
                    legend=LEGEND_BELOW, margin=dict(t=40, b=90),
                )
                fig_w.update_yaxes(title_text="건/명", secondary_y=False)
                fig_w.update_yaxes(title_text="%", secondary_y=True, range=[0, 110])
                st.plotly_chart(fig_w, use_container_width=True)

                if "48시간미확인률" in weekly.columns:
                    fig_w2 = go.Figure()
                    fig_w2.add_trace(go.Scatter(x=weekly["week"], y=weekly["48시간미확인률"], name="48시간미확인률",
                                                mode="lines+markers", line=dict(color="#D32F2F", width=2),
                                                marker=dict(size=6),
                                                hovertemplate="%{y:.1f}%<extra>48h미확인</extra>"))
                    fig_w2.update_layout(title=f"주차별 48시간 미확인률{biz_label}", height=280,
                                         xaxis=dict(type="category", tickangle=-45),
                                         yaxis=dict(title="%", range=[0, max(5, weekly["48시간미확인률"].max() * 1.3)]),
                                         margin=dict(t=40, b=70))
                    st.plotly_chart(fig_w2, use_container_width=True)
            else:
                st.info("주차별 집계 데이터가 없습니다.")

            # ── 사업구분별 지자체 안부확인율 추이 (시트 원본 데이터 활용)
            if selected_biz != "전체":
                mun_weekly = data.get("checkin_mun_weekly", pd.DataFrame())
                if not mun_weekly.empty and "시작일" in mun_weekly.columns:
                    mun_biz = biz_filter_df(mun_weekly, selected_biz, col="지자체명")
                    if not mun_biz.empty:
                        mun_biz = mun_biz.copy()
                        mun_biz["주차"] = mun_biz["시작일"].apply(date_to_week_label)
                        # 기간 필터 (date_start/date_end 기반)
                        mun_biz = mun_biz[
                            (mun_biz["시작일"] >= date_start) &
                            (mun_biz["시작일"] <= date_end)
                        ]
                        mun_biz = mun_biz.sort_values("주차")

                        if not mun_biz.empty:
                            # 사업구분 전체 집계 (분자합/분모합)
                            agg_w = mun_biz.groupby("주차").agg({"분모": "sum", "분자": "sum"}).reset_index()
                            agg_w["안부확인율"] = (
                                agg_w["분자"] / agg_w["분모"].replace(0, float("nan")) * 100
                            ).round(1).fillna(0)

                            st.markdown(f"**{selected_biz} 지자체별 주차별 안부확인율**")
                            _biz_colors = [
                                "#42A5F5", "#66BB6A", "#AB47BC", "#EC407A",
                                "#26C6DA", "#FFA726", "#8D6E63", "#78909C",
                            ]
                            fig_mw = go.Figure()
                            for i, (mun_name, mdf) in enumerate(mun_biz.groupby("지자체명")):
                                mdf_w = mdf.groupby("주차").agg({"분모": "sum", "분자": "sum"}).reset_index()
                                mdf_w["rate"] = (
                                    mdf_w["분자"] / mdf_w["분모"].replace(0, float("nan")) * 100
                                ).round(1).fillna(0)
                                fig_mw.add_trace(go.Scatter(
                                    x=mdf_w["주차"], y=mdf_w["rate"],
                                    mode="lines+markers", name=mun_name,
                                    line=dict(width=1.5, color=_biz_colors[i % len(_biz_colors)]),
                                    marker=dict(size=4), opacity=0.8,
                                    hovertemplate=f"<b>%{{x}}</b><br>{mun_name}: %{{y:.1f}}%<extra></extra>",
                                ))
                            fig_mw.add_trace(go.Scatter(
                                x=agg_w["주차"], y=agg_w["안부확인율"],
                                mode="lines+markers", name=f"{selected_biz} 평균",
                                line=dict(color="#FF6F00", width=3),
                                marker=dict(size=7, symbol="diamond"),
                                hovertemplate=f"<b>%{{x}}</b><br>{selected_biz} 평균: %{{y:.1f}}%<extra></extra>",
                            ))
                            fig_mw.update_layout(
                                title=f"주차별 {selected_biz} 지자체별 안부확인율 추이",
                                height=460, hovermode="x unified",
                                xaxis=dict(type="category", tickangle=-45),
                                yaxis=dict(title="안부확인율 (%)", range=[0, 110]),
                                margin=dict(t=40, b=90),
                                legend=dict(orientation="h", yanchor="top", y=-0.2, xanchor="center", x=0.5),
                            )
                            st.plotly_chart(fig_mw, use_container_width=True)

        # ── Tab 7: 일자별 전체 데이터 테이블 (Google Sheets gid=261480368 형태)
        with tab7:
            st.markdown("**일자별 안부확인 전체 데이터** (Google Sheets `복약확인알림(전체)` 시트와 동일)")

            # 일자별 기간 필터
            date_list = sorted(daily["date"].unique())
            if len(date_list) > 1:
                d_col1, d_col2 = st.columns(2)
                with d_col1:
                    date_from = st.selectbox("시작일", date_list, index=max(0, len(date_list)-14), key="daily_from")
                with d_col2:
                    date_to = st.selectbox("종료일", date_list, index=len(date_list)-1, key="daily_to")
                daily_filtered = daily[(daily["date"] >= date_from) & (daily["date"] <= date_to)].copy()
            else:
                daily_filtered = daily.copy()

            # 표시 컬럼 정리
            display_cols = {
                "date": "날짜",
                "target_user_count": "전체 대상자",
                "complete_user_count": "안부확인 완료자",
                "impossible_user_count": "안부미확인자",
                "uncheck_48hr_user_count": "48시간 미확인",
                "confirm_count": "①안부체크 응답자",
                "detect_motion_count": "②동작감지 이용자",
                "ai_care_response_count": "③AI케어 응답자",
                "call_response_count": "④안부확인콜 응답자",
                "alarm_send_count": "안부체크 발송수",
                "ai_care_generate_count": "AI케어 발송수",
                "call_generate_count": "안부확인콜 발송수",
                "안부미확인률": "안부미확인률(%)",
                "48시간미확인률": "48시간미확인률(%)",
                "안부체크응답률": "안부체크응답률(%)",
                "AI케어응답률": "AI케어응답률(%)",
                "안부확인콜응답률": "안부확인콜응답률(%)",
            }
            available_display = {k: v for k, v in display_cols.items() if k in daily_filtered.columns}
            daily_display = daily_filtered[list(available_display.keys())].rename(columns=available_display)
            daily_display = daily_display.sort_values("날짜", ascending=False)

            st.dataframe(
                daily_display,
                use_container_width=True,
                height=min(600, len(daily_display) * 35 + 50),
                column_config={
                    "안부미확인률(%)": st.column_config.NumberColumn(format="%.1f%%"),
                    "48시간미확인률(%)": st.column_config.NumberColumn(format="%.1f%%"),
                    "안부체크응답률(%)": st.column_config.NumberColumn(format="%.1f%%"),
                    "AI케어응답률(%)": st.column_config.NumberColumn(format="%.1f%%"),
                    "안부확인콜응답률(%)": st.column_config.NumberColumn(format="%.1f%%"),
                },
            )
            st.caption(f"총 {len(daily_display)}일 데이터")

        # ── Tab 8: 지자체별 데이터 테이블
        with tab8:
            st.markdown("**지자체별 안부확인 데이터** (Google Sheets `복약확인알림(전체지자체)` 시트와 동일)")

            # 날짜 선택
            all_dates = sorted(safety_db["date"].unique())
            if all_dates:
                selected_date = st.selectbox("날짜 선택", list(reversed(all_dates)), index=0, key="mun_date")
                mun_day = safety_db[safety_db["date"] == selected_date].copy()

                if not mun_day.empty:
                    # 지표 계산
                    t = mun_day["target_user_count"].replace(0, float("nan"))
                    mun_day["안부체크율(%)"] = (mun_day["confirm_count"] / mun_day["alarm_send_count"].replace(0, float("nan")) * 100).round(1).fillna(0)
                    mun_day["안부미확인률(%)"] = (mun_day["impossible_user_count"] / t * 100).round(1).fillna(0)
                    mun_day["48시간미확인률(%)"] = (mun_day.get("uncheck_48hr_user_count", 0) / t * 100).round(1).fillna(0)

                    # 표시용 정리
                    mun_display_cols = {
                        "agency_name": "지자체명",
                        "target_user_count": "대상자수",
                        "complete_user_count": "완료자수",
                        "impossible_user_count": "미확인자",
                        "confirm_count": "안부체크응답",
                        "detect_motion_count": "동작감지",
                        "ai_care_response_count": "AI케어응답",
                        "call_response_count": "콜응답",
                        "alarm_send_count": "체크발송수",
                        "ai_care_generate_count": "AI케어발송",
                        "call_generate_count": "콜발송",
                        "uncheck_48hr_user_count": "48시간미확인",
                        "안부체크율(%)": "안부체크율(%)",
                        "안부미확인률(%)": "안부미확인률(%)",
                        "48시간미확인률(%)": "48시간미확인률(%)",
                    }
                    available_mun = {k: v for k, v in mun_display_cols.items() if k in mun_day.columns}
                    mun_display = mun_day[list(available_mun.keys())].rename(columns=available_mun)
                    # 발송수 0인 지자체 제외
                    if "체크발송수" in mun_display.columns:
                        mun_display = mun_display[mun_display["체크발송수"] > 0]
                    mun_display = mun_display.sort_values("대상자수", ascending=False)

                    st.dataframe(
                        mun_display,
                        use_container_width=True,
                        height=min(600, len(mun_display) * 35 + 50),
                        column_config={
                            "안부체크율(%)": st.column_config.NumberColumn(format="%.1f%%"),
                            "안부미확인률(%)": st.column_config.NumberColumn(format="%.1f%%"),
                            "48시간미확인률(%)": st.column_config.NumberColumn(format="%.1f%%"),
                        },
                    )
                    st.caption(f"{selected_date} 기준 {len(mun_display)}개 지자체")

    else:
        # DB 데이터 없으면 Google Sheets 데이터 사용 (기존 방식)

        # ① 전체 평균 추이: 안부확인전체 시트 R열(안부미확인률) → 100 - 값, 주차별 평균
        # weekly_users 시작일 기준 날짜 범위로 묶어 Summary와 동일한 기준 사용
        cd_raw = data.get("checkin_daily", pd.DataFrame())
        wu_trend = data.get("weekly_users", pd.DataFrame())
        # 날짜 컬럼 탐색
        _cd_date_col = None
        for _c in cd_raw.columns:
            _cl = str(_c).replace("\n", "").strip().lower()
            if _cl in ("날짜", "date", "일자", "일") or "날짜" in _cl or "date" in _cl:
                _cd_date_col = _c
                break
        if not cd_raw.empty and "안부확인율" in cd_raw.columns and _cd_date_col is not None:
            cd_trend = cd_raw[cd_raw["안부확인율"] > 0].copy()
            cd_trend["_dt"] = pd.to_datetime(cd_trend[_cd_date_col].astype(str), errors="coerce")
            cd_trend = cd_trend.dropna(subset=["_dt"]).sort_values("_dt")
            # weekly_users 기준 날짜→주차 매핑 (Summary와 동일 경계)
            if not wu_trend.empty and "주차" in wu_trend.columns and "시작일" in wu_trend.columns:
                _wmap = {}
                for _, _r in wu_trend.iterrows():
                    try:
                        _s = pd.to_datetime(str(_r["시작일"]), errors="coerce")
                        if pd.isna(_s):
                            continue
                        for _i in range(7):
                            _wmap[(_s + pd.Timedelta(days=_i)).strftime("%Y-%m-%d")] = str(_r["주차"])
                    except Exception:
                        pass
                cd_trend["_wk"] = cd_trend["_dt"].dt.strftime("%Y-%m-%d").map(_wmap)
                cd_trend = cd_trend[cd_trend["_wk"].notna()]
                avg_cr = cd_trend.groupby("_wk")["안부확인율"].mean().round(1).reset_index()
                avg_cr = avg_cr.rename(columns={"_wk": "시작일"})
            else:
                cd_trend["_날짜"] = cd_trend["_dt"].apply(lambda d: date_to_week_label(d.strftime("%Y-%m-%d")))
                avg_cr = cd_trend.groupby("_날짜")["안부확인율"].mean().round(1).reset_index()
                avg_cr = avg_cr.rename(columns={"_날짜": "시작일"})
            avg_cr = avg_cr[avg_cr["시작일"].astype(str).str.startswith("26-")]
            if not avg_cr.empty:
                fig = px.line(avg_cr, x="시작일", y="안부확인율", markers=True,
                              color_discrete_sequence=["#2F5496"])
                fig.update_layout(title="안부확인율 추이 (전체 평균)", height=350,
                                  hovermode="x unified", xaxis=dict(type="category"),
                                  yaxis=dict(title="안부확인율 (%)", range=[0, 100]),
                                  margin=dict(t=40, b=30))
                fig.update_traces(hovertemplate="<b>%{x}</b><br>평균 안부확인율: %{y:.1f}%<extra></extra>")
                st.plotly_chart(fig, use_container_width=True)
        else:
            st.info("안부확인 데이터가 없습니다. ('안부확인전체' 시트 R열 안부미확인률 필요)")

        # ② 지자체별 안부확인율 — C:AK/AL:BT 직접 계산 바 차트
        cr_direct2 = data.get("checkin_mun_rate_direct", pd.DataFrame())
        if not cr_direct2.empty and "안부확인율" in cr_direct2.columns:
            latest_date2 = cr_direct2["시작일"].iloc[0]
            latest_cr2 = cr_direct2.copy()
            latest_cr2["권역"] = latest_cr2["지자체명"].map(DETAIL_REGION).fillna("기타")
            latest_cr2 = latest_cr2.sort_values("안부확인율", ascending=True).copy()
            latest_cr2["지자체명_표시"] = latest_cr2["지자체명"].apply(_mun_label)
            st.markdown(f"**{latest_date2} 기준**  |  분모: C~AK열, 분자: AL~BT열")
            fig2 = px.bar(
                latest_cr2, y="지자체명_표시", x="안부확인율", orientation="h",
                color="권역", color_discrete_map=REGION_COLORS,
                custom_data=["분자", "분모"],
                height=min(520, max(320, len(latest_cr2) * 22)),
            )
            fig2.update_layout(
                title=f"지자체별 안부확인율 ({latest_date2})",
                legend=LEGEND_BELOW, margin=dict(t=40, b=70),
                xaxis=dict(range=[0, 105], title="안부확인율 (%)"),
                yaxis=dict(tickfont=dict(size=11)),
            )
            fig2.update_traces(
                texttemplate="%{x:.1f}%", textposition="outside",
                textfont=dict(size=12),
                hovertemplate=(
                    "<b>%{y}</b><br>"
                    "안부확인율: %{x:.1f}%<br>"
                    "분자(안부확인): %{customdata[0]:,}명<br>"
                    "분모(전체발송): %{customdata[1]:,}명"
                    "<extra></extra>"
                ),
            )
            st.plotly_chart(fig2, use_container_width=True)


# ============================================================
# ❤ 심혈관체크
# ============================================================
elif page == "❤ 6.심혈관체크":
    st.markdown('<div class="section-header">❤ 심혈관체크</div>', unsafe_allow_html=True)
    selected_biz = biz_selector("심혈관")
    st.divider()

    p_start, p_end = page_week_range_selector("cardio", weeks)

    tab1, tab2 = st.tabs(["이용자수 추이", "검사횟수 추이"])

    with tab1:
        cardio_users = data.get("weekly_심혈관이용자", pd.DataFrame())
        cardio_user_raw = sheets.get("심혈관이용자", pd.DataFrame())
        if not cardio_user_raw.empty:
            cu = cardio_user_raw.copy()
            _wc, _sum_col, _rc = None, None, None
            for c in cu.columns:
                cl = str(c).replace("\n", "").strip()
                if "주차" in cl and _wc is None: _wc = c
                elif ("이용자합계" in cl or ("합계" in cl and "이용자" in cl)) and _sum_col is None: _sum_col = c
                elif ("전체이용비중" in cl or "이용비중" in cl) and _rc is None: _rc = c
            # C열 합계 우선
            cardio_total_c = data.get("total_심혈관이용자", pd.DataFrame())
            if _wc:
                cu = filter_by_week_range(cu, _wc, p_start, p_end, weeks)
                cu = shorten_dates_in_df(cu, _wc)
                if not cardio_total_c.empty:
                    ct = filter_by_week_range(cardio_total_c, "주차", p_start, p_end, weeks)
                    ct = shorten_dates_in_df(ct, "주차")
                    cu = cu.copy()
                    ct_map = dict(zip(ct["주차"], ct["값"].apply(safe_numeric)))
                    cu["_bar"] = cu[_wc].map(ct_map).fillna(cu[_sum_col].apply(safe_numeric) if _sum_col else 0)
                    bar_col_use = "_bar"
                else:
                    bar_col_use = _sum_col
                if bar_col_use:
                    if selected_biz != "전체":
                        _biz_cnt, _ = biz_agg_raw(cu, selected_biz, _wc)
                        if _biz_cnt is not None:
                            cu = cu.copy()
                            cu["_bar"] = _biz_cnt
                            bar_col_use = "_bar"
                            # 시트 내장 이용비중 대신 실제 가입완료 인원으로 직접 계산
                            _biz_completed = biz_filter_df(data.get("registration", pd.DataFrame()), selected_biz)["가입완료"].apply(safe_numeric).sum()
                            if _biz_completed > 0:
                                cu["_rc"] = (_biz_cnt / _biz_completed * 100).round(1)
                                _rc = "_rc"
                    plot_bar_rate_dual(cu, _wc, bar_col_use, "이용자수", "#EF5350",
                                       _rc, "이용비중", "#FF6F00",
                                       "심혈관체크 이용자수 + 이용비중")
            # ── 지자체별 이용자비중 추이 (구글 시트 AI~BK열 직접 사용) ──────────────────────────
            mrt_cardio = biz_filter_df(extract_mun_ratio_trend(cardio_user_raw), selected_biz)
            if not mrt_cardio.empty:
                mrt_cardio = filter_by_week_range(mrt_cardio, "주차", p_start, p_end, weeks)
                _active_c = mrt_cardio.groupby("지자체명")["값"].sum()
                _active_c = _active_c[_active_c > 0].index.tolist()
                mrt_cardio = mrt_cardio[mrt_cardio["지자체명"].isin(_active_c)]
                if not mrt_cardio.empty:
                    plot_municipality_lines(mrt_cardio, "지자체별 심혈관체크 이용자비중 추이 (%)", metric_label="이용자비중(%)")
        else:
            st.info("심혈관 이용자 데이터가 없습니다.")

    with tab2:
        cardio_exam = data.get("weekly_심혈관검사", pd.DataFrame())
        cardio_exam_raw = sheets.get("심혈관검사횟수", pd.DataFrame())
        if not cardio_exam_raw.empty:
            ce = cardio_exam_raw.copy()
            _wc, _sum_col, _awc = None, None, None
            for c in ce.columns:
                cl = str(c).replace("\n", "").strip()
                if "주차" in cl and _wc is None: _wc = c
                elif "합계" in cl and _sum_col is None: _sum_col = c
                elif "1인" in cl and "주평균" in cl and _awc is None: _awc = c
            cardio_exam_total_c = data.get("total_심혈관검사", pd.DataFrame())
            if _wc:
                ce = filter_by_week_range(ce, _wc, p_start, p_end, weeks)
                ce = shorten_dates_in_df(ce, _wc)
                if not cardio_exam_total_c.empty:
                    cet = filter_by_week_range(cardio_exam_total_c, "주차", p_start, p_end, weeks)
                    cet = shorten_dates_in_df(cet, "주차")
                    ct_map = dict(zip(cet["주차"], cet["값"].apply(safe_numeric)))
                    ce["_bar"] = ce[_wc].map(ct_map).fillna(ce[_sum_col].apply(safe_numeric) if _sum_col else 0)
                    bar_col_use = "_bar"
                else:
                    bar_col_use = _sum_col
                if bar_col_use:
                    if selected_biz != "전체":
                        _biz_cnt, _ = biz_agg_raw(ce, selected_biz, _wc)
                        if _biz_cnt is not None:
                            ce = ce.copy()
                            ce["_bar"] = _biz_cnt
                            bar_col_use = "_bar"
                    plot_bar_rate_dual(ce, _wc, bar_col_use, "검사횟수", "#EF5350",
                                       _awc, "1인 주평균", "#455A64",
                                       "심혈관 검사횟수 + 1인 주평균", bar_unit="회", line_unit="회")
            cf = filter_by_week_range(cardio_exam, "주차", p_start, p_end, weeks) if not cardio_exam.empty else pd.DataFrame()
            cf = biz_filter_df(cf, selected_biz)
            if not cf.empty:
                plot_municipality_lines(cf, "지자체별 심혈관 검사횟수 추이", metric_label="검사횟수")
        else:
            st.info("심혈관 검사횟수 데이터가 없습니다.")


# ============================================================
# 💊 복약관리
# ============================================================
elif page == "💊 8.복약관리":
    st.markdown('<div class="section-header">💊 복약관리</div>', unsafe_allow_html=True)
    st.divider()
    p_start, p_end = page_week_range_selector("med", weeks)

    # 데이터 준비
    med_users = data.get("weekly_복약등록회원", pd.DataFrame())
    med_count = data.get("weekly_복약등록건수", pd.DataFrame())
    reg = data.get("registration", pd.DataFrame())

    # 전체 가입완료자 수 (비율 계산용)
    total_registered = 0
    if not reg.empty and "가입완료" in reg.columns:
        total_registered = int(reg["가입완료"].apply(safe_numeric).sum())

    tab1, tab2, tab3, tab4 = st.tabs(["활성 이용자 복약 이용자수", "활성 이용자 복약 등록건수", "지자체별 비중", "상세 데이터"])

    with tab1:
        med_raw = sheets.get("복약등록회원", pd.DataFrame())
        if not med_raw.empty:
            mr = med_raw.copy()
            _wc, _sum_col, _ratio_col = None, None, None
            for c in mr.columns:
                cl = str(c).replace("\n", "").strip()
                if "주차" in cl: _wc = c
                elif "이용자" in cl and "합계" in cl: _sum_col = c
                elif cl == "비율" or ("비율" in cl and "WoW" not in cl and "1인" not in cl): _ratio_col = c
            if _wc and _sum_col:
                mr[_sum_col] = mr[_sum_col].apply(safe_numeric)
                if _ratio_col:
                    mr[_ratio_col] = mr[_ratio_col].apply(safe_numeric)
                mr = filter_by_week_range(mr, _wc, p_start, p_end, weeks)
                mr = shorten_dates_in_df(mr, _wc)
                mr = mr[mr[_wc].astype(str).str.strip() != ""]
                fig = make_subplots(specs=[[{"secondary_y": True}]])
                fig.add_trace(go.Bar(
                    x=mr[_wc], y=mr[_sum_col],
                    name="이용자 수 합계", marker_color="#424242",
                    text=mr[_sum_col].apply(lambda x: f"{x:,.0f}"),
                    textposition="outside", textfont=dict(size=13),
                    hovertemplate="<b>%{x}</b><br>이용자수: %{y:,}명<extra></extra>"
                ), secondary_y=False)
                if _ratio_col:
                    fig.add_trace(go.Scatter(
                        x=mr[_wc], y=mr[_ratio_col],
                        name="비율", mode="lines+markers+text",
                        line=dict(color="#FF6F00", width=2),
                        text=mr[_ratio_col].apply(lambda x: f"{x:.0f}%"),
                        textposition="top center", textfont=dict(size=13, color="#FF6F00"),
                        hovertemplate="<b>%{x}</b><br>비율: %{y:.1f}%<extra></extra>"
                    ), secondary_y=True)
                fig.update_layout(
                    title="활성 이용자 복약 이용자수",
                    height=450, hovermode="x unified",
                    xaxis=dict(type="category"),
                    legend=LEGEND_BELOW, margin=dict(t=40, b=70), bargap=0.3,
                )
                fig.update_yaxes(title_text="이용자수 (명)", secondary_y=False)
                if _ratio_col:
                    max_ratio = mr[_ratio_col].max()
                    fig.update_yaxes(title_text="비율 (%)", secondary_y=True,
                                     range=[0, max(25, max_ratio * 1.3) if max_ratio > 0 else 25])
                st.plotly_chart(fig, use_container_width=True)
            else:
                st.info("복약 등록 회원수 시트에서 합계/비율 컬럼을 찾을 수 없습니다.")
        else:
            st.info("복약 등록 회원수 데이터가 없습니다.")

    with tab2:
        med_count_raw = sheets.get("복약등록건수", pd.DataFrame())
        med_count_c = data.get("total_복약등록건수", pd.DataFrame())
        if not med_count_raw.empty:
            mc_raw = med_count_raw.copy()
            _wc, _sum_col, _rc = None, None, None
            for c in mc_raw.columns:
                cl = str(c).replace("\n", "").strip()
                if "주차" in cl and _wc is None: _wc = c
                elif "합계" in cl and _sum_col is None: _sum_col = c
                elif ("전체이용비중" in cl or "이용비중" in cl) and _rc is None: _rc = c
            if _wc:
                mc_raw = filter_by_week_range(mc_raw, _wc, p_start, p_end, weeks)
                mc_raw = shorten_dates_in_df(mc_raw, _wc)
                if not med_count_c.empty:
                    mct = filter_by_week_range(med_count_c, "주차", p_start, p_end, weeks)
                    mct = shorten_dates_in_df(mct, "주차")
                    ct_map = dict(zip(mct["주차"], mct["값"].apply(safe_numeric)))
                    mc_raw["_bar"] = mc_raw[_wc].map(ct_map).fillna(mc_raw[_sum_col].apply(safe_numeric) if _sum_col else 0)
                    bar_col_use = "_bar"
                else:
                    bar_col_use = _sum_col
                if bar_col_use:
                    plot_bar_rate_dual(mc_raw, _wc, bar_col_use, "등록건수", "#66BB6A",
                                       _rc, "전체이용비중", "#FF6F00",
                                       "활성 이용자 복약 등록건수 + 전체이용비중", bar_unit="건")
        elif not med_count.empty:
            mf = filter_by_week_range(med_count, "주차", p_start, p_end, weeks)
            total = mf.groupby("주차")["값"].sum().reset_index()
            total.columns = ["주차", "등록건수합계"]
            total = shorten_dates_in_df(total, "주차")
            plot_bar_rate_dual(total, "주차", "등록건수합계", "등록건수", "#66BB6A",
                               None, None, None,
                               "활성 이용자 복약 등록건수", bar_unit="건")
        else:
            st.info("복약 등록건수 데이터가 없습니다.")

    with tab3:
        # 지자체별 비중 추이
        if not med_users.empty:
            mf = filter_by_week_range(med_users, "주차", p_start, p_end, weeks)
            reg_dict = {}
            if not reg.empty and "지자체명" in reg.columns and "가입완료" in reg.columns:
                for _, r in reg.iterrows():
                    reg_dict[str(r["지자체명"]).strip()] = safe_numeric(r.get("가입완료", 0))
            if reg_dict:
                mf_ratio = mf.copy()
                def _fuzzy(d, key, default=0):
                    if key in d: return d[key]
                    for k, v in d.items():
                        if key in k or k in key: return v
                    return default
                mf_ratio["가입완료"] = mf_ratio["지자체명"].map(lambda x: _fuzzy(reg_dict, x, 0))
                mf_ratio["비중"] = (mf_ratio["값"] / mf_ratio["가입완료"].replace(0, float("nan")) * 100).round(1).fillna(0)
                mf_ratio["값"] = mf_ratio["비중"]
                plot_municipality_lines(mf_ratio, "지자체별 복약 등록 비중 (가입자 대비 %)", metric_label="비중(%)")

    with tab4:
        if not med_users.empty:
            st.markdown("**복약 등록 회원수 원본**")
            st.dataframe(med_users, use_container_width=True, height=300)
        if not med_count.empty:
            st.markdown("**복약 등록건수 원본**")
            st.dataframe(med_count, use_container_width=True, height=300)


# ============================================================
# 🎮 맞고 & 게임
# ============================================================
elif page == "📊 3.안부체크율":
    st.markdown('<div class="section-header">📊 안부체크율</div>', unsafe_allow_html=True)
    selected_biz = biz_selector("안부체크율")
    st.divider()

    # ── 전체 안부체크율(OFF 제외) 추이 — gid=261480368 AB열 (전체일 때만 표시)
    cd_all = data.get("checkin_daily", pd.DataFrame())
    if selected_biz == "전체" and not cd_all.empty and "안부체크율" in cd_all.columns and "날짜" in cd_all.columns:
        cd_plot = cd_all[cd_all["안부체크율"].apply(safe_numeric) > 0].copy()
        # 최근 90일만 표시 (최신 날짜가 오른쪽에 꽉 차도록)
        cd_plot = cd_plot.sort_values("날짜").tail(90).reset_index(drop=True)
        if not cd_plot.empty:
            # x축 틱 수 제한: 최대 18개만 표시
            n_pts = len(cd_plot)
            tick_step = max(1, n_pts // 18)
            tick_vals = cd_plot["날짜"].tolist()[::tick_step]

            fig_ab = go.Figure()
            fig_ab.add_trace(go.Scatter(
                x=cd_plot["날짜"], y=cd_plot["안부체크율"].apply(safe_numeric),
                mode="lines+markers", name="안부체크율(OFF 제외)",
                line=dict(color="#2F5496", width=2.5),
                marker=dict(size=5),
                fill="tozeroy", fillcolor="rgba(47,84,150,0.07)",
                hovertemplate="<b>%{x}</b><br>안부체크율: %{y:.1f}%<extra>OFF 제외</extra>",
            ))
            fig_ab.update_layout(
                title=f"안부체크율 전체 추이 (OFF 제외, AB열 기준) — 최근 {n_pts}일",
                height=360, hovermode="x unified",
                xaxis=dict(
                    type="category", title="",
                    tickmode="array", tickvals=tick_vals,
                    tickangle=-45, tickfont=dict(size=11),
                    range=[-0.5, n_pts - 0.5],
                    automargin=True,
                ),
                yaxis=dict(title="안부체크율 (%)", range=[0, 100]),
                margin=dict(t=45, b=90, r=80),
            )
            st.plotly_chart(fig_ab, use_container_width=True)
            st.markdown("---")

    cr_check_direct = biz_filter_df(data.get("checkin_mun_check_direct", pd.DataFrame()), selected_biz)
    checkin_rate = biz_filter_df(data.get("checkin_municipality_rate", pd.DataFrame()), selected_biz)

    # biz 선택 시: 사업구분별 주차별 집계 안부체크율 차트
    if selected_biz != "전체" and not checkin_rate.empty and "안부체크율" in checkin_rate.columns:
        _cr_biz = checkin_rate.copy()
        _send_c = "안부체크발송" if "안부체크발송" in _cr_biz.columns else None
        _resp_c = "안부체크응답" if "안부체크응답" in _cr_biz.columns else None
        _off_c  = "off대상자"   if "off대상자"   in _cr_biz.columns else None

        if _send_c and _resp_c:
            for _c in [_send_c, _resp_c]:
                _cr_biz[_c] = _cr_biz[_c].apply(safe_numeric)
            if _off_c:
                _cr_biz[_off_c] = _cr_biz[_off_c].apply(safe_numeric).fillna(0)
            _agg_d = {_send_c: "sum", _resp_c: "sum"}
            if _off_c:
                _agg_d[_off_c] = "sum"
            _wbiz = _cr_biz.groupby("시작일").agg(_agg_d).reset_index().sort_values("시작일")
            _denom = (_wbiz[_send_c] - _wbiz.get(_off_c, 0)).replace(0, float("nan"))
            _wbiz["_rate"] = (_wbiz[_resp_c] / _denom * 100).round(1).fillna(0)
        else:
            _wbiz = _cr_biz.groupby("시작일")["안부체크율"].mean().reset_index().sort_values("시작일")
            _wbiz.columns = ["시작일", "_rate"]

        _wbiz = _wbiz[_wbiz["_rate"] > 0]
        if not _wbiz.empty:
            fig_biz_cr = go.Figure()
            fig_biz_cr.add_trace(go.Scatter(
                x=_wbiz["시작일"], y=_wbiz["_rate"],
                mode="lines+markers", name=f"{selected_biz} 안부체크율",
                line=dict(color="#2F5496", width=2.5), marker=dict(size=6),
                fill="tozeroy", fillcolor="rgba(47,84,150,0.09)",
                hovertemplate="<b>%{x}</b><br>안부체크율: %{y:.1f}%<extra></extra>",
            ))
            fig_biz_cr.update_layout(
                title=f"{selected_biz} 주차별 안부체크율 추이",
                height=340, hovermode="x unified",
                xaxis=dict(type="category", tickangle=-45, title=""),
                yaxis=dict(title="안부체크율 (%)", range=[0, 100]),
                margin=dict(t=40, b=70),
            )
            st.plotly_chart(fig_biz_cr, use_container_width=True)
            st.markdown("---")

    # 권역별 시계열 탭용 cr 구성 (old data source)
    cr = pd.DataFrame()
    regions = []
    if not checkin_rate.empty and "안부체크율" in checkin_rate.columns:
        cr = checkin_rate[checkin_rate["안부체크율"].notna()].copy()
        cr["권역"] = cr["지자체명"].map(DETAIL_REGION).fillna("기타")

        # 📅 날짜 기간 선택기 — 원본 날짜 전체 기준 (dedup 이전)
        all_dates = sorted(cr["시작일"].unique())
        default_idx = max(0, len(all_dates) - 16)  # 최근 16주 기본 표시

        with st.expander("📅 기간 설정 (펼쳐서 변경)", expanded=False):
            dc1, dc2 = st.columns(2)
            with dc1:
                cr_date_start = st.selectbox("시작일", all_dates, index=default_idx, key="cr_date_start")
            with dc2:
                start_idx = all_dates.index(cr_date_start) if cr_date_start in all_dates else 0
                end_options = all_dates[start_idx:]
                cr_date_end = st.selectbox("종료일", end_options, index=len(end_options)-1, key="cr_date_end")
            st.caption(f"선택 기간: {cr_date_start} ~ {cr_date_end}")

        cr = cr[(cr["시작일"] >= cr_date_start) & (cr["시작일"] <= cr_date_end)]
        cr["_week_key"] = cr["시작일"].apply(date_to_week_label)
        if "안부체크율_원본" in cr.columns:
            cr["_has_orig"] = cr["안부체크율_원본"].fillna(0).gt(0).astype(int)
        else:
            cr["_has_orig"] = 0
        cr = cr.sort_values(["_has_orig", "시작일"])
        cr = cr.drop_duplicates(subset=["_week_key", "지자체명"], keep="last")
        cr = cr.drop(columns=["_week_key", "_has_orig"])
        cr = cr.sort_values("시작일")
        cr = week_label_df(cr, "시작일")

        _reg_df = data.get("registration", pd.DataFrame())
        _active_from_reg = []
        if not _reg_df.empty and "지자체명" in _reg_df.columns:
            _active_from_reg = _reg_df["지자체명"].dropna().str.strip().tolist()

        if _active_from_reg:
            def _match_active_reg(name):
                name_n = str(name).replace(" ", "")
                for a in _active_from_reg:
                    a_n = str(a).replace(" ", "")
                    if name_n == a_n or name_n in a_n or a_n in name_n:
                        return True
                return False
            cr = cr[cr["지자체명"].apply(_match_active_reg)]

        regions = sorted(cr["권역"].unique())

    # 탭 구성 — 전체 비교는 항상 표시 (PH:QJ 직접 데이터)
    tab_labels = ["전체 비교"] + regions
    tabs = st.tabs(tab_labels)

    # 전체 비교 탭 — PH:QJ 직접 데이터 사용 (29개 지자체)
    with tabs[0]:
        if not cr_check_direct.empty and "안부체크율" in cr_check_direct.columns:
            latest_snap = cr_check_direct.copy()
            latest_snap["권역"] = latest_snap["지자체명"].map(DETAIL_REGION).fillna("기타")
            latest_snap = latest_snap.sort_values("안부체크율", ascending=True).copy()
            latest_snap["지자체명_표시"] = latest_snap["지자체명"].apply(_mun_label)
            st.markdown(f"**{selected_week} 기준 — 전체 {len(latest_snap)}개 지자체**")
            fig = px.bar(latest_snap, y="지자체명_표시", x="안부체크율", orientation="h",
                         color="권역", color_discrete_map=REGION_COLORS,
                         height=min(600, max(400, len(latest_snap) * 22)))
            fig.update_layout(
                title=f"지자체별 안부체크율 ({selected_week})",
                legend=LEGEND_BELOW, margin=dict(t=40, b=70),
                xaxis=dict(range=[0, 105]),
                yaxis=dict(tickfont=dict(size=11)),
            )
            fig.update_traces(
                texttemplate="%{x:.1f}%", textposition="outside",
                textfont=dict(size=12),
                hovertemplate="<b>%{y}</b><br>안부체크율: %{x:.1f}%<extra></extra>",
            )
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.info("안부체크율 데이터 없음 (🔄 데이터 새로고침 후 다시 시도)")

    # 권역별 탭 (시계열 데이터 있을 때만)
    if regions:
        for i, region in enumerate(regions):
            with tabs[i + 1]:
                region_data = cr[cr["권역"] == region].copy()
                mun_list = region_data["지자체명"].unique().tolist()
                st.markdown(f"**{region}** — {', '.join(mun_list)}")
                region_data["지자체명"] = region_data["지자체명"].apply(_mun_label)

                # 지자체별 고유 색상 매핑 (label 적용 후 기준)
                _region_color_map = {row: MUNICIPALITY_COLORS.get(
                    str(row).lstrip("★○ "), MUNICIPALITY_COLORS.get(str(row), "#888888")
                ) for row in region_data["지자체명"].unique()}
                fig = px.line(region_data, x="시작일", y="안부체크율", color="지자체명",
                              markers=True, color_discrete_map=_region_color_map)
                fig.update_layout(
                    title=f"{region} 안부체크율 추이", height=400,
                    hovermode="x unified",
                    xaxis=dict(
                        type="category", title="",
                        tickangle=-45, tickfont=dict(size=11),
                    ),
                    yaxis=dict(title="안부체크율 (%)", range=[0, 100]),
                    legend=LEGEND_BELOW, margin=dict(t=40, b=90),
                )
                fig.update_traces(hovertemplate="<b>%{x}</b><br>%{y:.1f}%<extra>%{fullData.name}</extra>")
                st.plotly_chart(fig, use_container_width=True)

                # 해당 권역 최신 바 차트
                dates = list(dict.fromkeys(region_data["시작일"].tolist()))  # 시간순 정렬 보존
                if dates:
                    latest = (region_data[region_data["시작일"] == dates[-1]]
                              .drop_duplicates(subset="지자체명", keep="last")
                              .sort_values("안부체크율", ascending=True)
                              .copy())
                    latest["지자체명"] = latest["지자체명"].apply(_mun_label)
                    fig2 = px.bar(latest, y="지자체명", x="안부체크율", orientation="h",
                                  color_discrete_sequence=[REGION_COLORS.get(region, "#666")],
                                  height=min(400, max(200, len(latest) * 26)))
                    fig2.update_layout(
                        title=f"{region} 최신 안부체크율",
                        margin=dict(t=40, b=10),
                        xaxis=dict(range=[0, 105]),
                        yaxis=dict(tickfont=dict(size=11)),
                    )
                    fig2.update_traces(
                        texttemplate="%{x:.1f}%", textposition="outside",
                        textfont=dict(size=12),
                    )
                    st.plotly_chart(fig2, use_container_width=True)

    # 추가 지표 (시계열 데이터 있을 때만)
    if not checkin_rate.empty:
        extra_metrics = [m for m in ["안부확인율", "48미확인율", "안부콜응답률"] if m in checkin_rate.columns]
        if extra_metrics:
            st.markdown("---")
            selected_extra = st.selectbox("추가 지표 보기", extra_metrics)
            er = checkin_rate[checkin_rate[selected_extra].notna()].copy()
            er = er.sort_values("시작일")
            er = week_label_df(er, "시작일")
            if not er.empty:
                er["권역"] = er["지자체명"].map(DETAIL_REGION).fillna("기타")
                sel_region = st.radio("권역 선택", ["전체"] + sorted(er["권역"].unique().tolist()), horizontal=True, key="extra_region")
                if sel_region != "전체":
                    er = er[er["권역"] == sel_region]
                er = er.copy()
                er["지자체명"] = er["지자체명"].apply(_mun_label)
                fig3 = px.line(er, x="시작일", y=selected_extra, color="지자체명", markers=True)
                fig3.update_layout(
                    title=f"지자체별 {selected_extra} 추이", height=400,
                    hovermode="x unified", xaxis=dict(type="category"),
                    yaxis=dict(title="%"),
                    legend=LEGEND_BELOW_LARGE,
                )
                st.plotly_chart(fig3, use_container_width=True)


# ============================================================
# 🔄 4.안부체크 변경
# ============================================================
elif page == "🔄 4.안부체크 변경(베이직)":
    st.markdown('<div class="section-header">🔄 안부체크 변경건</div>', unsafe_allow_html=True)
    selected_biz = biz_selector("안부변경베이직")
    st.divider()
    st.markdown("""
    <div class="insight-box">
    개인정보가 포함된 데이터입니다.<br>
    데이터 소스: <a href="https://docs.google.com/spreadsheets/d/15UZ9dZjYdD24PdWoSvrFWpQCM-T0vhc_yy9wrMunSNc/edit?gid=851523453" target="_blank">지자체별 안부체크 변경건 시트</a>
    </div>
    """, unsafe_allow_html=True)

    # Google Sheets에서 안부체크횟수 시트 (gid=851523453) 직접 가져오기
    checkin_change_raw = sheets.get("안부체크횟수", pd.DataFrame())

    if not checkin_change_raw.empty:
        df = checkin_change_raw.copy()

        # 시작일 컬럼 찾기
        date_col = None
        for c in df.columns:
            cl = str(c).replace("\n", "").strip()
            if "시작일" in cl:
                date_col = c
                break
        if date_col is None:
            date_col = df.columns[0]

        # 날짜 목록 추출
        all_change_dates = sorted([str(d).strip() for d in df[date_col].dropna().unique() if str(d).strip() and str(d).strip() != "nan"])

        tab_basic, tab_input = st.tabs(["📊 베이직 (안부상태 변경)", "✏ 데이터 입력"])

        with tab_basic:
            # === 베이직: 안부상태 변경건 ===

            # 📅 기간 선택 (기본: 25-52주차 = 2025-12-28 근사)
            basic_default = "2025-12-28"
            basic_idx = 0
            for i, d in enumerate(all_change_dates):
                if d >= basic_default:
                    basic_idx = i
                    break
            with st.expander("📅 베이직 기간 설정 (펼쳐서 변경)", expanded=False):
                bc1, bc2 = st.columns(2)
                with bc1:
                    basic_start = st.selectbox("시작일", all_change_dates, index=basic_idx, key="basic_change_start")
                with bc2:
                    bs_idx = all_change_dates.index(basic_start) if basic_start in all_change_dates else 0
                    basic_end_opts = all_change_dates[bs_idx:]
                    basic_end = st.selectbox("종료일", basic_end_opts, index=len(basic_end_opts)-1, key="basic_change_end")
                st.caption(f"베이직 기간: {basic_start} ~ {basic_end}")

            # 기간 필터 적용
            df_basic = df[(df[date_col].astype(str) >= basic_start) & (df[date_col].astype(str) <= basic_end)].copy()
            # datetime 파싱으로 정확한 날짜 정렬
            df_basic["_sort"] = pd.to_datetime(df_basic[date_col].astype(str), errors="coerce")
            df_basic = df_basic.sort_values("_sort").drop(columns=["_sort"])
            df_basic = shorten_dates_in_df(df_basic, date_col)

            # 총합, 총 안부상태변경률 컬럼 찾기
            total_col = None
            total_rate_col = None
            mun_change_cols = []  # 지자체별 변경건 컬럼
            mun_rate_cols = []    # 지자체별 변경률 컬럼

            for c in df_basic.columns:
                cl = str(c).replace("\n", "").strip()
                if cl == "총합":
                    total_col = c
                elif "총 안부상태변경률" in cl or cl == "총 안부상태변경률":
                    total_rate_col = c
                elif "안부상태변경률" in cl and "총" not in cl:
                    mun_rate_cols.append(c)
                elif any(kw in cl for kw in MUNICIPALITY_KEYWORDS) and "KTT" not in cl and "관제" not in cl and "발송" not in cl and "변경률" not in cl:
                    mun_change_cols.append(c)

            # 사업구분 필터: 지자체 컬럼만 선별
            mun_change_cols = biz_filter_wide_cols(mun_change_cols, selected_biz)
            mun_rate_cols   = biz_filter_wide_cols(mun_rate_cols, selected_biz)

            # 1. 총 안부상태 변경건 + 변경률 추이
            if total_col:
                df_basic[total_col] = df_basic[total_col].apply(safe_numeric)
                fig = make_subplots(specs=[[{"secondary_y": True}]])
                fig.add_trace(go.Bar(
                    x=df_basic[date_col], y=df_basic[total_col], name="총 변경건",
                    marker_color="#FF6F00", opacity=0.7,
                    hovertemplate="%{y:,.0f}건<extra>총 변경건</extra>"
                ), secondary_y=False)

                if total_rate_col:
                    df_basic[total_rate_col] = df_basic[total_rate_col].apply(safe_numeric)
                    fig.add_trace(go.Scatter(
                        x=df_basic[date_col], y=df_basic[total_rate_col], name="변경률",
                        mode="lines+markers", line=dict(color="#D32F2F", width=2),
                        hovertemplate="%{y:.1f}%<extra>변경률</extra>"
                    ), secondary_y=True)

                fig.update_layout(
                    title="총 안부상태 변경건 및 변경률", height=400,
                    xaxis=dict(type="category"), hovermode="x unified",
                    margin=dict(t=40, b=60, l=40, r=40),
                    legend=dict(orientation="h", yanchor="top", y=-0.15, xanchor="center", x=0.5),
                )
                fig.update_yaxes(title_text="변경건", secondary_y=False)
                fig.update_yaxes(title_text="%", secondary_y=True)
                st.plotly_chart(fig, use_container_width=True)

            # 2. 지자체별 안부상태 변경건 추이
            if mun_change_cols:
                rows = []
                for _, row in df_basic.iterrows():
                    d = str(row.get(date_col, "")).strip()
                    if not d or d == "nan":
                        continue
                    for mc in mun_change_cols:
                        mun = extract_municipality_name(mc)
                        val = safe_numeric(row.get(mc, 0))
                        if val > 0:
                            rows.append({"날짜": d, "지자체명": mun, "변경건": val})
                if rows:
                    mun_df = pd.DataFrame(rows)
                    # 날짜를 datetime으로 파싱하여 정확히 정렬
                    mun_df["_sort"] = pd.to_datetime("20" + mun_df["날짜"], errors="coerce")
                    mun_df = mun_df.sort_values("_sort").drop(columns=["_sort"])
                    sorted_dates = mun_df["날짜"].unique().tolist()
                    mun_df["지자체명"] = mun_df["지자체명"].apply(_mun_label)
                    fig2 = px.line(mun_df, x="날짜", y="변경건", color="지자체명", markers=True)
                    fig2.update_layout(
                        title="지자체별 안부상태 변경건 추이", height=400,
                        xaxis=dict(type="category", categoryorder="array", categoryarray=sorted_dates),
                        hovermode="x unified",
                        margin=dict(t=40, b=60, l=40, r=60),
                        legend=LEGEND_BELOW_LARGE,
                    )
                    st.plotly_chart(fig2, use_container_width=True)


        with tab_input:
            # 수동 입력 폼 (개인정보 데이터)
            st.markdown('<div class="section-header">데이터 직접 입력</div>', unsafe_allow_html=True)
            st.caption("개인정보가 포함된 안부상태 변경 데이터를 직접 입력합니다.")

            with st.form("checkin_change_form", clear_on_submit=True):
                fc1, fc2 = st.columns(2)
                with fc1:
                    change_date = st.date_input("날짜")
                    change_agency = st.selectbox("지자체", [
                        "경기도청", "서초구청", "진천군청", "음성군청", "강북구청",
                        "금정구청", "증평군청", "포천시청", "경남사회서비스원",
                        "강릉시청", "강원사회서비스원", "충북사회서비스원",
                        "독거노인지원종합센터", "희망나래장애인복지관", "홍천군청",
                        "충남사회서비스원", "삼척시청", "마포구청", "광진구청",
                    ])
                with fc2:
                    change_count = st.number_input("변경건수", min_value=0, value=0)
                    change_memo = st.text_input("메모", placeholder="상태 변경 사유")

                if st.form_submit_button("💾 저장", type="primary", use_container_width=True):
                    from local_db import get_connection as _gc
                    conn = _gc()
                    try:
                        conn.execute("""
                        INSERT INTO raw_generic (data_type, date, agency_name, raw_json)
                        VALUES (?, ?, ?, ?)
                        ON CONFLICT(data_type, date, agency_name) DO UPDATE SET raw_json=excluded.raw_json
                        """, ("안부체크변경", str(change_date), change_agency,
                              f'{{"count": {change_count}, "memo": "{change_memo}"}}'))
                        conn.commit()
                        st.success(f"✅ {change_agency} {change_date} 변경건 {change_count}건 저장!")
                    except Exception as e:
                        st.error(f"저장 실패: {e}")
                    conn.close()

    else:
        st.info("안부체크 변경건 데이터가 없습니다.")


# ============================================================
# 🛡 5.관제 현황
# ============================================================
elif page == "🛡 5.안부체크 변경(세이프)":
    st.markdown('<div class="section-header">🛡 안부체크 변경(세이프) — KT 관제</div>', unsafe_allow_html=True)
    selected_biz = biz_selector("안부변경세이프")
    st.divider()

    # ── KT 관제 차트 (안부체크횟수 시트에서) ─────────────────────────
    _safe_raw = sheets.get("안부체크횟수", pd.DataFrame())
    if not _safe_raw.empty:
        _sf = _safe_raw.copy()
        _sf_date_col = None
        for _c in _sf.columns:
            if "시작일" in str(_c).replace("\n", "").strip():
                _sf_date_col = _c
                break
        if _sf_date_col is None:
            _sf_date_col = _sf.columns[0]

        _all_safe_dates = sorted([str(d).strip() for d in _sf[_sf_date_col].dropna().unique()
                                   if str(d).strip() and str(d).strip() != "nan"])

        _safe_default = "2026-03-01"
        _safe_idx = next((i for i, d in enumerate(_all_safe_dates) if d >= _safe_default), 0)
        with st.expander("📅 기간 설정 (펼쳐서 변경)", expanded=False):
            _sc1, _sc2 = st.columns(2)
            with _sc1:
                _safe_start = st.selectbox("시작일", _all_safe_dates, index=_safe_idx, key="p5_safe_start")
            with _sc2:
                _ss_idx = _all_safe_dates.index(_safe_start) if _safe_start in _all_safe_dates else 0
                _safe_end_opts = _all_safe_dates[_ss_idx:]
                _safe_end = st.selectbox("종료일", _safe_end_opts, index=len(_safe_end_opts)-1, key="p5_safe_end")
            st.caption(f"기간: {_safe_start} ~ {_safe_end}")

        _df_safe = _sf[(_sf[_sf_date_col].astype(str) >= _safe_start) &
                       (_sf[_sf_date_col].astype(str) <= _safe_end)].copy()
        _df_safe = shorten_dates_in_df(_df_safe, _sf_date_col)

        kt_total_col = kt_send_col = kt_rate_col = kt_mgmt_rate_col = kt_disp_rate_col = None
        kt_mun_rate_cols = []
        for _c in _df_safe.columns:
            _cl_raw = str(_c).replace("\n", "").strip()
            _cl_low = _cl_raw.replace(" ", "").lower()
            if _cl_raw == "KT 관제 수":
                kt_total_col = _c
            elif _cl_raw == "전체 발송수":
                kt_send_col = _c
            elif _cl_raw == "KT 관제 대응률":
                kt_rate_col = _c
            elif "KT관제 대응률" in _cl_raw and _cl_raw != "KT 관제 대응률":
                kt_mun_rate_cols.append(_c)
            if "kt관제율" in _cl_low:
                kt_mgmt_rate_col = _c
            elif "kt출동율" in _cl_low or "kt출동률" in _cl_low:
                kt_disp_rate_col = _c

        # 사업구분 필터: 지자체별 관제율 컬럼 선별
        kt_mun_rate_cols = biz_filter_wide_cols(kt_mun_rate_cols, selected_biz)

        # KT 관제율 · 출동율 추이
        if kt_mgmt_rate_col or kt_disp_rate_col:
            st.markdown('<div class="section-header">📡 KT 관제율 · 출동율</div>', unsafe_allow_html=True)
            _fig_kt_rate = go.Figure()
            if kt_mgmt_rate_col:
                _df_safe[kt_mgmt_rate_col] = _df_safe[kt_mgmt_rate_col].apply(safe_numeric)
                _kt_mgmt = _df_safe[_df_safe[kt_mgmt_rate_col] > 0]
                _fig_kt_rate.add_trace(go.Scatter(
                    x=_kt_mgmt[_sf_date_col], y=_kt_mgmt[kt_mgmt_rate_col],
                    name="KT 관제율", mode="lines+markers",
                    line=dict(color="#1565C0", width=2.5), marker=dict(size=7),
                    hovertemplate="<b>%{x}</b><br>KT 관제율: <b>%{y:.1f}%</b><extra></extra>"
                ))
            if kt_disp_rate_col:
                _df_safe[kt_disp_rate_col] = _df_safe[kt_disp_rate_col].apply(safe_numeric)
                _kt_disp = _df_safe[_df_safe[kt_disp_rate_col] > 0]
                if not _kt_disp.empty:
                    _fig_kt_rate.add_trace(go.Scatter(
                        x=_kt_disp[_sf_date_col], y=_kt_disp[kt_disp_rate_col],
                        name="KT 출동율", mode="lines+markers",
                        line=dict(color="#E65100", width=2.5, dash="dot"), marker=dict(size=7, symbol="diamond"),
                        hovertemplate="<b>%{x}</b><br>KT 출동율: <b>%{y:.1f}%</b><extra></extra>"
                    ))
            _fig_kt_rate.update_layout(
                title="KT 관제율 · 출동율 추이", height=380,
                xaxis=dict(type="category", title=""), yaxis=dict(title="%", ticksuffix="%"),
                hovermode="x unified", margin=dict(t=40, b=60, l=50, r=20),
                legend=dict(orientation="h", yanchor="top", y=-0.18, xanchor="center", x=0.5),
            )
            st.plotly_chart(_fig_kt_rate, use_container_width=True)

        # KT 관제 수 + 전체 발송수 + 대응률
        if kt_total_col and kt_send_col:
            _df_safe[kt_total_col] = _df_safe[kt_total_col].apply(safe_numeric)
            _df_safe[kt_send_col] = _df_safe[kt_send_col].apply(safe_numeric)
            _fig3 = make_subplots(specs=[[{"secondary_y": True}]])
            _fig3.add_trace(go.Bar(x=_df_safe[_sf_date_col], y=_df_safe[kt_send_col], name="전체 발송수",
                                   marker_color="#B0BEC5", opacity=0.6,
                                   hovertemplate="%{y:,.0f}건<extra>전체 발송수</extra>"), secondary_y=False)
            _fig3.add_trace(go.Bar(x=_df_safe[_sf_date_col], y=_df_safe[kt_total_col], name="KT 관제 수",
                                   marker_color="#1565C0",
                                   hovertemplate="%{y:,.0f}건<extra>KT 관제 수</extra>"), secondary_y=False)
            if kt_rate_col:
                _df_safe[kt_rate_col] = _df_safe[kt_rate_col].apply(safe_numeric)
                _fig3.add_trace(go.Scatter(x=_df_safe[_sf_date_col], y=_df_safe[kt_rate_col], name="KT 관제 대응률",
                                           mode="lines+markers", line=dict(color="#D32F2F", width=2),
                                           hovertemplate="%{y:.1f}%<extra>KT 관제 대응률</extra>"), secondary_y=True)
            _fig3.update_layout(title="세이프 - KT 관제 현황", height=400,
                                xaxis=dict(type="category"), hovermode="x unified", barmode="group",
                                margin=dict(t=40, b=60, l=40, r=40),
                                legend=dict(orientation="h", yanchor="top", y=-0.15, xanchor="center", x=0.5))
            _fig3.update_yaxes(title_text="건수", secondary_y=False)
            _fig3.update_yaxes(title_text="%", secondary_y=True)
            st.plotly_chart(_fig3, use_container_width=True)

        # 지자체별 KT 관제/출동 수
        _ktt_ctrl_cols = [_c for _c in _df_safe.columns if str(_c).replace("\n","").strip().startswith("KTT_관제_")]
        _ktt_disp_cols = [_c for _c in _df_safe.columns if str(_c).replace("\n","").strip().startswith("KTT_출동_")]
        _EXCL_MUNS = {"마포구청", "광진구청"}
        if _ktt_ctrl_cols:
            _ctrl_rows, _disp_rows = [], []
            for _, _row in _df_safe.iterrows():
                _d = str(_row.get(_sf_date_col, "")).strip()
                if not _d or _d == "nan":
                    continue
                for _cc in _ktt_ctrl_cols:
                    _mun = str(_cc).replace("\n","").strip().replace("KTT_관제_","").strip()
                    if _mun not in _EXCL_MUNS:
                        _ctrl_rows.append({"날짜": _d, "지자체명": _mun, "관제수": safe_numeric(_row.get(_cc, 0))})
                for _dc in _ktt_disp_cols:
                    _mun = str(_dc).replace("\n","").strip().replace("KTT_출동_","").strip()
                    if _mun not in _EXCL_MUNS:
                        _disp_rows.append({"날짜": _d, "지자체명": _mun, "출동수": safe_numeric(_row.get(_dc, 0))})
            if _ctrl_rows:
                _ctrl_muns = {r["지자체명"] for r in _ctrl_rows}
                _disp_muns = {r["지자체명"] for r in _disp_rows}
                for _mun in _ctrl_muns - _disp_muns:
                    for _d in sorted({r["날짜"] for r in _ctrl_rows}):
                        _disp_rows.append({"날짜": _d, "지자체명": _mun, "출동수": 0})

            st.markdown('<div class="section-header">지자체별 KT 관제 수</div>', unsafe_allow_html=True)
            if _ctrl_rows:
                _ctrl_df = pd.DataFrame(_ctrl_rows)
                _mun_total = _ctrl_df.groupby("지자체명")["관제수"].sum().sort_values(ascending=False)
                _mun_order = _mun_total.index.tolist()
                _ctrl_color_map = {_mun_label(m): MUNICIPALITY_COLORS.get(m, "#888888") for m in _mun_order}
                _ctrl_df["지자체명"] = _ctrl_df["지자체명"].apply(_mun_label)
                _mun_order_lbl = [_mun_label(m) for m in _mun_order]
                _fig4 = px.bar(_ctrl_df, x="날짜", y="관제수", color="지자체명", barmode="stack", height=500,
                               color_discrete_map=_ctrl_color_map,
                               category_orders={"지자체명": _mun_order_lbl[::-1]})
                _fig4.update_layout(title="지자체별 주차별 KT 관제 수", xaxis=dict(type="category", title=""),
                                    yaxis=dict(title="관제 수 (건)"), hovermode="closest",
                                    legend=dict(orientation="h", yanchor="top", y=-0.18, xanchor="center", x=0.5,
                                                font=dict(size=10), tracegroupgap=2),
                                    margin=dict(t=40, b=160, r=20))
                _fig4.update_traces(hovertemplate="<b>%{fullData.name}</b><br>날짜: %{x}<br>관제 수: <b>%{y:,.0f}건</b><extra></extra>")
                st.plotly_chart(_fig4, use_container_width=True)
                _latest_d = _ctrl_df["날짜"].max()
                _latest_ctrl = _ctrl_df[_ctrl_df["날짜"] == _latest_d].sort_values("관제수", ascending=True)
                if not _latest_ctrl.empty:
                    _fig4b = px.bar(_latest_ctrl, y="지자체명", x="관제수", orientation="h",
                                    color="관제수", color_continuous_scale="Blues",
                                    height=max(280, len(_latest_ctrl)*28), text="관제수")
                    _fig4b.update_layout(title=f"최신 주차({_latest_d}) 지자체별 KT 관제 수",
                                         xaxis=dict(title="관제 수 (건)"), yaxis=dict(title=""),
                                         coloraxis_showscale=False, margin=dict(t=40, b=10, l=10, r=60))
                    _fig4b.update_traces(textposition="outside", texttemplate="%{x:,.0f}건",
                                         hovertemplate="<b>%{y}</b><br>관제 수: <b>%{x:,.0f}건</b><extra></extra>")
                    st.plotly_chart(_fig4b, use_container_width=True)

            st.markdown('<div class="section-header">지자체별 KT 출동 수</div>', unsafe_allow_html=True)
            if _disp_rows:
                _disp_df = pd.DataFrame(_disp_rows)
                _mun_total_d = _disp_df.groupby("지자체명")["출동수"].sum().sort_values(ascending=False)
                if not _mun_total_d.empty:
                    _mun_order_d = _mun_total_d.index.tolist()
                    _disp_color_map = {_mun_label(m): MUNICIPALITY_COLORS.get(m, "#888888") for m in _mun_order_d}
                    _disp_df["지자체명"] = _disp_df["지자체명"].apply(_mun_label)
                    _mun_order_d_lbl = [_mun_label(m) for m in _mun_order_d]
                    _fig5 = px.bar(_disp_df, x="날짜", y="출동수", color="지자체명", barmode="stack", height=500,
                                   color_discrete_map=_disp_color_map,
                                   category_orders={"지자체명": _mun_order_d_lbl[::-1]})
                    _fig5.update_layout(title="지자체별 주차별 KT 출동 수", xaxis=dict(type="category", title=""),
                                        yaxis=dict(title="출동 수 (건)"), hovermode="closest",
                                        legend=dict(orientation="h", yanchor="top", y=-0.18, xanchor="center", x=0.5,
                                                    font=dict(size=10), tracegroupgap=2),
                                        margin=dict(t=40, b=160, r=20))
                    _fig5.update_traces(hovertemplate="<b>%{fullData.name}</b><br>날짜: %{x}<br>출동 수: <b>%{y:,.0f}건</b><extra></extra>")
                    st.plotly_chart(_fig5, use_container_width=True)
                    _latest_d2 = _disp_df["날짜"].max()
                    _latest_disp = _disp_df[_disp_df["날짜"] == _latest_d2].sort_values("출동수", ascending=True)
                    if not _latest_disp.empty:
                        _fig5b = px.bar(_latest_disp, y="지자체명", x="출동수", orientation="h",
                                        color="출동수", color_continuous_scale="Oranges",
                                        height=max(280, len(_latest_disp)*28), text="출동수")
                        _fig5b.update_layout(title=f"최신 주차({_latest_d2}) 지자체별 KT 출동 수",
                                             xaxis=dict(title="출동 수 (건)"), yaxis=dict(title=""),
                                             coloraxis_showscale=False, margin=dict(t=40, b=10, l=10, r=60))
                        _fig5b.update_traces(textposition="outside", texttemplate="%{x:,.0f}건",
                                             hovertemplate="<b>%{y}</b><br>출동 수: <b>%{x:,.0f}건</b><extra></extra>")
                        st.plotly_chart(_fig5b, use_container_width=True)

    st.markdown("---")
    st.markdown('<div class="section-header">📊 관제 현황 대시보드</div>', unsafe_allow_html=True)
    import streamlit.components.v1 as _components
    from pathlib import Path as _Path
    import datetime as _dt
    # update.ps1 이 빌드 후 앱 폴더로 복사해 둔 사본을 우선 사용, 없으면 원본 경로
    _html = _Path(__file__).parent / "gwanje_dashboard.html"
    if not _html.exists():
        _html = _Path(r"C:\Users\NHN\_bmad\waplat-gwanje\dashboard.html")
    if _html.exists():
        _mt = _dt.datetime.fromtimestamp(_html.stat().st_mtime).strftime("%Y-%m-%d %H:%M")
        st.caption(f"📄 출처: WAPLAT 관제 시트 · 최종 갱신 {_mt} · 주간 자동화: update.ps1")
        _components.html(_html.read_text(encoding="utf-8"), height=1200, scrolling=True)
    else:
        st.warning("관제 현황 데이터가 아직 생성되지 않았습니다. `_bmad/waplat-gwanje/update.ps1`을 먼저 실행하세요.")


# ============================================================
# 🎮 12.맞고(와플랫+게스트)
# ============================================================
elif page == "🎮 12.맞고(와플랫+게스트)":
    st.markdown('<div class="section-header">🎮 맞고 (와플랫+게스트)</div>', unsafe_allow_html=True)
    p_start, p_end = page_week_range_selector("matgo_all", weeks)
    st.divider()

    # 집계형 시트에서 직접 가져오기
    matgo_all_raw = sheets.get("맞고와플게스트", pd.DataFrame())
    if not matgo_all_raw.empty:
        matgo_all = matgo_all_raw.copy()
        # 주차 컬럼 찾기
        week_col = None
        for c in matgo_all.columns:
            if "주차" in str(c):
                week_col = c
                break
        num_cols = [c for c in matgo_all.columns if c not in [week_col, "시작일"] and c is not None]
        for c in num_cols:
            matgo_all[c] = matgo_all[c].apply(safe_numeric)

        if week_col:
            matgo_all = filter_by_week_range(matgo_all, week_col, p_start, p_end, weeks)
            matgo_all = shorten_dates_in_df(matgo_all, week_col)

        tab1, tab2, tab3 = st.tabs(["이용자수·플레이판수", "플레이시간", "상세 데이터"])
        with tab1:
            if week_col:
                # 이용자수 + 1인당 플레이판수 듀얼 차트
                user_col = next((c for c in num_cols if "이용자" in str(c)), None)
                play_col = next((c for c in num_cols if "플레이판수" in str(c) and "1인당" not in str(c)), None)
                per_play_col = next((c for c in num_cols if "1인당" in str(c) and "플레이판" in str(c)), None)

                if user_col and play_col:
                    # 이용자수(막대, 좌) + 플레이판수(꺾은선, 우) 듀얼 Y축
                    fig = make_subplots(specs=[[{"secondary_y": True}]])
                    fig.add_trace(go.Bar(
                        x=matgo_all[week_col], y=matgo_all[user_col], name="이용자수",
                        marker_color="#FF6F00", opacity=0.85,
                        text=matgo_all[user_col].apply(lambda v: f"{int(v):,}" if pd.notna(v) else ""),
                        textposition="outside", textfont=dict(size=9),
                        hovertemplate="%{y:,}명<extra>이용자수</extra>"
                    ), secondary_y=False)
                    fig.add_trace(go.Scatter(
                        x=matgo_all[week_col], y=matgo_all[play_col], name="플레이판수",
                        mode="lines+markers+text",
                        line=dict(color="#FFB74D", width=2),
                        text=matgo_all[play_col].apply(lambda v: f"{int(v):,}" if pd.notna(v) else ""),
                        textposition="top center", textfont=dict(size=9, color="#E65100"),
                        hovertemplate="%{y:,}판<extra>플레이판수</extra>"
                    ), secondary_y=True)
                    if per_play_col:
                        fig.add_trace(go.Scatter(
                            x=matgo_all[week_col], y=matgo_all[per_play_col],
                            name="1인당 플레이판수", mode="lines+markers",
                            line=dict(color="#D32F2F", width=2, dash="dot"),
                            hovertemplate="%{y:.1f}판<extra>1인당</extra>"
                        ), secondary_y=True)
                    fig.update_layout(title="맞고(와플랫+게스트) 이용 현황", height=420,
                                      hovermode="x unified",
                                      xaxis=dict(type="category"),
                                      legend=LEGEND_BELOW, margin=dict(t=40, b=70), bargap=0.3)
                    fig.update_yaxes(title_text="이용자수 (명)", secondary_y=False)
                    fig.update_yaxes(title_text="플레이판수 (판)", secondary_y=True, showgrid=False)
                    st.plotly_chart(fig, use_container_width=True)

        with tab2:
            if week_col:
                # "수정" 버전 우선 사용 (시:분:초 형식이 아닌 숫자 형태)
                time_col = next((c for c in num_cols if "수정 플레이시간" in str(c) and "1인당" not in str(c)), None)
                if time_col is None:
                    time_col = next((c for c in num_cols if "플레이시간" in str(c) and "게스트" not in str(c) and "합" not in str(c) and "1인당" not in str(c) and "수정" not in str(c)), None)
                per_time_col = next((c for c in num_cols if "수정 1인당" in str(c) and "플레이시간" in str(c)), None)
                if per_time_col is None:
                    per_time_col = next((c for c in num_cols if "1인당" in str(c) and "플레이시간" in str(c) and "수정" not in str(c)), None)

                if time_col:
                    fig = make_subplots(specs=[[{"secondary_y": True}]])
                    fig.add_trace(go.Bar(x=matgo_all[week_col], y=matgo_all[time_col], name="플레이시간",
                                         marker_color="#4E342E",
                                         hovertemplate="%{y:,.0f}시간<extra>전체</extra>"), secondary_y=False)
                    if per_time_col:
                        fig.add_trace(go.Scatter(x=matgo_all[week_col], y=matgo_all[per_time_col],
                                                  name="1인당 플레이시간", mode="lines+markers",
                                                  line=dict(color="#E91E63", width=2),
                                                  hovertemplate="%{y:.1f}시간<extra>1인당</extra>"), secondary_y=True)
                    fig.update_layout(title="맞고(와플랫+게스트) 플레이 시간", height=400,
                                      hovermode="x unified", xaxis=dict(type="category"),
                                      legend=LEGEND_BELOW, margin=dict(t=40, b=70))
                    st.plotly_chart(fig, use_container_width=True)

        with tab3:
            st.dataframe(matgo_all, use_container_width=True, height=400)
    else:
        st.info("맞고(와플랫+게스트) 데이터가 없습니다.")


# ============================================================
# 🃏 11.맞고(와플랫)
# ============================================================
elif page == "🃏 11.맞고(와플랫)":
    st.markdown('<div class="section-header">🃏 맞고 (와플랫)</div>', unsafe_allow_html=True)
    p_start, p_end = page_week_range_selector("matgo_waplat", weeks)
    selected_biz = biz_selector("맞고와플랫")
    st.divider()

    tab1, tab2, tab3 = st.tabs(["이용자수", "플레이 판수", "플레이 시간"])

    with tab1:
        matgo_user_raw = sheets.get("맞고이용자", pd.DataFrame())
        df = data.get("weekly_맞고이용자", pd.DataFrame())
        if not matgo_user_raw.empty:
            mu = matgo_user_raw.copy()
            _wc, _sum_col, _rc = None, None, None
            for c in mu.columns:
                cl = str(c).replace("\n", "").strip()
                if "주차" in cl and _wc is None: _wc = c
                elif ("이용자합계" in cl or ("합계" in cl and "이용자" in cl)) and _sum_col is None: _sum_col = c
                elif ("전체이용비중" in cl or "이용비중" in cl) and _rc is None: _rc = c
            if _wc:
                mu = filter_by_week_range(mu, _wc, p_start, p_end, weeks)
                mu = shorten_dates_in_df(mu, _wc)
                if _sum_col:
                    _biz_cnt, _ = biz_agg_raw(mu, selected_biz, _wc)
                    if _biz_cnt is not None:
                        mu = mu.copy(); mu["_bar"] = _biz_cnt; _sum_col = "_bar"
                        # 시트 내장 이용비중 대신 실제 가입완료 인원으로 직접 계산
                        _biz_completed = biz_filter_df(data.get("registration", pd.DataFrame()), selected_biz)["가입완료"].apply(safe_numeric).sum()
                        if _biz_completed > 0:
                            mu["_rc"] = (_biz_cnt / _biz_completed * 100).round(1); _rc = "_rc"
                    plot_bar_rate_dual(mu, _wc, _sum_col, "이용자수", "#42A5F5",
                                       _rc, "이용비중" if selected_biz != "전체" else "전체이용비중", "#FF6F00",
                                       "맞고(와플랫) 이용자수 + 이용비중")
            if not df.empty:
                dff = biz_filter_df(filter_by_week_range(df, "주차", p_start, p_end, weeks), selected_biz)
                plot_municipality_lines(dff, "지자체별 맞고 이용자수", metric_label="이용자수")
            # ── 지자체별 이용자비중 추이 (AI~BK열) ──────────────────────────
            st.markdown("---")
            mrt_matgo = extract_mun_ratio_trend(matgo_user_raw)
            if not mrt_matgo.empty:
                mrt_matgo = biz_filter_df(filter_by_week_range(mrt_matgo, "주차", p_start, p_end, weeks), selected_biz)
                _active_m = mrt_matgo.groupby("지자체명")["값"].sum()
                _active_m = _active_m[_active_m > 0].index.tolist()
                mrt_matgo = mrt_matgo[mrt_matgo["지자체명"].isin(_active_m)]
                if not mrt_matgo.empty:
                    plot_municipality_lines(mrt_matgo, "지자체별 맞고(와플랫) 이용자비중 추이 (%)", metric_label="이용자비중(%)")
        else:
            st.info("데이터 없음")

    with tab2:
        matgo_play_raw = sheets.get("맞고플레이판수", pd.DataFrame())
        df = data.get("weekly_맞고플레이판수", pd.DataFrame())
        if not matgo_play_raw.empty:
            mp = matgo_play_raw.copy()
            _wc, _sum_col, _awc = None, None, None
            for c in mp.columns:
                cl = str(c).replace("\n", "").strip()
                if "주차" in cl and _wc is None: _wc = c
                elif "합계" in cl and _sum_col is None: _sum_col = c
                elif "1인" in cl and "주평균" in cl and _awc is None: _awc = c
            if _wc:
                mp = filter_by_week_range(mp, _wc, p_start, p_end, weeks)
                mp = shorten_dates_in_df(mp, _wc)
                if _sum_col:
                    _biz_cnt, _ = biz_agg_raw(mp, selected_biz, _wc)
                    if _biz_cnt is not None:
                        mp = mp.copy(); mp["_bar"] = _biz_cnt; _sum_col = "_bar"; _awc = None
                    plot_bar_rate_dual(mp, _wc, _sum_col, "플레이판수", "#42A5F5",
                                       _awc, "1인 주평균", "#455A64",
                                       "맞고(와플랫) 플레이판수 + 1인 주평균", bar_unit="판", line_unit="판")
            if not df.empty:
                dff = biz_filter_df(filter_by_week_range(df, "주차", p_start, p_end, weeks), selected_biz)
                plot_municipality_lines(dff, "지자체별 플레이 판수", metric_label="판수")
        else:
            st.info("데이터 없음")

    with tab3:
        df = data.get("weekly_맞고플레이시간", pd.DataFrame())
        if not df.empty:
            dff = biz_filter_df(filter_by_week_range(df, "주차", p_start, p_end, weeks), selected_biz)
            if selected_biz == "전체":
                total = dff.pipe(weekly_total)
                total = total.rename(columns={"값": "플레이시간합계"})
                plot_weekly_series(total, "주차", "플레이시간합계", "맞고(와플랫) 플레이 시간 추이", "#BF360C")
            plot_municipality_lines(dff, "지자체별 플레이 시간", metric_label="시간")
        else:
            st.info("데이터 없음")


# ============================================================
# 👤 13.맞고(게스트)
# ============================================================
elif page == "👤 13.맞고(게스트)":
    st.markdown('<div class="section-header">👤 맞고 (게스트)</div>', unsafe_allow_html=True)
    p_start, p_end = page_week_range_selector("matgo_guest", weeks)
    st.divider()

    # 집계형 시트에서 직접 가져오기
    guest_raw = sheets.get("맞고게스트", pd.DataFrame())
    if not guest_raw.empty:
        guest_df = guest_raw.copy()
        week_col = None
        for c in guest_df.columns:
            if "주차" in str(c):
                week_col = c
                break
        num_cols = [c for c in guest_df.columns if c not in [week_col, "시작일"] and c is not None]
        for c in num_cols:
            guest_df[c] = guest_df[c].apply(safe_numeric)

        if week_col:
            guest_df = filter_by_week_range(guest_df, week_col, p_start, p_end, weeks)
            guest_df = shorten_dates_in_df(guest_df, week_col)

        tab1, tab2 = st.tabs(["추이 차트", "상세 데이터"])
        with tab1:
            if week_col:
                user_col = next((c for c in num_cols if "이용자" in str(c) and "WoW" not in str(c) and "1인" not in str(c)), None)
                play_col = next((c for c in num_cols if "플레이 판수" in str(c) or "플레이판수" in str(c)), None)
                time_col = next((c for c in num_cols if "플레이 시간" in str(c) and "1인" not in str(c) and "초" not in str(c)), None)
                per_play_col = next((c for c in num_cols if "1인당" in str(c) and "플레이 판" in str(c)), None)

                if user_col:
                    fig = make_subplots(specs=[[{"secondary_y": True}]])
                    fig.add_trace(go.Bar(x=guest_df[week_col], y=guest_df[user_col], name="이용자수",
                                         marker_color="#795548",
                                         hovertemplate="%{y:,}명<extra>이용자수</extra>"), secondary_y=False)
                    if play_col:
                        fig.add_trace(go.Bar(x=guest_df[week_col], y=guest_df[play_col], name="플레이판수",
                                             marker_color="#BCAAA4",
                                             hovertemplate="%{y:,}판<extra>플레이판수</extra>"), secondary_y=False)
                    if per_play_col:
                        fig.add_trace(go.Scatter(x=guest_df[week_col], y=guest_df[per_play_col],
                                                  name="1인당 플레이판수", mode="lines+markers",
                                                  line=dict(color="#D32F2F", width=2),
                                                  hovertemplate="%{y:.1f}판<extra>1인당</extra>"), secondary_y=True)
                    fig.update_layout(title="맞고(게스트) 이용 현황", height=420,
                                      hovermode="x unified", barmode="group",
                                      xaxis=dict(type="category"),
                                      legend=LEGEND_BELOW, margin=dict(t=40, b=70))
                    fig.update_yaxes(title_text="명/판", secondary_y=False)
                    if per_play_col:
                        fig.update_yaxes(title_text="1인당 판수", secondary_y=True)
                    st.plotly_chart(fig, use_container_width=True)

        with tab2:
            st.dataframe(guest_df, use_container_width=True, height=400)
    else:
        st.info("맞고(게스트) 데이터가 없습니다.")


# ============================================================
# 😰 스트레스체크
# ============================================================
elif page == "😰 7.스트레스체크":
    st.markdown('<div class="section-header">😰 스트레스체크</div>', unsafe_allow_html=True)
    selected_biz = biz_selector("스트레스")
    st.divider()

    p_start, p_end = page_week_range_selector("stress", weeks)

    tab1, tab2 = st.tabs(["이용자수 추이", "수행횟수 추이"])

    with tab1:
        stress_users = data.get("weekly_스트레스이용자", pd.DataFrame())
        stress_user_raw = sheets.get("스트레스이용자", pd.DataFrame())
        if not stress_user_raw.empty:
            su = stress_user_raw.copy()
            _wc, _sum_col, _rc = None, None, None
            for c in su.columns:
                cl = str(c).replace("\n", "").strip()
                if "주차" in cl and _wc is None: _wc = c
                elif ("이용자합계" in cl or ("합계" in cl and "이용자" in cl)) and _sum_col is None: _sum_col = c
                elif ("전체이용비중" in cl or "이용비중" in cl) and _rc is None: _rc = c
            stress_total_c = data.get("total_스트레스이용자", pd.DataFrame())
            if _wc:
                su = filter_by_week_range(su, _wc, p_start, p_end, weeks)
                su = shorten_dates_in_df(su, _wc)
                if not stress_total_c.empty:
                    stt = filter_by_week_range(stress_total_c, "주차", p_start, p_end, weeks)
                    stt = shorten_dates_in_df(stt, "주차")
                    ct_map = dict(zip(stt["주차"], stt["값"].apply(safe_numeric)))
                    su["_bar"] = su[_wc].map(ct_map).fillna(su[_sum_col].apply(safe_numeric) if _sum_col else 0)
                    bar_col_use = "_bar"
                else:
                    bar_col_use = _sum_col
                if bar_col_use:
                    if selected_biz != "전체":
                        _biz_cnt, _ = biz_agg_raw(su, selected_biz, _wc)
                        if _biz_cnt is not None:
                            su = su.copy()
                            su["_bar"] = _biz_cnt
                            bar_col_use = "_bar"
                            # 시트 내장 이용비중 대신 실제 가입완료 인원으로 직접 계산
                            _biz_completed = biz_filter_df(data.get("registration", pd.DataFrame()), selected_biz)["가입완료"].apply(safe_numeric).sum()
                            if _biz_completed > 0:
                                su["_rc"] = (_biz_cnt / _biz_completed * 100).round(1)
                                _rc = "_rc"
                    plot_bar_rate_dual(su, _wc, bar_col_use, "이용자수", "#AB47BC",
                                       _rc, "이용비중", "#FF6F00",
                                       "스트레스체크 이용자수 + 이용비중")
            # ── 지자체별 이용자비중 추이 (구글 시트 AI~BK열 직접 사용) ──────────────────────────
            mrt_stress = biz_filter_df(extract_mun_ratio_trend(stress_user_raw), selected_biz)
            if not mrt_stress.empty:
                mrt_stress = filter_by_week_range(mrt_stress, "주차", p_start, p_end, weeks)
                _active_s = mrt_stress.groupby("지자체명")["값"].sum()
                _active_s = _active_s[_active_s > 0].index.tolist()
                mrt_stress = mrt_stress[mrt_stress["지자체명"].isin(_active_s)]
                if not mrt_stress.empty:
                    plot_municipality_lines(mrt_stress, "지자체별 스트레스체크 이용자비중 추이 (%)", metric_label="이용자비중(%)")
        else:
            st.info("스트레스체크 이용자 데이터가 없습니다.")

    with tab2:
        stress_count = data.get("weekly_스트레스수행횟수", pd.DataFrame())
        stress_exam_raw = sheets.get("스트레스수행횟수", pd.DataFrame())
        if not stress_exam_raw.empty:
            se = stress_exam_raw.copy()
            _wc, _sum_col, _awc = None, None, None
            for c in se.columns:
                cl = str(c).replace("\n", "").strip()
                if "주차" in cl and _wc is None: _wc = c
                elif "합계" in cl and _sum_col is None: _sum_col = c
                elif "1인" in cl and "주평균" in cl and _awc is None: _awc = c
            stress_exam_total_c = data.get("total_스트레스수행횟수", pd.DataFrame())
            if _wc:
                se = filter_by_week_range(se, _wc, p_start, p_end, weeks)
                se = shorten_dates_in_df(se, _wc)
                if not stress_exam_total_c.empty:
                    sett = filter_by_week_range(stress_exam_total_c, "주차", p_start, p_end, weeks)
                    sett = shorten_dates_in_df(sett, "주차")
                    ct_map = dict(zip(sett["주차"], sett["값"].apply(safe_numeric)))
                    se["_bar"] = se[_wc].map(ct_map).fillna(se[_sum_col].apply(safe_numeric) if _sum_col else 0)
                    bar_col_use = "_bar"
                else:
                    bar_col_use = _sum_col
                if bar_col_use:
                    if selected_biz != "전체":
                        _biz_cnt, _ = biz_agg_raw(se, selected_biz, _wc)
                        if _biz_cnt is not None:
                            se = se.copy()
                            se["_bar"] = _biz_cnt
                            bar_col_use = "_bar"
                    plot_bar_rate_dual(se, _wc, bar_col_use, "수행횟수", "#AB47BC",
                                       _awc, "1인 주평균", "#455A64",
                                       "스트레스체크 수행횟수 + 1인 주평균", bar_unit="회", line_unit="회")
            sf = filter_by_week_range(stress_count, "주차", p_start, p_end, weeks) if not stress_count.empty else pd.DataFrame()
            sf = biz_filter_df(sf, selected_biz)
            if not sf.empty:
                plot_municipality_lines(sf, "지자체별 스트레스체크 수행횟수 추이", metric_label="수행횟수")
        else:
            st.info("스트레스체크 수행횟수 데이터가 없습니다.")



# ============================================================
# 🤖 AI 생활지원사
# ============================================================
elif page == "🤖 AI 생활지원사":
    st.markdown('<div class="section-header">🤖 AI 생활지원사</div>', unsafe_allow_html=True)

    ai_df = data.get("ai_funnel", pd.DataFrame())
    if not ai_df.empty and "주차" in ai_df.columns:
        ai_df = ai_df.copy()

        # 컬럼 찾기 (줄바꿈 제거 후 매칭)
        cols_map = {}
        for c in ai_df.columns:
            cl = str(c).replace("\n", "").replace(" ", "").strip()
            if cl == "회원수": cols_map["회원수"] = c
            elif cl == "인트로": cols_map["인트로율"] = c
            elif "인트로" in cl and "단계" in cl and "회원" in cl and "건수" not in cl: cols_map["인트로수"] = c
            elif cl == "프로그램완료" and "회원" not in cl: cols_map["완료율"] = c
            elif "프로그램" in cl and "완료" in cl and "회원" in cl and "건수" not in cl: cols_map["완료수"] = c
            elif cl == "서비스": cols_map["서비스율"] = c
            elif "서비스" in cl and "제안" in cl and "단계" in cl and "회원" in cl and "건수" not in cl: cols_map["서비스수"] = c

        # 숫자 변환 + NaN → 0
        for k, c in cols_map.items():
            if c in ai_df.columns:
                ai_df[c] = ai_df[c].apply(safe_numeric).fillna(0)

        # 이상 데이터 제거 (주차 형식이 아닌 행)
        ai_df = ai_df[ai_df["주차"].astype(str).str.match(r"^\d{2}-\d{2}$", na=False)]
        # 주차 정리 + X축 짧게
        ai_chart = shorten_dates_in_df(ai_df, "주차")

        # 26-08~10 데이터 없는 주차 표시를 위한 처리
        existing_weeks = set(ai_chart["주차"].tolist())
        missing_weeks = [w for w in ["26-08", "26-09", "26-10"] if w not in existing_weeks]

        tab3, tab5, tab1, tab2, tab4 = st.tabs(["📅 월별 추이", "🏛 지자체별 비교", "참여율 추이 (%)", "참여 인원 (명)", "상세 데이터"])

        with tab1:
            # 주차별 인트로 참여율만 표시 (전화 수신 = 주간 지표)
            # 프로그램 완료·서비스 이용률은 월간 지표이므로 월별 탭에서 확인
            fig = go.Figure()

            if "인트로율" in cols_map:
                _intro_cnt = ai_chart[cols_map["인트로수"]].apply(safe_numeric).fillna(0).round().astype(int) if "인트로수" in cols_map else pd.Series([0]*len(ai_chart))
                fig.add_trace(go.Scatter(
                    x=ai_chart["주차"], y=ai_chart[cols_map["인트로율"]],
                    name="인트로 참여율 (전화 수신)",
                    mode="lines+markers",
                    line=dict(color="#7B1FA2", width=2.5),
                    customdata=_intro_cnt,
                    hovertemplate="<b>%{x}</b><br>인트로: %{y:.1f}% (%{customdata}명 수신)<extra></extra>"
                ))

            # 지자체 구분 annotation
            fig.add_annotation(x="26-07", y=55, text="← 독거노인지원종합센터 | 데이터 없음 (26-08~10) | 삼척시청 →",
                               showarrow=False, font=dict(size=9, color="#999"),
                               bgcolor="rgba(255,255,255,0.8)", bordercolor="#ccc", borderwidth=1)

            # 각 주차별 회원수 표시 (X축 아래 annotation)
            if "회원수" in cols_map:
                for _, row in ai_chart.iterrows():
                    w = row["주차"]
                    m = int(row[cols_map["회원수"]])
                    if m > 0:
                        fig.add_annotation(
                            x=w, y=0, yref="y", yshift=-20,
                            text=f"{m}명", showarrow=False,
                            font=dict(size=8, color="#666"),
                        )

            fig.update_layout(
                title="주차별 인트로 참여율 (전화 수신률) — 프로그램·서비스 지표는 월별 탭 참고",
                height=460, hovermode="x unified",
                xaxis=dict(type="category", title=""),
                yaxis=dict(title="인트로 참여율 (%)", range=[-5, 100]),
                legend=LEGEND_BELOW, margin=dict(t=50, b=80),
            )
            st.info("📞 인트로(전화 수신)는 주1회 통화 기반의 주간 지표입니다. 프로그램 완료율·서비스 이용률은 월간 지표이므로 **월별 추이 탭**에서 확인하세요.", icon=None)
            st.plotly_chart(fig, use_container_width=True)

        with tab2:
            # 주차별 인트로 수신 인원수 (전화 받은 명수)
            fig2 = make_subplots(specs=[[{"secondary_y": True}]])

            if "회원수" in cols_map:
                fig2.add_trace(go.Scatter(
                    x=ai_chart["주차"], y=ai_chart[cols_map["회원수"]],
                    name="전체 회원수", mode="lines+markers",
                    line=dict(color="#9E9E9E", width=1, dash="dot"),
                    hovertemplate="%{y:,}명<extra>전체 회원수</extra>"
                ), secondary_y=True)

            if "인트로수" in cols_map:
                _intro_vals = ai_chart[cols_map["인트로수"]].apply(safe_numeric).fillna(0)
                fig2.add_trace(go.Bar(
                    x=ai_chart["주차"], y=_intro_vals,
                    name="인트로 수신 인원 (전화 받음)", marker_color="#7B1FA2",
                    text=_intro_vals.apply(lambda x: f"{int(x)}명" if x > 0 else ""),
                    textposition="outside", textfont=dict(size=10),
                    hovertemplate="<b>%{x}</b><br>전화 수신: <b>%{y:,.0f}명</b><extra></extra>"
                ), secondary_y=False)

            # 데이터 없는 구간 표시
            fig2.add_annotation(x="26-07", y=0, yref="paper", yshift=10,
                               text="← 독거노인 | 데이터 없음 (08~10) | 삼척시청 →",
                               showarrow=False, font=dict(size=9, color="#999"),
                               bgcolor="rgba(255,255,255,0.8)")

            fig2.update_layout(
                title="주차별 인트로 수신 인원 (전화 받은 명수)",
                height=450, hovermode="x unified",
                barmode="group", xaxis=dict(type="category"),
                legend=LEGEND_BELOW, margin=dict(t=40, b=80),
                bargap=0.3,
            )
            fig2.update_yaxes(title_text="수신 인원 (명)", secondary_y=False)
            fig2.update_yaxes(title_text="전체 회원수", secondary_y=True)
            st.plotly_chart(fig2, use_container_width=True)

        with tab3:
            st.markdown(
                "<div style='font-size:0.85rem;color:#555;padding:0.4rem 0 0.8rem'>"
                "📊 막대: <b>인트로 참여율</b> · <b>서비스 제안율</b> — 주차별 &nbsp;|&nbsp; "
                "➖ 선: <b>프로그램 완료율</b> — 주차별 누적"
                "</div>",
                unsafe_allow_html=True,
            )

            import re as _re

            _MUN_COLORS_M = {
                "삼척시청": "#1565C0",
                "양양군청": "#2E7D32",
                "정선군청": "#E65100",
                "고성군청": "#00838F",
                "다살림재가노인지원서비스센터": "#8D6E63",
                "계양구청": "#283593",
            }
            _SVC_COLORS_M = {
                "#1565C0": "#42A5F5",
                "#2E7D32": "#43A047",
                "#E65100": "#F4511E",
                "#00838F": "#26C6DA",
                "#8D6E63": "#A1887F",
                "#283593": "#3949AB",
            }
            _MUN_ORDER_M = ["삼척시청", "양양군청", "정선군청", "고성군청", "다살림재가노인지원서비스센터", "계양구청"]

            # 데이터 소스 — 외부 시트 C열 기준
            ai_mun_weekly  = data.get("ai_municipality_ext", pd.DataFrame())
            ai_mun_monthly = data.get("ai_mun_monthly", {})

            def _get_month_num(period_str: str) -> int:
                """'11주차 (3월 8일~14일)' → 3"""
                m = _re.search(r'(\d+)월', str(period_str))
                return int(m.group(1)) if m else 0

            def _get_last_month_num(period_str: str) -> int:
                """'18주차 (4월 26일~5월 2일)' → 5 (마지막 월 기준 — 토요일 진행 지자체용)"""
                months = _re.findall(r'(\d+)월', str(period_str))
                return int(months[-1]) if months else 0

            def _short_period(period_str: str) -> str:
                """'11주차 (3월 8일~14일)' → '11주차'"""
                m = _re.search(r'(\d+주차)', str(period_str))
                return m.group(1) if m else str(period_str)

            if ai_mun_weekly.empty:
                st.info("주차별 데이터가 없습니다.")
            else:
                _weekly_muns = ai_mun_weekly[ai_mun_weekly["지자체"].isin(_MUN_ORDER_M)].copy()
                _muns_m = [m for m in _MUN_ORDER_M if m in _weekly_muns["지자체"].values]

                if not _muns_m:
                    st.info("삼척/양양/정선 데이터가 없습니다.")
                else:
                    def _month_from_gubn(s):
                        m = _re.search(r'(\d+)월', str(s))
                        return int(m.group(1)) if m else 0

                    for mun in _muns_m:
                        color_base = _MUN_COLORS_M.get(mun, "#607D8B")
                        svc_color  = _SVC_COLORS_M.get(color_base, "#90A4AE")

                        mun_wk = _weekly_muns[_weekly_muns["지자체"] == mun].copy()
                        if mun_wk.empty:
                            continue

                        # X 라벨: 외부 시트 C열 직접 사용 (ai_municipality_ext에서 이미 필터됨)
                        mun_wk["_xlbl"]  = mun_wk["월간구분"]
                        mun_wk["_month"] = mun_wk["월간구분"].apply(_month_from_gubn)

                        # 이 지자체의 X 순서 + 월별 그룹 (각자 독립)
                        _mun_x_labels = mun_wk["_xlbl"].tolist()
                        _mun_month_groups = {}
                        for _xl, _xm in zip(mun_wk["_xlbl"].tolist(), mun_wk["_month"].tolist()):
                            _mun_month_groups.setdefault(int(_xm), []).append(_xl)

                        x_labels = _mun_x_labels
                        periods  = mun_wk["기간"].tolist()

                        alarm_users = mun_wk["receiveAlarmUserCount"].apply(safe_numeric).fillna(0)
                        intro_vals  = mun_wk["intro(%)"].apply(safe_numeric).fillna(0)
                        svc_vals    = mun_wk["service proposal(%)"].apply(safe_numeric).fillna(0)
                        intro_cnts  = mun_wk["intro"].apply(safe_numeric).fillna(0).round().astype(int)
                        svc_cnts    = mun_wk["service proposal"].apply(safe_numeric).fillna(0).round().astype(int) if "service proposal" in mun_wk.columns else (svc_vals / 100 * alarm_users).round().astype(int)

                        fig_mm = go.Figure()

                        # ── 주별 막대: 인트로 참여율
                        fig_mm.add_trace(go.Bar(
                            x=x_labels, y=intro_vals,
                            name="인트로 참여율",
                            marker_color=color_base,
                            text=[f"{v:.0f}%<br>({c}명)" if v > 0 else ""
                                  for v, c in zip(intro_vals, intro_cnts)],
                            textposition="outside", textfont=dict(size=10),
                            customdata=list(zip(intro_cnts, periods)),
                            hovertemplate=(
                                "<b>%{customdata[1]}</b><br>"
                                "인트로: <b>%{y:.1f}%</b> (%{customdata[0]}명)"
                                "<extra></extra>"
                            ),
                        ))

                        # ── 주별 막대: 서비스 제안율
                        fig_mm.add_trace(go.Bar(
                            x=x_labels, y=svc_vals,
                            name="서비스 제안율",
                            marker_color=svc_color,
                            text=[f"{v:.0f}%<br>({c}명)" if v > 0 else ""
                                  for v, c in zip(svc_vals, svc_cnts)],
                            textposition="outside", textfont=dict(size=10),
                            customdata=list(zip(svc_cnts, periods)),
                            hovertemplate=(
                                "<b>%{customdata[1]}</b><br>"
                                "서비스: <b>%{y:.1f}%</b> (%{customdata[0]}명)"
                                "<extra></extra>"
                            ),
                        ))

                        # ── 월별 누적 프로그램 완료율 (월 내 program complete cumsum / 가입인원)
                        mun_wk["_prog_cnt"] = mun_wk["program complete"].apply(safe_numeric).fillna(0) if "program complete" in mun_wk.columns else pd.Series(0, index=mun_wk.index)
                        mun_wk["_prog_cum_cnt"] = mun_wk.groupby("_month")["_prog_cnt"].cumsum()
                        _last_gaip = mun_wk.groupby("_month")["가입인원"].transform("last").apply(safe_numeric).replace(0, float("nan")).fillna(1)
                        mun_wk["_prog_cum"] = (mun_wk["_prog_cum_cnt"] / _last_gaip * 100).clip(upper=100)
                        prog_cum = mun_wk["_prog_cum"]
                        if prog_cum.max() > 0:
                            prog_cnts = (prog_cum / 100 * _last_gaip).round().astype(int)
                            fig_mm.add_trace(go.Bar(
                                x=x_labels, y=prog_cum,
                                name="프로그램 완료율 (누적)",
                                legendgroup="prog",
                                marker_color="rgba(233,30,99,0.75)",
                                marker_line=dict(color="#E91E63", width=1.5),
                                text=[f"{v:.0f}%<br>({c}명)" if v > 0 else ""
                                      for v, c in zip(prog_cum, prog_cnts)],
                                textposition="outside", textfont=dict(size=10, color="#880E4F"),
                                customdata=list(zip(prog_cnts, periods)),
                                hovertemplate=(
                                    "<b>%{customdata[1]}</b><br>"
                                    "누적 완료율: <b>%{y:.1f}%</b> (%{customdata[0]}명)"
                                    "<extra></extra>"
                                ),
                            ))

                        # ── 월 구분선 + 월 레이블 (이 지자체 기준, 기존 법칙 그대로)
                        _mun_months_sorted = sorted(_mun_month_groups.keys())
                        for _mi, _rm in enumerate(_mun_months_sorted):
                            if _mi > 0:
                                _first_lbl = _mun_month_groups[_rm][0]
                                _bi = _mun_x_labels.index(_first_lbl) if _first_lbl in _mun_x_labels else 0
                                fig_mm.add_vline(
                                    x=_bi - 0.5,
                                    line=dict(color="#78909C", width=1.5, dash="dot"),
                                )
                            _lbls = _mun_month_groups[_rm]
                            _mid = _lbls[len(_lbls) // 2]
                            fig_mm.add_annotation(
                                x=_mid, y=138,
                                text=f"<b>{_rm}월</b>",
                                showarrow=False,
                                font=dict(size=12, color="#37474F"),
                                bgcolor="rgba(236,239,241,0.92)",
                                bordercolor="#90A4AE",
                                borderpad=4, borderwidth=1,
                            )

                        fig_mm.update_layout(
                            title=dict(text=mun, font=dict(size=16, color=color_base)),
                            height=480, barmode="group", bargap=0.2,
                            hovermode="x unified",
                            xaxis=dict(
                                type="category",
                                categoryorder="array",
                                categoryarray=_mun_x_labels,
                                title="",
                            ),
                            yaxis=dict(title="참여율 (%)", range=[0, 148]),
                            legend=dict(orientation="h", yanchor="top", y=-0.20,
                                        xanchor="center", x=0.5),
                            margin=dict(t=55, b=90, r=20),
                        )
                        st.plotly_chart(fig_mm, use_container_width=True)

        with tab4:
            st.dataframe(ai_df, use_container_width=True, height=400)

        with tab5:
            # ── 지자체별 비교 (신규 시트: gid=887906400) ──────────────────
            MUN_COLORS = {
                "삼척시청": "#1565C0",   # 파랑
                "양양군청": "#2E7D32",   # 초록
                "정선군청": "#E65100",   # 주황
                "고성군청": "#00838F",   # 청록
                "다살림재가노인지원서비스센터": "#8D6E63",   # 갈색
                "계양구청": "#283593",   # 남색
            }

            ai_mun = data.get("ai_municipality", pd.DataFrame())

            if ai_mun.empty:
                st.info("지자체별 데이터가 없습니다. (gid=887906400)")
            else:
                mun_df    = ai_mun.copy()
                agg_df    = mun_df[mun_df["지자체"] == "통합"].copy()   # 통합 행
                mun_only  = mun_df[mun_df["지자체"] != "통합"].copy()   # 지자체별 행
                periods   = mun_only["기간"].unique().tolist()
                muns      = mun_only["지자체"].unique().tolist()

                # ── 지자체 계약 시작 정보 ─────────────────────────────────
                mun_info = (
                    mun_only.groupby("지자체", sort=False)
                    .agg(계약시작주차=("기간", "first"), 알람요일=("알람요일", "first"))
                    .reset_index()
                )
                # ── 지자체별 현황 카드 (계약정보 + 최신 실적 통합) ───────────────────
                # 계약 시작주차 lookup: mun_info에서 추출
                mun_start_map = {
                    row["지자체"]: (row["계약시작주차"] or "-")
                    for _, row in mun_info.iterrows()
                }
                latest_period = periods[-1] if periods else None
                if latest_period:
                    latest = mun_only[mun_only["기간"] == latest_period].copy()
                    st.markdown(
                        f"<div style='font-weight:700;font-size:0.95rem;"
                        f"color:#1E293B;margin:0.3rem 0 0.8rem'>"
                        f"📌 {latest_period} 기준 지자체별 현황</div>",
                        unsafe_allow_html=True,
                    )
                    cols_kpi = st.columns(len(latest))
                    for i, (_, row) in enumerate(latest.iterrows()):
                        mun = row["지자체"]
                        color = MUN_COLORS.get(mun, "#607D8B")
                        alarm = row.get("알람요일") or "미정"
                        intro_pct  = safe_numeric(row.get("intro(%)", 0))
                        svc_pct    = safe_numeric(row.get("service proposal(%)", 0))
                        prog_pct   = safe_numeric(row.get("program(%)", 0))
                        alarm_user  = int(safe_numeric(row.get("receiveAlarmUserCount", 0)))
                        joined_user = int(safe_numeric(row.get("가입인원", 0)))
                        intro_cnt  = int(safe_numeric(row.get("intro", 0))) if safe_numeric(row.get("intro", 0)) > 0 else int(round(intro_pct / 100 * alarm_user)) if alarm_user > 0 else 0
                        svc_cnt    = int(round(svc_pct / 100 * alarm_user)) if alarm_user > 0 else 0
                        prog_cnt   = int(round(prog_pct / 100 * alarm_user)) if alarm_user > 0 else 0
                        raw_period = mun_start_map.get(mun, "-")
                        start_date = raw_period.split("~")[0].strip() if "~" in raw_period else raw_period
                        with cols_kpi[i]:
                            st.markdown(
                                f"""<div style="background:{color};border-radius:14px;
                                    padding:1.2rem 1rem;color:white;text-align:center;
                                    box-shadow:0 4px 12px rgba(0,0,0,0.15)">
                                  <b style="font-size:1.05rem">{mun}</b><br>
                                  <div style="font-size:0.78rem;opacity:0.9;margin:0.3rem 0 0.5rem">
                                    📅 {start_date} &nbsp;|&nbsp; 🔔 {alarm} 알람
                                  </div>
                                  <hr style="border:none;border-top:1px solid rgba(255,255,255,0.35);margin:0 0 0.6rem">
                                  <div style="font-size:0.78rem;opacity:0.85;margin-bottom:0.2rem">
                                    👥 가입인원 {joined_user}명 &nbsp;|&nbsp; 📨 알람도달 {alarm_user}명
                                  </div>
                                  <div style="font-size:1.9rem;font-weight:900;line-height:1.1">{intro_pct:.0f}%</div>
                                  <div style="font-size:0.72rem;opacity:0.9;margin-bottom:0.5rem">📊 인트로 참여율 <b>({intro_cnt}명)</b></div>
                                  <div style="font-size:0.82rem">
                                    🛎️ 서비스 {svc_pct:.0f}% ({svc_cnt}명) &nbsp;|&nbsp; ✅ 프로그램 {prog_pct:.0f}% ({prog_cnt}명)
                                  </div>
                                </div>""",
                                unsafe_allow_html=True,
                            )

                st.markdown("---")

                # ── 지자체별 월별 지표 차트 (ai_mun_monthly: 삼척/양양/정선 전용 시트) ──
                _MM_COLORS = {
                    "인트로참여율":  "#7B1FA2",
                    "서비스제안율":  "#00897B",
                    "프로그램완료율": "#E91E63",
                }
                _MM_LABELS = {
                    "인트로참여율":  "인트로 참여율",
                    "서비스제안율":  "서비스 제안율",
                    "프로그램완료율": "프로그램 완료율",
                }
                ai_mun_monthly = data.get("ai_mun_monthly", {})

                # 표시 순서: 삼척 → 양양 → 정선
                _MUN_ORDER = ["삼척시청", "양양군청", "정선군청"]
                _muns_to_show = [m for m in _MUN_ORDER if m in ai_mun_monthly] + \
                                [m for m in ai_mun_monthly if m not in _MUN_ORDER]

                if _muns_to_show:
                    for mun in _muns_to_show:
                        am = ai_mun_monthly[mun]
                        color_base = MUN_COLORS.get(mun, "#607D8B")
                        alarm_day = mun_only[mun_only["지자체"] == mun]["알람요일"].iloc[0] \
                                    if mun in mun_only["지자체"].values and "알람요일" in mun_only.columns else ""
                        mun_label = f"{mun} ({alarm_day})" if alarm_day else mun

                        fig_mm = go.Figure()
                        bar_colors = {
                            "인트로참여율":  color_base,
                            "서비스제안율":  "#42A5F5" if color_base == "#1565C0" else
                                            "#43A047" if color_base == "#2E7D32" else "#F4511E",
                            "프로그램완료율": "#90CAF9" if color_base == "#1565C0" else
                                            "#A5D6A7" if color_base == "#2E7D32" else "#FFAB91",
                        }
                        for col, label in _MM_LABELS.items():
                            if col not in am.columns:
                                continue
                            vals = am[col].apply(safe_numeric).fillna(0)
                            # 회원수 기반 인원 계산
                            if "회원수" in am.columns:
                                mbr = am["회원수"].apply(safe_numeric).fillna(0)
                                cnts = (vals / 100 * mbr).round().astype(int)
                            else:
                                cnts = pd.Series([0] * len(vals))
                            fig_mm.add_trace(go.Bar(
                                x=am["월"], y=vals,
                                name=label,
                                marker_color=bar_colors.get(col, "#607D8B"),
                                text=[f"{v:.0f}%\n({c}명)" if v > 0 else "" for v, c in zip(vals, cnts)],
                                textposition="outside",
                                textfont=dict(size=11),
                                customdata=cnts,
                                hovertemplate=f"<b>%{{x}}</b><br>{label}: <b>%{{y:.1f}}%</b> (%{{customdata}}명)<extra></extra>",
                            ))
                        fig_mm.update_layout(
                            title=dict(text=mun_label, font=dict(size=16, color=color_base)),
                            height=400,
                            barmode="group",
                            bargap=0.25,
                            hovermode="x unified",
                            xaxis=dict(type="category", title=""),
                            yaxis=dict(title="참여율 (%)", range=[0, 120]),
                            legend=dict(orientation="h", yanchor="top", y=-0.18,
                                        xanchor="center", x=0.5),
                            margin=dict(t=50, b=90),
                        )
                        st.plotly_chart(fig_mm, use_container_width=True)
                else:
                    st.info("월별 데이터를 불러오는 중입니다. 잠시 후 새로고침해주세요.")

                # ── 상세 데이터 표 ─────────────────────────────────────────
                with st.expander("상세 데이터 보기"):
                    st.dataframe(mun_df, use_container_width=True, height=350)

    else:
        st.info("AI 생활지원사 데이터가 없습니다.")


# ============================================================
# 📊 콘텐츠·생활·날씨
# ============================================================
elif page == "🩺 9.건강상담":
    st.markdown('<div class="section-header">🩺 건강상담</div>', unsafe_allow_html=True)
    selected_biz = biz_selector("건강상담")
    st.divider()
    p_start, p_end = page_week_range_selector("health", weeks)

    health_df = data.get("건강상담", pd.DataFrame())
    if not health_df.empty:
        # 주차 기간 필터링
        for c in health_df.columns:
            if "주차" in str(c):
                health_df = filter_by_week_range(health_df, c, p_start, p_end, weeks)
                break
        # 주차 컬럼 찾기
        week_col = None
        for c in health_df.columns:
            if "주차" in str(c):
                week_col = c
                break

        # 숫자 컬럼 변환
        num_cols = []
        for c in health_df.columns:
            if c not in [week_col, "시작일"] and c is not None:
                health_df[c] = health_df[c].apply(safe_numeric)
                if health_df[c].sum() > 0:
                    num_cols.append(c)

        hc_mun = biz_filter_df(data.get("건강상담지자체", pd.DataFrame()), selected_biz, col="지자체")
        tab1, tab2, tab3 = st.tabs(["이용 추이", "지자체별 서비스 현황", "상세 데이터"])

        with tab1:
            if week_col:
                health_chart = shorten_dates_in_df(health_df, week_col)
            else:
                health_chart = health_df.copy()

            # 이용 건수(막대) + 전체이용비중(꺾은선) 이중축
            if week_col and "실제 이용 건수" in health_chart.columns:
                plot_bar_rate_dual(
                    health_chart, week_col,
                    bar_col="실제 이용 건수",   bar_label="실제 이용 건수", bar_color="#26C6DA",
                    line_col="전체이용비중",    line_label="전체 이용비중", line_color="#FF6F00",
                    title="건강상담 이용 건수 & 전체 이용비중",
                    bar_unit="건", line_unit="%",
                )

            if week_col and "전화버튼클릭건수" in health_chart.columns:
                fig = go.Figure()
                if "메뉴클릭건수" in health_chart.columns:
                    fig.add_trace(go.Scatter(x=health_chart[week_col], y=health_chart["메뉴클릭건수"],
                                             name="메뉴클릭", mode="lines+markers", line=dict(color="#2196F3")))
                fig.add_trace(go.Scatter(x=health_chart[week_col], y=health_chart["전화버튼클릭건수"],
                                         name="전화버튼클릭", mode="lines+markers", line=dict(color="#FF6F00")))
                if "아웃바운드\n성공건수" in health_chart.columns:
                    fig.add_trace(go.Scatter(x=health_chart[week_col], y=health_chart["아웃바운드\n성공건수"],
                                             name="아웃바운드 성공", mode="lines+markers", line=dict(color="#00C853")))
                fig.update_layout(title="건강상담 클릭/통화 추이", height=350,
                                  hovermode="x unified", xaxis=dict(type="category"),
                                  legend=dict(orientation="h", yanchor="top", y=-0.15, xanchor="center", x=0.5))
                st.plotly_chart(fig, use_container_width=True)

            # ── 서비스 유형별 이용건수 (tab1 인라인) ──────────────────────
            if not hc_mun.empty:
                _hc = hc_mun.copy()
                _SERVICE_COLS = [c for c in ["전문의료진상담", "병원안내", "일반상담", "진료예약"] if c in _hc.columns]
                _SERVICE_COLORS = {"전문의료진상담": "#26C6DA", "병원안내": "#5B73E8", "일반상담": "#66BB6A", "진료예약": "#FF6F00"}
                _hc["주차"] = _hc["날짜"].apply(date_to_week_label)
                if p_start and p_end and weeks:
                    _s = weeks.index(p_start) if p_start in weeks else 0
                    _e = weeks.index(p_end) if p_end in weeks else len(weeks) - 1
                    _week_order = weeks[_s:_e + 1]
                    _hc_f = _hc[_hc["주차"].isin(set(_week_order))].copy()
                else:
                    _week_order = list(dict.fromkeys(_hc.sort_values("날짜")["주차"].tolist()))
                    _hc_f = _hc.copy()
                _hc_week = _hc_f.groupby("주차")[_SERVICE_COLS].sum().reset_index()
                if _week_order:
                    _all_w = pd.DataFrame({"주차": _week_order})
                    _hc_week = _all_w.merge(_hc_week, on="주차", how="left").fillna(0)
                    _hc_week["주차"] = pd.Categorical(_hc_week["주차"], categories=_week_order, ordered=True)
                    _hc_week = _hc_week.sort_values("주차")
                _fig_st = go.Figure()
                for _sc in _SERVICE_COLS:
                    _fig_st.add_trace(go.Bar(
                        x=_hc_week["주차"].astype(str), y=_hc_week[_sc], name=_sc,
                        marker_color=_SERVICE_COLORS.get(_sc, "#999"),
                        text=_hc_week[_sc].apply(lambda v: f"{int(v)}" if v > 0 else ""),
                        textposition="inside", textfont=dict(size=11, color="white"),
                        hovertemplate=f"<b>%{{x}}</b><br>{_sc}: %{{y:,}}건<extra></extra>",
                    ))
                _fig_st.update_layout(
                    barmode="stack",
                    title="주차별 건강상담 서비스 유형별 이용건수",
                    height=400, hovermode="x unified",
                    xaxis=dict(type="category", tickangle=-45, tickfont=dict(size=11)),
                    yaxis=dict(title="이용건수"),
                    legend=LEGEND_BELOW, margin=dict(t=45, b=90),
                )
                st.plotly_chart(_fig_st, use_container_width=True, key="hc_service_type_tab1")

        # ── 탭 2: 지자체별 서비스 유형별 현황 ────────────────────────────
        with tab2:
            # 캐시 데이터가 비어있으면 직접 재시도
            if hc_mun.empty:
                with st.spinner("건강상담 지자체 데이터 로딩 중..."):
                    try:
                        from sheets_data import fetch_sheet, get_health_consult_by_municipality as _get_hc_mun
                        _raw = fetch_sheet("867975933")
                        if not _raw.empty:
                            hc_mun = _get_hc_mun({"건강상담지자체": _raw})
                    except Exception as _e:
                        st.error(f"데이터 로딩 오류: {_e}")
            if hc_mun.empty:
                st.info("건강상담 지자체 서비스 이용 데이터가 아직 없습니다.")
            else:
                SERVICE_COLS_HC = [c for c in ["전문의료진상담", "병원안내", "일반상담", "진료예약"]
                                   if c in hc_mun.columns]
                SERVICE_COLORS_HC = {
                    "전문의료진상담": "#26C6DA",
                    "병원안내":       "#5B73E8",
                    "일반상담":       "#66BB6A",
                    "진료예약":       "#FF6F00",
                }

                # 일별 날짜 → ISO 주차 레이블 추가 (2026-04-13 → "26-16")
                hc_mun = hc_mun.copy()
                hc_mun["주차"] = hc_mun["날짜"].apply(date_to_week_label)

                # ── 탭1과 동일한 X축(주차 범위) 계산 ──
                if p_start and p_end and weeks:
                    _s = weeks.index(p_start) if p_start in weeks else 0
                    _e = weeks.index(p_end)   if p_end   in weeks else len(weeks) - 1
                    week_order = weeks[_s:_e + 1]           # 탭1과 동일한 전체 주차 목록
                    hc_filtered = hc_mun[hc_mun["주차"].isin(set(week_order))].copy()
                else:
                    _tmp = hc_mun.sort_values("날짜")
                    week_order = list(dict.fromkeys(_tmp["주차"].tolist()))
                    hc_filtered = hc_mun.copy()

                if hc_filtered.empty and not hc_mun.empty:
                    hc_filtered = hc_mun.copy()             # fallback: 전체 데이터

                # ① 주차별 서비스 유형 스택 바 (전체 합산 — 일별 데이터를 주 단위로 집계)
                hc_week = hc_filtered.groupby("주차")[SERVICE_COLS_HC].sum().reset_index()
                # 탭1 X축과 동일하게: 데이터 없는 주차도 0으로 채워 표시
                if week_order:
                    _all_w = pd.DataFrame({"주차": week_order})
                    hc_week = _all_w.merge(hc_week, on="주차", how="left").fillna(0)
                    hc_week["주차"] = pd.Categorical(hc_week["주차"], categories=week_order, ordered=True)
                    hc_week = hc_week.sort_values("주차")

                fig_stack = go.Figure()
                for sc in SERVICE_COLS_HC:
                    fig_stack.add_trace(go.Bar(
                        x=hc_week["주차"].astype(str), y=hc_week[sc], name=sc,
                        marker_color=SERVICE_COLORS_HC.get(sc, "#999"),
                        text=hc_week[sc].apply(lambda v: f"{int(v)}" if v > 0 else ""),
                        textposition="inside", textfont=dict(size=11, color="white"),
                        hovertemplate=f"<b>%{{x}}</b><br>{sc}: %{{y:,}}건<extra></extra>",
                    ))
                fig_stack.update_layout(
                    barmode="stack",
                    title="주차별 건강상담 서비스 유형별 이용건수",
                    height=400, hovermode="x unified",
                    xaxis=dict(type="category", tickangle=-45, tickfont=dict(size=11)),
                    yaxis=dict(title="이용건수"),
                    legend=LEGEND_BELOW, margin=dict(t=45, b=90),
                )
                st.plotly_chart(fig_stack, use_container_width=True, key="hc_service_type_tab2")

                # ② 실제 데이터가 있는 가장 최근 주차 기준 지자체별 현황
                if not hc_filtered.empty and week_order:
                    _weeks_with_data = [w for w in reversed(week_order)
                                        if w in hc_filtered["주차"].values]
                    latest_week = _weeks_with_data[0] if _weeks_with_data else None
                    if latest_week:
                        latest_hc = hc_filtered[hc_filtered["주차"] == latest_week].groupby("지자체")[SERVICE_COLS_HC].sum().reset_index()
                        latest_hc["합계"] = latest_hc[SERVICE_COLS_HC].sum(axis=1)
                        latest_hc = latest_hc[latest_hc["합계"] > 0].sort_values("합계", ascending=True)
                        if not latest_hc.empty:
                            st.markdown(f"**📌 {latest_week} 기준 지자체별 서비스 이용현황** (이용 있는 지자체만 표시)")
                            fig_mun_hc = go.Figure()
                            for sc in SERVICE_COLS_HC:
                                fig_mun_hc.add_trace(go.Bar(
                                    y=latest_hc["지자체"], x=latest_hc[sc],
                                    name=sc, orientation="h",
                                    marker_color=SERVICE_COLORS_HC.get(sc, "#999"),
                                    text=latest_hc[sc].apply(lambda v: f"{int(v)}" if v > 0 else ""),
                                    textposition="inside", textfont=dict(size=11, color="white"),
                                    hovertemplate=f"<b>%{{y}}</b><br>{sc}: %{{x:,}}건<extra></extra>",
                                ))
                            fig_mun_hc.update_layout(
                                barmode="stack",
                                title=f"지자체별 서비스 유형별 이용건수 ({latest_week})",
                                height=max(300, len(latest_hc) * 38),
                                xaxis=dict(title="이용건수"),
                                yaxis=dict(title=""),
                                legend=LEGEND_BELOW, margin=dict(t=45, b=90),
                            )
                            st.plotly_chart(fig_mun_hc, use_container_width=True)
                        else:
                            st.info(f"📭 {latest_week} 주차에는 이용 데이터가 없습니다.")

                # ③ 지자체별 주차별 라인 차트 (서비스 유형 선택 — 주 단위 집계)
                st.markdown("---")
                sel_svc = st.selectbox("📊 서비스 유형별 지자체 추이",
                                       SERVICE_COLS_HC + ["합계"],
                                       key="hc_svc_select")
                svc_col = sel_svc if sel_svc != "합계" else None
                _svc_agg = sel_svc if sel_svc in SERVICE_COLS_HC else SERVICE_COLS_HC
                hc_line = hc_filtered.groupby(["주차", "지자체"])[SERVICE_COLS_HC].sum().reset_index()
                hc_line["합계"] = hc_line[SERVICE_COLS_HC].sum(axis=1)
                if sel_svc in hc_line.columns:
                    # 선택 서비스에서 값이 하나라도 있는 지자체만 표시 (hover 정리)
                    _nonzero = hc_line.groupby("지자체")[sel_svc].sum()
                    _active_muns = _nonzero[_nonzero > 0].index.tolist()
                    hc_line = hc_line[hc_line["지자체"].isin(_active_muns)]
                    # 주차 정렬 순서 유지 (이미 계산된 week_order 재사용)
                    week_order2 = week_order
                    hc_line["주차"] = pd.Categorical(hc_line["주차"], categories=week_order2, ordered=True)
                    hc_line = hc_line.sort_values(["주차", "지자체"])
                    fig_line = px.line(
                        hc_line, x="주차", y=sel_svc, color="지자체",
                        markers=True,
                        color_discrete_sequence=px.colors.qualitative.Set2,
                        title=f"지자체별 {sel_svc} 주간 추이",
                    )
                    fig_line.update_layout(
                        height=430, hovermode="x unified",
                        xaxis=dict(type="category", tickangle=-45, tickfont=dict(size=11)),
                        yaxis=dict(title="이용건수"),
                        legend=LEGEND_BELOW, margin=dict(t=45, b=100),
                    )
                    fig_line.update_traces(
                        hovertemplate="<b>%{x}</b><br>%{y:,}건<extra>%{fullData.name}</extra>"
                    )
                    st.plotly_chart(fig_line, use_container_width=True)

                # ④ 원본 테이블
                with st.expander("📋 원본 데이터 보기"):
                    st.dataframe(hc_filtered.sort_values(["날짜", "지자체"]),
                                 use_container_width=True, height=350)

        with tab3:
            st.dataframe(health_df, use_container_width=True, height=400)
    else:
        st.info("건강상담 데이터가 없습니다.")


# ============================================================
# 💬 생활상담
# ============================================================
elif page == "💬 10.생활상담":
    st.markdown('<div class="section-header">💬 생활상담</div>', unsafe_allow_html=True)

    p_start, p_end = page_week_range_selector("life", weeks)

    life_df = data.get("생활상담", pd.DataFrame())
    if not life_df.empty:
        for c in life_df.columns:
            if "주차" in str(c):
                life_df = filter_by_week_range(life_df, c, p_start, p_end, weeks)
                break
        week_col = None
        for c in life_df.columns:
            if "주차" in str(c):
                week_col = c
                break

        num_cols = []
        for c in life_df.columns:
            if c not in [week_col, "시작일"] and c is not None:
                life_df[c] = life_df[c].apply(safe_numeric)
                if life_df[c].sum() > 0:
                    num_cols.append(c)

        tab1, tab2 = st.tabs(["이용 추이", "상세 데이터"])

        with tab1:
            if week_col:
                life_chart = shorten_dates_in_df(life_df, week_col)

                # 차트1: 메뉴클릭건수(막대) + 전체이용비중(꺾은선) — 건강상담과 동일 양식
                if "메뉴클릭건수" in life_chart.columns:
                    plot_bar_rate_dual(
                        life_chart, week_col,
                        bar_col="메뉴클릭건수",  bar_label="메뉴클릭건수", bar_color="#8D6E63",
                        line_col="전체이용비중", line_label="전체 이용비중", line_color="#FF6F00",
                        title="생활상담 메뉴클릭건수 & 전체 이용비중",
                        bar_unit="건", line_unit="%",
                    )

                # 차트2: 전화버튼 클릭자수 — 건강상담과 동일 양식
                phone_cols = [c for c in life_chart.columns
                              if "전화" in str(c) and c != "전체이용비중"]
                if phone_cols:
                    fig2 = go.Figure()
                    phone_colors = ["#D32F2F", "#E57373", "#FF8A65"]
                    for i, c in enumerate(phone_cols):
                        fig2.add_trace(go.Scatter(
                            x=life_chart[week_col], y=life_chart[c],
                            name=c, mode="lines+markers",
                            line=dict(color=phone_colors[i % len(phone_colors)])
                        ))
                    fig2.update_layout(
                        title="📞 전화버튼 클릭자수 추이", height=320,
                        hovermode="x unified", xaxis=dict(type="category"),
                        legend=dict(orientation="h", yanchor="top", y=-0.15, xanchor="center", x=0.5)
                    )
                    st.plotly_chart(fig2, use_container_width=True)

        with tab2:
            st.dataframe(life_df, use_container_width=True, height=400)
    else:
        st.info("생활상담 데이터가 없습니다.")


# ============================================================
# 🚶 걸음수
# ============================================================
elif page == "🚶 걸음수":
    st.markdown('<div class="section-header">🚶 걸음수</div>', unsafe_allow_html=True)
    selected_biz = biz_selector("걸음수")
    st.divider()

    _df_steps_raw = sheets.get("걸음수현황", pd.DataFrame())

    # 테스트/내부 계정 제거
    _STEPS_EXCLUDE = {"WAPLAT", "ai생활지원사테스트", "한전MCS"}

    if _df_steps_raw.empty or "date" not in _df_steps_raw.columns:
        st.info("걸음수 데이터가 없습니다.")
    else:
        _df_steps = _df_steps_raw[~_df_steps_raw["agencyName"].isin(_STEPS_EXCLUDE)].copy()
        _df_steps = biz_filter_df(_df_steps, selected_biz, col="agencyName")
        _df_steps["date"] = pd.to_datetime(_df_steps["date"], errors="coerce")
        _df_steps = _df_steps.dropna(subset=["date"]).sort_values("date")
        _df_steps["memberCnt"]    = _df_steps["memberCnt"].apply(safe_numeric)
        _df_steps["dailyStepAvg"] = _df_steps["dailyStepAvg"].apply(safe_numeric)
        _df_steps["totalSteps"]   = _df_steps["memberCnt"] * _df_steps["dailyStepAvg"]

        # ── 날짜 범위 필터 ────────────────────────────────────────────
        _min_date = _df_steps["date"].min().date()
        _max_date = _df_steps["date"].max().date()
        _col_date_l, _col_date_r = st.columns([2, 3])
        with _col_date_l:
            _sel_range = st.date_input(
                "기간 선택",
                value=(_min_date, _max_date),
                min_value=_min_date,
                max_value=_max_date,
            )
        if isinstance(_sel_range, (list, tuple)) and len(_sel_range) == 2:
            _df_filt = _df_steps[
                (_df_steps["date"].dt.date >= _sel_range[0]) &
                (_df_steps["date"].dt.date <= _sel_range[1])
            ]
        else:
            _df_filt = _df_steps
        # 6월 1일 데이터 제외
        _df_filt = _df_filt[~((_df_filt["date"].dt.month == 6) & (_df_filt["date"].dt.day == 1))]

        # ── KPI 카드 (최신 주차 vs 전주차) ──────────────────────────
        _daily_agg = _df_filt.groupby("date").agg(
            _users=("memberCnt", "sum"),
            _steps=("totalSteps", "sum"),
        ).reset_index()
        _daily_agg["_avg_per_person"] = (
            _daily_agg["_steps"] / _daily_agg["_users"].replace(0, float("nan"))
        )

        _max_dt   = _daily_agg["date"].max()
        _curr_start = _max_dt - pd.Timedelta(days=6)
        _prev_end   = _curr_start - pd.Timedelta(days=1)
        _prev_start = _prev_end - pd.Timedelta(days=6)

        _curr_d = _daily_agg[_daily_agg["date"] >= _curr_start]
        _prev_d = _daily_agg[(_daily_agg["date"] >= _prev_start) & (_daily_agg["date"] <= _prev_end)]

        _kpi_users = int(round(_curr_d["_users"].mean())) if not _curr_d.empty else 0
        _kpi_avg   = _curr_d["_avg_per_person"].mean() if not _curr_d.empty else 0

        _prev_users = int(round(_prev_d["_users"].mean())) if not _prev_d.empty else None
        _prev_avg   = _prev_d["_avg_per_person"].mean() if not _prev_d.empty else None

        def _delta_html(curr, prev, unit, fmt=".0f"):
            if prev is None or prev == 0:
                return ""
            diff = curr - prev
            sign = "▲" if diff >= 0 else "▼"
            color = "#4CAF50" if diff >= 0 else "#F44336"
            return f'<span style="font-size:13px;color:{color};margin-left:8px">{sign} {abs(diff):{fmt}}{unit} vs 전주</span>'

        _curr_lbl = f"{_curr_start.strftime('%m/%d')}~{_max_dt.strftime('%m/%d')}"
        _prev_lbl = f"{_prev_start.strftime('%m/%d')}~{_prev_end.strftime('%m/%d')}" if _prev_d is not None and not _prev_d.empty else "-"
        _filt_start = _df_filt["date"].min().strftime("%m/%d")
        _filt_end   = _df_filt["date"].max().strftime("%m/%d")
        st.caption(f"기간: {_filt_start} ~ {_filt_end} | 최신 주차: {_curr_lbl}  전주차: {_prev_lbl}")

        k1, k2 = st.columns(2)
        with k1:
            st.markdown(f"""<div class="metric-card">
                <div class="label">이용자 수 (일 평균)</div>
                <div class="value">{_kpi_users:,}<span class="unit">명</span>{_delta_html(_kpi_users, _prev_users, "명", ",.0f")}</div>
            </div>""", unsafe_allow_html=True)
        with k2:
            st.markdown(f"""<div class="metric-card-orange">
                <div class="label">1인 평균</div>
                <div class="value">{_kpi_avg:,.0f}<span class="unit">보</span>{_delta_html(_kpi_avg, _prev_avg, "보", ",.0f")}</div>
            </div>""", unsafe_allow_html=True)

        st.markdown("---")

        # ── 지자체 색상 매핑 ─────────────────────────────────────────
        _all_agencies = sorted(_df_filt["agencyName"].unique().tolist())
        _steps_color_map = {m: MUNICIPALITY_COLORS.get(m, "#90CAF9") for m in _all_agencies}

        # ── 일자별 합산 차트 (탭 위) ─────────────────────────────────
        _daily_sum_users = _df_filt.groupby("date")["memberCnt"].sum().reset_index()
        _daily_sum_users["date_str"] = _daily_sum_users["date"].dt.strftime("%m/%d")
        _daily_sum_steps = _df_filt.groupby("date")["totalSteps"].sum().reset_index()
        _daily_sum_steps["date_str"] = _daily_sum_steps["date"].dt.strftime("%m/%d")
        _daily_avg_steps = _daily_sum_steps.copy()
        _daily_avg_steps["avgStep"] = (_daily_sum_steps["totalSteps"] / _daily_sum_users["memberCnt"].replace(0, float("nan"))).fillna(0)

        _sc1, _sc2 = st.columns(2)
        with _sc1:
            _fig_sum_u = go.Figure(go.Bar(
                x=_daily_sum_users["date_str"],
                y=_daily_sum_users["memberCnt"],
                marker_color="#42A5F5",
                text=_daily_sum_users["memberCnt"],
                textposition="outside",
                texttemplate="%{y}명",
                hovertemplate="날짜: <b>%{x}</b><br>이용자: <b>%{y:,}명</b><extra></extra>",
            ))
            _fig_sum_u.update_layout(
                title="일자별 이용자 수 (합산)",
                xaxis=dict(title="", tickangle=-45),
                yaxis=dict(title="이용자 수 (명)"),
                height=380,
                margin=dict(t=50, b=80, r=20),
            )
            st.plotly_chart(_fig_sum_u, use_container_width=True)
        with _sc2:
            _fig_avg_s = go.Figure(go.Bar(
                x=_daily_avg_steps["date_str"],
                y=_daily_avg_steps["avgStep"],
                marker_color="#FFA726",
                text=_daily_avg_steps["avgStep"],
                textposition="outside",
                texttemplate="%{y:,.0f}",
                hovertemplate="날짜: <b>%{x}</b><br>1인 평균: <b>%{y:,.0f}보</b><extra></extra>",
            ))
            _fig_avg_s.update_layout(
                title="일자별 1인 평균 걸음수",
                xaxis=dict(title="", tickangle=-45),
                yaxis=dict(title="걸음수 (보)"),
                height=380,
                margin=dict(t=50, b=80, r=20),
            )
            st.plotly_chart(_fig_avg_s, use_container_width=True)

        st.markdown("---")

        tab1, tab2 = st.tabs(["👥 일자별 이용자 수 (지자체별)", "🦶 일자별 평균 걸음수 (지자체별)"])

        # ── Tab1: 일자별 지자체별 이용자 수 ─────────────────────────
        with tab1:
            # 지자체별 라인 차트
            fig_line_cnt = px.line(
                _df_filt, x="date", y="memberCnt", color="agencyName",
                color_discrete_map=_steps_color_map,
                markers=True,
                title="일자별 지자체별 이용자 추이 (line)",
            )
            fig_line_cnt.update_layout(
                xaxis=dict(title="날짜"),
                yaxis=dict(title="이용자 수 (명)"),
                legend=dict(orientation="h", yanchor="top", y=-0.25, xanchor="center", x=0.5, font=dict(size=10)),
                margin=dict(t=50, b=160, r=20),
                height=480,
            )
            st.plotly_chart(fig_line_cnt, use_container_width=True)

        # ── Tab2: 일자별 지자체별 평균 걸음수 ───────────────────────
        with tab2:
            fig_line_avg = px.line(
                _df_filt, x="date", y="dailyStepAvg", color="agencyName",
                color_discrete_map=_steps_color_map,
                markers=True,
                title="일자별 지자체별 1인 평균 걸음수 (dailyStepAvg)",
            )
            fig_line_avg.update_layout(
                xaxis=dict(title="날짜"),
                yaxis=dict(title="평균 걸음수 (보)"),
                legend=dict(orientation="h", yanchor="top", y=-0.25, xanchor="center", x=0.5, font=dict(size=10)),
                margin=dict(t=50, b=160, r=20),
                height=480,
            )
            fig_line_avg.update_traces(hovertemplate="<b>%{fullData.name}</b><br>날짜: %{x|%m/%d}<br>평균: <b>%{y:,.0f}보</b><extra></extra>")
            st.plotly_chart(fig_line_avg, use_container_width=True)

            # 총 걸음수 라인 (memberCnt × dailyStepAvg)
            fig_line_total = px.line(
                _df_filt, x="date", y="totalSteps", color="agencyName",
                color_discrete_map=_steps_color_map,
                markers=True,
                title="일자별 지자체별 총 걸음수 (memberCnt × dailyStepAvg)",
            )
            fig_line_total.update_layout(
                xaxis=dict(title="날짜"),
                yaxis=dict(title="총 걸음수 (보)"),
                legend=dict(orientation="h", yanchor="top", y=-0.25, xanchor="center", x=0.5, font=dict(size=10)),
                margin=dict(t=50, b=160, r=20),
                height=480,
            )
            fig_line_total.update_traces(hovertemplate="<b>%{fullData.name}</b><br>날짜: %{x|%m/%d}<br>총 걸음수: <b>%{y:,.0f}보</b><extra></extra>")
            st.plotly_chart(fig_line_total, use_container_width=True)
